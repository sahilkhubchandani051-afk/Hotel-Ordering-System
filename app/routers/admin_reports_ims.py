from fastapi import APIRouter, Request, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
# from app.main_fixed import get_db_connection, templates, require_admin (Deferred to avoid circular import)
from app.services.inventory_service import InventoryService
from datetime import datetime
import io
import json
from openpyxl import Workbook
from typing import Optional, List

router = APIRouter(prefix="/admin/reports/ims", tags=["IMS Reports"])

@router.get("/stock-summary", response_class=HTMLResponse)
async def report_stock_summary(
    request: Request,
    search: Optional[str] = None,
    category: Optional[str] = None,
    location: str = "all", # all, shop, godown
    attention_only: Optional[str] = None,
    export: Optional[str] = None
):
    from app.main import get_db_connection, templates, require_admin
    require_admin(request)
    
    # Validation helpers
    def parse_bool(v):
        return str(v).lower() in ('true', '1', 'on', 'yes') if v else False
        
    attention_only = parse_bool(attention_only)
    export = parse_bool(export)
    
    conn = get_db_connection()
    service = InventoryService(conn)
    
    # Base query for products and categories
    query = """
        SELECT 
            p.id, p.name, p.unit, c.name as category, p.reorder_level_shop, p.reorder_level_godown,
            COALESCE(SUM(CASE WHEN l.location_id = 1 THEN l.qty_change ELSE 0 END), 0) as godown_stock,
            COALESCE(SUM(CASE WHEN l.location_id = 2 THEN l.qty_change ELSE 0 END), 0) as shop_stock
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        LEFT JOIN inventory_ledger l ON p.id = l.product_id
    """
    
    where_clauses = []
    params = []
    
    if search:
        where_clauses.append("LOWER(p.name) LIKE ?")
        params.append(f"%{search.lower()}%")
    
    if category:
        where_clauses.append("c.name = ?")
        params.append(category)
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
        
    query += " GROUP BY p.id, p.name, p.unit, c.name, p.reorder_level_shop, p.reorder_level_godown"
    
    # Fetch all to apply computed filters (Status, Location filter)
    rows = conn.execute(query, params).fetchall()
    
    data = []
    for r in rows:
        item = dict(r)
        item['category'] = item['category'] or 'Uncategorized'
        item['total_stock'] = item['godown_stock'] + item['shop_stock']
        
        # Determine status
        # Shop Status
        if item['shop_stock'] == 0: item['shop_status'] = 'OUT'
        elif item['shop_stock'] <= item['reorder_level_shop']: item['shop_status'] = 'LOW'
        else: item['shop_status'] = 'OK'
        
        # Godown Status
        if item['godown_stock'] == 0: item['godown_status'] = 'OUT'
        elif item['godown_stock'] <= item['reorder_level_godown']: item['godown_status'] = 'LOW'
        else: item['godown_status'] = 'OK'

        # Overall Status for "Needs Attention"
        is_attention = item['shop_status'] != 'OK' or item['godown_status'] != 'OK'
        
        # Apply filters in memory for simplicity
        if attention_only and not is_attention:
            continue
        
        # Location filter logic is mostly UI-side but we prepare data
        data.append(item)

    # Categories for filter dropdown
    # Categories for filter dropdown
    categories = [r['name'] for r in conn.execute("SELECT name FROM categories ORDER BY name").fetchall()]
    conn.close()

    if export:
        return export_stock_summary(data)

    return templates.TemplateResponse("admin_reports_ims_stock_summary.html", {
        "request": request,
        "data": data,
        "categories": categories,
        "filters": {
            "search": search,
            "category": category,
            "location": location,
            "attention_only": attention_only
        }
    })

@router.get("/stock-movement", response_class=HTMLResponse)
async def report_stock_movement(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    product_id: Optional[str] = None,
    location_id: Optional[str] = None,
    type: Optional[str] = None,
    export: Optional[str] = None
):
    from app.main import get_db_connection, templates, require_admin
    require_admin(request)
    
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    
    # Int conversion
    if product_id and str(product_id).isdigit():
        product_id = int(product_id)
    else:
        product_id = None
        
    if location_id and str(location_id).isdigit():
        location_id = int(location_id)
    else:
        location_id = None
        
    conn = get_db_connection()
    
    query = """
        SELECT 
            l.*, p.name as product_name, p.unit, loc.name as location_name
        FROM inventory_ledger l
        JOIN products p ON l.product_id = p.id
        JOIN inventory_locations loc ON l.location_id = loc.id
        WHERE 1=1
    """
    params = []
    
    if date_from:
        query += " AND l.created_at >= ?"
        params.append(f"{date_from} 00:00:00")
    if date_to:
        query += " AND l.created_at <= ?"
        params.append(f"{date_to} 23:59:59")
    if product_id:
        query += " AND l.product_id = ?"
        params.append(product_id)
    if location_id:
        query += " AND l.location_id = ?"
        params.append(location_id)
    if type:
        query += " AND l.transaction_type = ?"
        params.append(type)
        
    query += " ORDER BY l.created_at DESC"
    
    rows = conn.execute(query, params).fetchall()
    data = [dict(r) for r in rows]
    
    products = conn.execute("SELECT id, name FROM products ORDER BY name").fetchall()
    locations = conn.execute("SELECT id, name FROM inventory_locations").fetchall()
    conn.close()

    if export:
        return export_stock_movement(data)

    return templates.TemplateResponse("admin_reports_ims_stock_movement.html", {
        "request": request,
        "data": data,
        "products": products,
        "locations": locations,
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "product_id": product_id,
            "location_id": location_id,
            "type": type
        }
    })

@router.get("/sales-vs-stock", response_class=HTMLResponse)
async def report_sales_vs_stock(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    product_id: Optional[str] = None,
    export: Optional[str] = None
):
    from app.main import get_db_connection, templates, require_admin
    require_admin(request)
    
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    
    if product_id and str(product_id).isdigit():
        product_id = int(product_id)
    else:
        product_id = None
        
    conn = get_db_connection()
    
    # 1. Get Products
    p_query = "SELECT id, name, unit FROM products"
    if product_id:
        p_query += f" WHERE id = {product_id}"
    products = conn.execute(p_query).fetchall()
    
    data = []
    for p in products:
        pid = p['id']
        
        # Period Opening Balance (Sum of changes before date_from)
        opening_stock = 0
        if date_from:
            res = conn.execute("SELECT SUM(qty_change) FROM inventory_ledger WHERE product_id = ? AND created_at < ?", (pid, f"{date_from} 00:00:00")).fetchone()
            opening_stock = float(res[0] or 0)
            
        # Sales in period (Direct Sale + Invoice Sales)
        # We look at inventory_ledger with type 'SALE_OUT' or 'SALE'
        sale_query = "SELECT SUM(ABS(qty_change)), transaction_type FROM inventory_ledger WHERE product_id = ? AND transaction_type IN ('SALE', 'SALE_OUT')"
        sale_params = [pid]
        if date_from:
            sale_query += " AND created_at >= ?"
            sale_params.append(f"{date_from} 00:00:00")
        if date_to:
            sale_query += " AND created_at <= ?"
            sale_params.append(f"{date_to} 23:59:59")
        
        sale_query += " GROUP BY transaction_type"
        sale_rows = conn.execute(sale_query, sale_params).fetchall()
        
        total_sales = 0
        direct_sales = 0
        invoice_sales = 0
        
        for sr in sale_rows:
            qty = float(sr[0] or 0)
            total_sales += qty
            if sr['transaction_type'] == 'SALE_OUT': # This is our new type for invoices/direct
                # Note: Currently both are SALE_OUT if finalized via create_invoice. 
                # If we need granularity we'd check ledger.reference_type
                pass
        
        # Alternative: Refined sale granularity
        refined_sale_query = """
            SELECT SUM(ABS(qty_change)) as qty, reference_type 
            FROM inventory_ledger 
            WHERE product_id = ? AND transaction_type IN ('SALE', 'SALE_OUT')
        """
        if date_from: refined_sale_query += f" AND created_at >= '{date_from} 00:00:00'"
        if date_to: refined_sale_query += f" AND created_at <= '{date_to} 23:59:59'"
        refined_sale_query += " GROUP BY reference_type"
        
        refined_rows = conn.execute(refined_sale_query, [pid]).fetchall()
        for rr in refined_rows:
            qty = float(rr['qty'] or 0)
            if rr['reference_type'] == 'INVOICE': invoice_sales += qty
            # Direct sales also use INVOICE ref in Step 9
        
        # Closing Stock (Current total stock)
        closing_res = conn.execute("SELECT SUM(qty_change) FROM inventory_ledger WHERE product_id = ?", (pid,)).fetchone()
        closing_stock = float(closing_res[0] or 0)
        
        # Net Consumed (Sum of all negative qty_change in period)
        cons_query = "SELECT SUM(ABS(qty_change)) FROM inventory_ledger WHERE product_id = ? AND qty_change < 0"
        cons_params = [pid]
        if date_from:
            cons_query += " AND created_at >= ?"
            cons_params.append(f"{date_from} 00:00:00")
        if date_to:
            cons_query += " AND created_at <= ?"
            cons_params.append(f"{date_to} 23:59:59")
        cons_res = conn.execute(cons_query, cons_params).fetchone()
        net_consumed = float(cons_res[0] or 0)

        data.append({
            "name": p['name'],
            "unit": p['unit'],
            "opening_stock": opening_stock,
            "total_sales": total_sales,
            "invoice_sales": invoice_sales, # Approximation for now
            "direct_sales": total_sales - invoice_sales, # Placeholder
            "closing_stock": closing_stock,
            "net_consumed": net_consumed
        })

    product_list = conn.execute("SELECT id, name FROM products ORDER BY name").fetchall()
    conn.close()

    if export:
        return export_sales_vs_stock(data)

    return templates.TemplateResponse("admin_reports_ims_sales_vs_stock.html", {
        "request": request,
        "data": data,
        "products": product_list,
        "filters": {
            "date_from": date_from,
            "date_to": date_to,
            "product_id": product_id
        }
    })

@router.get("/reorder-planning", response_class=HTMLResponse)
async def report_reorder_planning(
    request: Request,
    search: Optional[str] = None,
    category: Optional[str] = None,
    export: Optional[str] = None
):
    from app.main import get_db_connection, templates, require_admin
    require_admin(request)
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    
    conn = get_db_connection()
    
    query = """
        SELECT 
            p.id, p.name, p.unit, p.reorder_level_shop, p.reorder_level_godown,
            c.name as category_name,
            COALESCE(SUM(CASE WHEN l.location_id = 1 THEN l.qty_change ELSE 0 END), 0) as godown_stock,
            COALESCE(SUM(CASE WHEN l.location_id = 2 THEN l.qty_change ELSE 0 END), 0) as shop_stock
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        LEFT JOIN inventory_ledger l ON p.id = l.product_id
    """
    
    where_clauses = []
    params = []
    
    if search:
        where_clauses.append("LOWER(p.name) LIKE ?")
        params.append(f"%{search.lower().strip()}%")
    
    if category:
        where_clauses.append("c.name = ?")
        params.append(category)
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
        
    query += """
        GROUP BY p.id, p.name, p.unit, p.reorder_level_shop, p.reorder_level_godown, c.name
        HAVING godown_stock <= reorder_level_godown OR shop_stock <= reorder_level_shop
        ORDER BY p.name
    """
    
    rows = conn.execute(query, params).fetchall()
    data = []
    for r in rows:
        item = dict(r)
        # Shop Gap
        item['shop_gap'] = max(0, item['reorder_level_shop'] - item['shop_stock'])
        item['shop_status'] = 'LOW' if item['shop_stock'] <= item['reorder_level_shop'] else 'OK'
        
        # Godown Gap
        item['godown_gap'] = max(0, item['reorder_level_godown'] - item['godown_stock'])
        item['godown_status'] = 'LOW' if item['godown_stock'] <= item['reorder_level_godown'] else 'OK'
        
        data.append(item)
    
    # Categories for filter dropdown
    categories = [r['name'] for r in conn.execute("SELECT name FROM categories ORDER BY name").fetchall()]
    conn.close()

    if export:
        return export_reorder_planning(data)

    return templates.TemplateResponse("admin_reports_ims_reorder_planning.html", {
        "request": request,
        "data": data,
        "categories": categories,
        "filters": {
            "search": search,
            "category": category
        }
    })

@router.get("/product-ledger-export/{product_id}", response_class=HTMLResponse)
async def report_product_ledger_export(
    product_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    location_id: Optional[str] = None,
    export: Optional[str] = None
):
    from app.main import get_db_connection, templates, require_admin
    
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    
    if location_id and str(location_id).isdigit():
        location_id = int(location_id)
    else:
        location_id = None
        
    conn = get_db_connection()
    service = InventoryService(conn)
    
    # Get Product info
    p = conn.execute("SELECT name, unit FROM products WHERE id = ?", (product_id,)).fetchone()
    if not p:
        conn.close()
        raise HTTPException(status_code=404, detail="Product not found")
        
    # Get Ledger Data
    entries = service.get_ledger(
        product_id=product_id,
        start_date=date_from,
        end_date=date_to,
        location_id=location_id
    )
    
    conn.close()

    if export:
        return export_product_ledger(p['name'], p['unit'], entries)
    
    # This route is mainly for export, but we could render a preview if needed.
    # For now, if no export flag, we just error out or redirect.
    return {"status": "Export logic only"}

# --- Export Helpers ---

def export_product_ledger(product_name, unit, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Ledger"
    
    ws.append([f"Product: {product_name}", f"Unit: {unit}"])
    ws.append([]) # Spacer
    
    headers = ["Date & Time", "Location", "Type", "Reference", "Qty In", "Qty Out", "Running Balance", "Remark"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row['created_at'], row['location_name'], row['transaction_type'], 
            row['reference_id'], row['qty_in'], row['qty_out'], row['running_balance'],
            row.get('remark', '')
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Ledger_{product_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_stock_summary(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Summary"
    
    headers = ["Product", "Unit", "Shop Stock", "Godown Stock", "Total Stock", "Shop Min", "Godown Min", "Shop Status", "Godown Status"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row['name'], row['unit'], row['shop_stock'], row['godown_stock'], 
            row['total_stock'], row['reorder_level_shop'], row['reorder_level_godown'],
            row['shop_status'], row['godown_status']
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Stock_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_stock_movement(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Movement"
    
    headers = ["Date & Time", "Product", "Location", "Type", "Reference", "Qty In", "Qty Out", "Balance After"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row['created_at'], row['product_name'], row['location_name'], 
            row['transaction_type'], row['reference_id'], row['qty_in'], row['qty_out'], row['running_balance']
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Stock_Movement_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_sales_vs_stock(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales vs Stock"
    
    headers = ["Product", "Unit", "Opening Stock", "Total Sales", "Invoice Sales", "Direct Sales", "Closing Stock", "Net Consumed"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row['name'], row['unit'], row['opening_stock'], row['total_sales'], 
            row['invoice_sales'], row['direct_sales'], row['closing_stock'], row['net_consumed']
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Sales_vs_Stock_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_reorder_planning(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reorder Planning"
    
    headers = ["Product", "Location", "Current Stock", "Min Level", "Gap", "Status"]
    ws.append(headers)
    
    for row in data:
        # Shop Row
        ws.append([row['name'], "SHOP", row['shop_stock'], row['reorder_level_shop'], row['shop_gap'], row['shop_status']])
        # Godown Row
        ws.append([row['name'], "GODOWN", row['godown_stock'], row['reorder_level_godown'], row['godown_gap'], row['godown_status']])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Reorder_Planning_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
