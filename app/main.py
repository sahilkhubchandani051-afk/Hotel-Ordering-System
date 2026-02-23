
import os
import sys
import asyncio
import json
import sqlite3
import pandas as pd
import io
from openpyxl import Workbook
try:
    import psycopg2
    from psycopg2 import extras
except ImportError:
    psycopg2 = None
    extras = None
import logging
import random
import string
import re
from datetime import datetime, timedelta, date
from typing import Optional, List, Dict
from pydantic import BaseModel
from app import utils
from app.utils import number_to_words
try:
    from deep_translator import GoogleTranslator
except ImportError:
    GoogleTranslator = None

import shutil
from fastapi import FastAPI, Request, Form, Response, Depends, HTTPException, status, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from app.routers import admin_inventory, admin_reports_ims
from app.services.inventory_service import InventoryService
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
# SUPABASE FIX APPLIED - mandatory import, no optional fallback
from supabase import create_client, Client as SupabaseClient


# Load environment variables
# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# --- Supabase Initialization ---
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
APP_BASE_URL = os.getenv("APP_BASE_URL", "http://localhost:8000")
VAPID_PUBLIC_KEY = os.getenv("VAPID_PUBLIC_KEY", "BJmqMoqaCjYA4670ufEJVq3tN7uWLjDRjDGOJ47jz_bewcWQQ997YpGr3idWa4O1Myutvr9gnJHMQ1XZNR7J0tk")
VAPID_PRIVATE_KEY = os.getenv("VAPID_PRIVATE_KEY", "4hqYVYWQ0Lek9qo2j3FPE-jQMYdie6vJ_No2b-0F4No")

# SUPABASE FIX APPLIED - mandatory initialization, hard fail if env vars missing
# Supports both SUPABASE_KEY and SUPABASE_SERVICE_ROLE_KEY for backward compatibility
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or SUPABASE_SERVICE_ROLE_KEY

if not SUPABASE_URL or not SUPABASE_KEY:
    raise Exception(
        "Supabase environment variables not configured. "
        "Set SUPABASE_URL and SUPABASE_KEY (or SUPABASE_SERVICE_ROLE_KEY) in your .env file."
    )

supabase: SupabaseClient = create_client(SUPABASE_URL, SUPABASE_KEY)
print("[OK] Supabase client initialized.")


# SUPABASE FIX APPLIED - connection validation functions
def check_supabase_connection():
    """Validates Supabase is reachable. Raises if connection fails."""
    try:
        supabase.table("users").select("*").limit(1).execute()
        print("\u2705 Supabase connected successfully")
    except Exception as e:
        raise Exception(f"Supabase connection failed: {e}")


def check_supabase_tables():
    """Warns if required Supabase tables are missing."""
    required_tables = ["users", "customers", "purchase_requisitions", "invoices"]
    missing = []
    for table in required_tables:
        try:
            supabase.table(table).select("*").limit(1).execute()
        except Exception:
            missing.append(table)
    if missing:
        print(f"\u26a0\ufe0f  WARNING: Supabase tables not found or inaccessible: {missing}. Run your migration SQL.")
    else:
        print("[OK] All required Supabase tables verified.")

# --- Translation Service ---

# --- Notification Helpers ---
def create_notification(conn, email, role, message, link=None):
    """Creates an in-app notification safely."""
    try:
        if not email or not message:
            return
            
        sql = """
            INSERT INTO in_app_notifications (user_email, user_role, message, link)
            VALUES (?, ?, ?, ?)
        """
        is_pg = os.getenv("DATABASE_URL") is not None
        if is_pg:
            sql = sql.replace("?", "%s")
            
        conn.execute(sql, (email, role, message, link))
        # Note: caller must commit
        
        # --- Push Notification ---
        try:
            push_sql = "SELECT endpoint, p256dh, auth FROM push_subscriptions WHERE user_email = ?"
            if is_pg:
                push_sql = push_sql.replace("?", "%s")
                
            subs = conn.execute(push_sql, (email,)).fetchall()
            
            if subs:
                from app.notifications import send_push_notification
                for sub in subs:
                    # Handle both dict-like (sqlite Row) and tuple/object access
                    # SQLite Row can access by name. specific logic for safety
                    endpoint = sub['endpoint'] if isinstance(sub, (dict, sqlite3.Row)) or hasattr(sub, '__getitem__') else sub[0]
                    p256dh = sub['p256dh'] if isinstance(sub, (dict, sqlite3.Row)) or hasattr(sub, '__getitem__') else sub[1]
                    auth = sub['auth'] if isinstance(sub, (dict, sqlite3.Row)) or hasattr(sub, '__getitem__') else sub[2]
                    
                    sub_info = {
                        "endpoint": endpoint,
                        "keys": {
                            "p256dh": p256dh,
                            "auth": auth
                        }
                    }
                    send_push_notification(sub_info, message, link or "/")
        except Exception as push_err:
            print(f"[WARNING] Push notification failed: {push_err}")
            
    except Exception as e:
        print(f"[ERROR] Failed to create notification: {e}")

def get_notifications(conn, email, limit=20):
    """Fetches unread and recent read notifications."""
    try:
        sql = """
            SELECT * FROM in_app_notifications 
            WHERE user_email = ? 
            ORDER BY created_at DESC 
            LIMIT ?
        """
        is_pg = os.getenv("DATABASE_URL") is not None
        if is_pg:
            sql = sql.replace("?", "%s")
            
        return conn.execute(sql, (email, limit)).fetchall()
    except Exception as e:
        print(f"[ERROR] Failed to fetch notifications: {e}")
        return []

def mark_notification_read(conn, notification_id, email):
    """Marks a notification as read safely."""
    try:
        sql = "UPDATE in_app_notifications SET is_read = 1 WHERE id = ? AND user_email = ?"
        is_pg = os.getenv("DATABASE_URL") is not None
        if is_pg:
            sql = sql.replace("?", "%s")
            
        conn.execute(sql, (notification_id, email))
        # Note: caller must commit
    except Exception as e:
        print(f"[ERROR] Failed to mark notification read: {e}")

def validate_image_url(url):
    if not url:
        return True
    try:
        import requests
        response = requests.head(url, timeout=3)
        content_type = response.headers.get('content-type', '')
        return content_type.startswith('image/')
    except:
        return False



# --- Configuration & Path Handling ---

def get_base_path():
    """ Returns the root directory of the project (parent of 'app'). """
    # This file is in app/main.py, so parent of parent of this file is the root
    current_file_path = os.path.abspath(__file__)
    app_dir = os.path.dirname(current_file_path)
    root_dir = os.path.dirname(app_dir)
    return root_dir

def get_db_path():
    """ Returns the path for the database.
        For Render, check if /app/data exists (persistent disk)
    """
    # Check for Render persistent disk
    render_data_path = "/app/data"
    if os.path.exists(render_data_path) and os.path.isdir(render_data_path):
        return render_data_path
    
    # Fallback to project root
    return get_base_path()

ROOT_PATH = get_base_path()
DB_FOLDER = get_db_path()
DB_PATH = os.path.join(DB_FOLDER, "orders.db")
DATABASE_URL = os.getenv("DATABASE_URL")
TEMPLATES_PATH = os.path.join(ROOT_PATH, "templates")
STATIC_PATH = os.path.join(ROOT_PATH, "static")

# Debugging paths and database on startup
print(f"=" * 60)
print(f"DATABASE CONFIGURATION")
print(f"=" * 60)
if DATABASE_URL:
    print(f"[OK] DATABASE_URL is SET")
    print(f"[OK] USING: PostgreSQL (Production)")
    print(f"[OK] URL starts with: {DATABASE_URL[:20]}...")
else:
    # Check if we're on Render (no persistent disk)
    if os.getenv("RENDER"):
        print(f"[WARN] RENDER detected but DATABASE_URL NOT SET")
        print(f"[OK] USING: In-Memory SQLite (NO DATA PERSISTENCE)")
        print(f"[WARN] All data will be lost on restart!")
    else:
        print(f"[INFO] DATABASE_URL is NOT SET")
        print(f"[OK] USING: SQLite (Local Development)")
        print(f"[OK] DB_PATH: {DB_PATH}")
print(f"=" * 60)

# Upload support
UPLOAD_DIR = os.path.join(DB_FOLDER, "static", "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Get configuration from environment variables
SESSION_SECRET = os.getenv("SESSION_SECRET_KEY", "SUPER_SECRET_KEY_CHANGE_ME")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@hotel.com")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")
ENVIRONMENT = os.getenv("ENVIRONMENT", "development")

app = FastAPI(title="Hotel Ordering System")
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)

@app.middleware("http")
async def debug_logging_middleware(request: Request, call_next):
    logging.error(f"DEBUG_MIDDLEWARE: {request.method} {request.url.path}")
    response = await call_next(request)
    return response
        
# Debug code removed

# Mount static files correctly
# Mount persistent uploads directory FIRST to catch /static/uploads/ requests
app.mount("/static/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Mount general static files
if os.path.exists(STATIC_PATH):
    app.mount("/static", StaticFiles(directory=STATIC_PATH), name="static")
else:
    # Fallback for some deployment structures
    os.makedirs(os.path.join(DB_FOLDER, "static"), exist_ok=True)
    app.mount("/static", StaticFiles(directory=os.path.join(DB_FOLDER, "static")), name="static")

# Include Routers
app.include_router(admin_inventory.router)
app.include_router(admin_reports_ims.router)





@app.post("/api/translate")
async def translate_text(request: Request):
    try:
        # Check admin auth
        user = get_current_user(request)
        if not user or user.get("role") != "admin":
             return JSONResponse(status_code=401, content={"error": "Not authorized"})

        data = await request.json()
        text = data.get('text', '')
        target = data.get('target', 'mr') # Default to Marathi
        
        if not text:
            return JSONResponse({'translatedText': ''})
            
        try:
            from deep_translator import GoogleTranslator
            # Use deep-translator
            translator = GoogleTranslator(source='auto', target=target)
            translated_text = translator.translate(text)
            return JSONResponse({'translatedText': translated_text, 'marathi': translated_text})
        except Exception as e:
            print(f"[WARN] Translation failed: {e}")
            # Fallback for common words if lib fails
            fallback_map = {
                'rice': 'तांदूळ',
                'wheat': 'गहू',
                'sugar': 'साखर',
                'oil': 'तेल',
                'milk': 'दूध',
                'water': 'पाणी',
                'salt': 'मीठ'
            }
            return JSONResponse({'translatedText': fallback_map.get(text.lower(), text), 'marathi': fallback_map.get(text.lower(), text)})

    except Exception as e:
        print(f"[ERR] Translation API error: {e}")
        return JSONResponse({'translatedText': '', 'error': str(e)})

# Verify templates directory exists
print(f"=" * 60)
print(f"TEMPLATE CONFIGURATION")
print(f"=" * 60)
print(f"ROOT_PATH: {ROOT_PATH}")
print(f"TEMPLATES_PATH: {TEMPLATES_PATH}")
print(f"Templates directory exists: {os.path.exists(TEMPLATES_PATH)}")
if os.path.exists(TEMPLATES_PATH):
    template_files = os.listdir(TEMPLATES_PATH)
    print(f"Template files found: {len(template_files)}")
    print(f"Sample files: {template_files[:5] if len(template_files) > 0 else 'NONE'}")
else:
    print(f"⚠️ WARNING: Templates directory NOT FOUND!")
print(f"=" * 60)

templates = Jinja2Templates(directory=TEMPLATES_PATH)


# --- Database Setup ---

class PostgreSQLCursorWrapper:
    def __init__(self, cursor, conn):
        self.cursor = cursor
        self.conn = conn
        self._lastrowid = None

    @property
    def lastrowid(self):
        return self._lastrowid

    def execute(self, sql, params=None):
        # Convert SQLite placeholders (?) to PostgreSQL placeholders (%s)
        if "?" in sql:
            sql = sql.replace("?", "%s")
        
        # PostgreSQL-specific case sensitivity for EXCLUDED
        if "ON CONFLICT" in sql.upper() and "excluded." in sql:
            sql = sql.replace("excluded.", "EXCLUDED.")
        
        # Append RETURNING id for INSERTs to simulate lastrowid
        is_insert = sql.strip().upper().startswith("INSERT")
        if is_insert and "RETURNING" not in sql.upper():
            # Avoid RETURNING id for tables that don't have an 'id' column
            no_id_tables = ['customers', 'settings', 'otp_codes']
            sql_upper = sql.upper()
            should_append = True
            for table in no_id_tables:
                if f"INTO {table.upper()}" in sql_upper:
                    should_append = False
                    break
            
            if should_append:
                sql = sql.rstrip().rstrip(';') + " RETURNING id"

        try:
            if params is not None:
                self.cursor.execute(sql, params)
            else:
                self.cursor.execute(sql)
            
            if is_insert:
                try:
                    row = self.cursor.fetchone()
                    if row:
                        self._lastrowid = row[0]
                except:
                    pass
            
            return self
        except Exception as e:
            # Check if this is a PostgreSQL transaction error
            # If the transaction is aborted, we might need a rollback
            # However, we'll let the higher level handle the decision to rollback or just log.
            # But usually, any error in PG means the transaction is aborted.
            print(f"PostgreSQL Cursor Error: {e}")
            print(f"Failed SQL: {sql}")
            
            # Simple heuristic: if we get a transaction error, we should probably ROLLBACK 
            # so subsequent commands DON'T fail with "current transaction is aborted"
            if "current transaction is aborted" in str(e).lower():
                try:
                    self.conn.rollback()
                    print("Transaction rolled back due to aborted state.")
                except:
                    pass
            raise e

    def fetchone(self):
        try:
            return self.cursor.fetchone()
        except Exception as e:
            print(f"PostgreSQL fetchone error: {e}")
            return None

    def fetchall(self):
        try:
            return self.cursor.fetchall()
        except Exception as e:
            print(f"PostgreSQL fetchall error: {e}")
            return []

    def __getattr__(self, name):
        return getattr(self.cursor, name)

class PostgreSQLWrapper:
    def __init__(self, conn):
        self.conn = conn
        
    def execute(self, sql, params=None):
        cur = self.cursor()
        return cur.execute(sql, params)

    def commit(self):
        try:
            self.conn.commit()
        except Exception as e:
            print(f"PostgreSQL commit error: {e}")
            try:
                self.conn.rollback()
            except:
                pass
            raise e

    def rollback(self):
        try:
            self.conn.rollback()
        except:
            pass

    def close(self):
        try:
            self.conn.close()
        except:
            pass

    def cursor(self):
        try:
            # use DictCursor for sqlite.Row-like behavior
            cur = self.conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            return PostgreSQLCursorWrapper(cur, self.conn)
        except Exception as e:
            # If cursor creation fails because transaction is aborted, try a rollback
            if "current transaction is aborted" in str(e).lower():
                try:
                    self.conn.rollback()
                    cur = self.conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                    return PostgreSQLCursorWrapper(cur, self.conn)
                except:
                    pass
            raise e

def get_db_connection():
    db_url = os.getenv("DATABASE_URL")
    if db_url:
        # PostgreSQL (Supabase / Render)
        print(f"[DB] Attempting PostgreSQL connection...")
        try:
            url = db_url
            # Fix legacy postgres:// prefix if present
            if url.startswith("postgres://"):
                url = url.replace("postgres://", "postgresql://", 1)
            
            # Add SSL mode for Supabase/Render
            if "?" in url:
                if "sslmode=" not in url:
                    url += "&sslmode=require"
            else:
                url += "?sslmode=require"
            
            conn = psycopg2.connect(url, connect_timeout=10)
            conn.autocommit = False # Ensure we use transactions
            print(f"[DB] OK PostgreSQL connected successfully")
            return PostgreSQLWrapper(conn)
        except Exception as e:
            print(f"[DB] ERROR CRITICAL ERROR connecting to PostgreSQL: {e}")
            print(f"[DB] DATABASE_URL exists but connection failed!")
            raise e
    else:
        # SQLite (Local or In-Memory for Render)
        try:
            # If on Render without DATABASE_URL, use in-memory database
            if os.getenv("RENDER"):
                print(f"[DB] WARN Using IN-MEMORY SQLite (data will be lost on restart)")
                conn = sqlite3.connect(":memory:", check_same_thread=False)
            else:
                print(f"[DB] Using file-based SQLite (local development)")
                conn = sqlite3.connect(DB_PATH)
            
            conn.row_factory = sqlite3.Row
            return conn
        except Exception as e:
            print(f"[DB] ERROR Error connecting to SQLite: {e}")
            raise e

def init_db():
    conn = get_db_connection()
    c = conn.cursor()
    
    is_pg = os.getenv("DATABASE_URL") is not None
    pk_type = "SERIAL PRIMARY KEY" if is_pg else "INTEGER PRIMARY KEY AUTOINCREMENT"
    supabase_user_id_type = "UUID" if is_pg else "TEXT"

    try:
        # 1. Categories (Independent) - MUST EXIST BEFORE PRODUCTS
        c.execute(f"CREATE TABLE IF NOT EXISTS categories (id {pk_type}, name TEXT UNIQUE NOT NULL)")
        
        # 2. Products (Depends on categories)
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS products (
                id {pk_type},
                name TEXT NOT NULL,
                category_id INTEGER,
                unit TEXT NOT NULL DEFAULT 'pcs',
                rate REAL NOT NULL DEFAULT 0,
                image_path TEXT,
                name_marathi TEXT,
                image_url TEXT,
                created_at TEXT,
                FOREIGN KEY (category_id) REFERENCES categories (id)
            )
        ''')
        # 3. Customers
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS customers (
                email TEXT PRIMARY KEY,
                password TEXT NOT NULL,
                name TEXT NOT NULL,
                phone TEXT,
                full_name TEXT,
                business_name TEXT,
                mobile TEXT,
                address TEXT,
                category TEXT,
                status TEXT DEFAULT 'Active',
                supabase_user_id {supabase_user_id_type},
                created_at TEXT
            )
        ''')
        
        # 4. Customer Categories (Updated for Delivery Routes)
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS customer_categories (
                id {pk_type},
                name TEXT UNIQUE NOT NULL,
                is_active INTEGER DEFAULT 1,
                delivery_days TEXT,
                parent_id INTEGER,
                level INTEGER DEFAULT 1,
                route_name TEXT
            )
        ''')
        
        # 5. Purchase Requisitions
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS purchase_requisitions (
                id {pk_type},
                customer_email TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'PR',
                created_at TEXT NOT NULL,
                items_json TEXT NOT NULL,
                admin_notes TEXT DEFAULT ''
            )
        ''')
        
        # 6. Purchase Orders
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS purchase_orders (
                id {pk_type},
                pr_id INTEGER,
                customer_email TEXT NOT NULL,
                created_at TEXT NOT NULL,
                total_amount REAL NOT NULL,
                items_json TEXT NOT NULL,
                amount_received REAL DEFAULT 0,
                invoice_receipt REAL DEFAULT 0,
                invoice_source TEXT DEFAULT 'PR',
                revision_of_id INTEGER DEFAULT NULL,
                display_id TEXT DEFAULT NULL,
                is_active INTEGER DEFAULT 1,
                revision_reason TEXT DEFAULT NULL,
                status TEXT DEFAULT 'Accepted',
                customer_name_snapshot TEXT,
                business_name_snapshot TEXT,
                address_snapshot TEXT,
                customer_category_snapshot TEXT,
                customer_email_snapshot TEXT,
                customer_mobile_snapshot TEXT,
                expected_delivery_date DATE,
                delivery_stage VARCHAR(50) DEFAULT 'ORDER_PLACED',
                delivery_status VARCHAR DEFAULT 'OPEN',
                tracking_status VARCHAR DEFAULT 'PO_CREATED',
                order_placed_at TIMESTAMP,
                packaged_at TIMESTAMP,
                shipped_at TIMESTAMP,
                out_for_delivery_at TIMESTAMP,
                delivered_at TIMESTAMP
            )
        ''')
            
        # 7. Settings
        c.execute('''
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        ''')
        
        # 8. Invoices
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS invoices (
                id {pk_type},
                po_id INTEGER NOT NULL,
                invoice_no TEXT NOT NULL,
                created_at TEXT NOT NULL,
                payment_mode TEXT NOT NULL,
                delivery_remarks TEXT,
                items_json TEXT NOT NULL,
                status TEXT DEFAULT 'PENDING',
                FOREIGN KEY (po_id) REFERENCES purchase_orders (id)
            )
        ''')

        # 9. Email Notifications Log
        if is_pg:
            c.execute('''
                CREATE TABLE IF NOT EXISTS email_notifications (
                    id SERIAL PRIMARY KEY,
                    event_type VARCHAR NOT NULL,
                    ref_id INTEGER NOT NULL,
                    recipient_role VARCHAR NOT NULL,
                    recipient_email VARCHAR NOT NULL,
                    status VARCHAR DEFAULT 'SENT',
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    error_message TEXT
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS email_notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    event_type TEXT NOT NULL,
                    ref_id INTEGER NOT NULL,
                    recipient_role TEXT NOT NULL,
                    recipient_email TEXT NOT NULL,
                    status TEXT DEFAULT 'SENT',
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    error_message TEXT
                )
            ''')
            
        # 10. In-App Notifications
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS in_app_notifications (
                id {pk_type},
                user_email TEXT NOT NULL,
                user_role TEXT NOT NULL,
                message TEXT NOT NULL,
                link TEXT,
                is_read INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 11. Push Subscriptions
        c.execute(f'''
            CREATE TABLE IF NOT EXISTS push_subscriptions (
                id {pk_type},
                user_email TEXT NOT NULL,
                endpoint TEXT UNIQUE NOT NULL,
                p256dh TEXT NOT NULL,
                auth TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 8. OTP Codes - REMOVED: Using Supabase native email verification
        
        # Initial Seeds
        if is_pg:
            c.execute("INSERT INTO settings (key, value) VALUES ('supplier_name', 'Hotel Supplier Inc.') ON CONFLICT (key) DO NOTHING")
        else:
            c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('supplier_name', 'Hotel Supplier Inc.')")
        conn.commit()
        print("[OK] Database initialization complete.")
    except Exception as e:
        print(f"[ERROR] Database initialization FAILED: {e}")
        try:
            conn.rollback()
        except:
            pass
    finally:
        conn.close()

def calculate_next_delivery_date(po_created_at: datetime, delivery_days: list) -> str:
    """
    Calculate next delivery date based on route delivery days.
    
    Args:
        po_created_at: PO creation timestamp
        delivery_days: List of weekday names (e.g., ["MONDAY", "FRIDAY"])
    
    Returns:
        Next delivery date as string (YYYY-MM-DD)
    
    Logic:
        - Convert delivery_days to weekday numbers (0=Monday, 6=Sunday)
        - Find next occurrence of any delivery day
        - If PO created on delivery day, check time:
          - Before cutoff (e.g., 10 AM) → same day
          - After cutoff → next delivery day
    """
    from datetime import datetime, timedelta
    import calendar
    
    # Convert delivery days to weekday numbers
    delivery_weekdays = []
    for day in delivery_days:
        day_map = {
            'MONDAY': 0, 'TUESDAY': 1, 'WEDNESDAY': 2, 'THURSDAY': 3,
            'FRIDAY': 4, 'SATURDAY': 5, 'SUNDAY': 6
        }
        if day in day_map:
            delivery_weekdays.append(day_map[day])
    
    if not delivery_weekdays:
        # Default to next day if no delivery days configured
        return (po_created_at + timedelta(days=1)).strftime('%Y-%m-%d')
    
    po_date = po_created_at.date()
    current_time = po_created_at.time()
    cutoff_time = datetime.strptime('10:00', '%H:%M').time()
    
    # Check each day from tomorrow onwards
    for i in range(1, 8):  # Check up to 7 days ahead
        check_date = po_date + timedelta(days=i)
        check_weekday = check_date.weekday()
        
        if check_weekday in delivery_weekdays:
            if i == 0 and current_time <= cutoff_time:
                # Same day delivery if before cutoff
                return check_date.strftime('%Y-%m-%d')
            else:
                # Future delivery day
                return check_date.strftime('%Y-%m-%d')
    
    # Fallback to next Monday
    days_until_monday = (7 - po_date.weekday()) % 7 or 7
    return (po_date + timedelta(days=days_until_monday)).strftime('%Y-%m-%d')

def get_current_delivery_stage(po: dict, current_date: date = None) -> tuple:
    """
    Calculate current delivery stage based on dates.
    
    Args:
        po: Purchase order dict with delivery timestamps
        current_date: Today's date (defaults to today)
    
    Returns:
        (stage, timestamp) tuple
    """
    from datetime import datetime, date, timedelta
    
    if current_date is None:
        current_date = date.today()
    
    # Check if invoice exists (delivered)
    if 'delivered_at' in po and po['delivered_at']:
        return ('DELIVERED', po['delivered_at'])
    
    # Get expected delivery date
    expected_date = None
    if 'expected_delivery_date' in po and po['expected_delivery_date']:
        expected_date = datetime.strptime(po['expected_delivery_date'], '%Y-%m-%d').date()
    
    if expected_date and current_date >= expected_date:
        return ('OUT_FOR_DELIVERY', expected_date.strftime('%Y-%m-%d 00:00'))
    elif 'packaged_at' in po and po['packaged_at']:
        return ('PACKED', po['packaged_at'])
    elif expected_date:
        return ('SCHEDULED', expected_date.strftime('%Y-%m-%d'))
    elif 'created_at' in po and po['created_at']:
        po_date = datetime.strptime(po['created_at'], '%Y-%m-%d %H:%M:%S').date()
        if current_date >= po_date + timedelta(days=1):
            return ('PACKED', (po_date + timedelta(days=1)).strftime('%Y-%m-%d 00:00'))
    
    # Default to order placed
    return ('ORDER_PLACED', po.get('created_at', ''))

def update_delivery_timestamps(po_id: int):
    """
    Update all delivery stage timestamps for a PO.
    Called daily via background job or on-demand.
    
    Logic:
        - Get current stage from get_current_delivery_stage()
        - Update corresponding timestamp column
        - Do NOT update if already DELIVERED
    """
    from datetime import datetime, date
    
    conn = get_db_connection()
    try:
        # Get PO details
        po = conn.execute("""
            SELECT expected_delivery_date, delivery_stage, created_at, order_placed_at, packaged_at, shipped_at, out_for_delivery_at
            FROM purchase_orders 
            WHERE id = ?
        """, (po_id,)).fetchone()
        
        if not po:
            return
        
        # Skip if already delivered
        if po.get('delivery_stage') == 'DELIVERED':
            return
        
        # Get current stage
        current_date = date.today()
        stage, timestamp = get_current_delivery_stage(po, current_date)
        
        # Update the appropriate timestamp
        updates = {}
        if stage == 'ORDER_PLACED' and not po.get('order_placed_at'):
            updates['order_placed_at'] = timestamp
        elif stage == 'SCHEDULED' and not po.get('expected_delivery_date'):
            # This usually is set manually or on PO creation, but sync here if needed
            pass
        elif stage == 'PACKED' and not po.get('packaged_at'):
            updates['packaged_at'] = timestamp
        elif stage == 'OUT_FOR_DELIVERY' and not po.get('out_for_delivery_at'):
            updates['out_for_delivery_at'] = timestamp
        
        # Update delivery stage
        updates['delivery_stage'] = stage
        
        if updates:
            set_clause = ", ".join([f"{k} = ?" for k in updates.keys()])
            values = list(updates.values()) + [po_id]
            
            conn.execute(f"""
                UPDATE purchase_orders 
                SET {set_clause}
                WHERE id = ?
            """, values)
            conn.commit()
            
            # NOTIFICATION: Delivery Stage Changed
            # Only notify if stage actually changed (ignore timestamp-only updates for now to reduce noise, 
            # or notify on important timestamp updates like 'Out for Delivery')
            old_stage = po.get('delivery_stage')
            if stage != old_stage:
                 # Get customer email
                 cust_email_row = conn.execute("SELECT customer_email FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
                 if cust_email_row:
                     cust_email = cust_email_row[0]
                     msg = f"Delivery Update: Order #{po_id} is now {stage.replace('_', ' ')}"
                     create_notification(conn, cust_email, "customer", msg, f"/customer/orders")
                     conn.commit()
            
    finally:
        conn.close()

def apply_migrations():
    """ Runs incremental schema updates safely. """
    conn = get_db_connection()
    c = conn.cursor()
    
    is_pg = DATABASE_URL is not None
    
    # Track migrations applied to avoid repeated scans
    if get_setting('migrations_applied') == 'v12':
        conn.close()
        return

    cols_to_add = {
        'customers': {
            'full_name': 'TEXT',
            'business_name': 'TEXT',
            'mobile': 'TEXT',
            'address': 'TEXT',
            'category': 'TEXT',
            'status': "TEXT DEFAULT 'Active'",
            'supabase_user_id': 'UUID',
            'created_at': "TEXT"
        },
        'products': {
            'image_path': 'TEXT',
            'category_id': 'INTEGER',
            'name_marathi': 'TEXT', 
            'image_url': 'TEXT',    
            'created_at': 'TEXT'    
        },
        'purchase_orders': {
            'amount_received': 'REAL DEFAULT 0',
            'invoice_source': "TEXT DEFAULT 'PR'",
            'revision_of_id': 'INTEGER DEFAULT NULL',
            'display_id': 'TEXT DEFAULT NULL',
            'is_active': 'INTEGER DEFAULT 1',
            'status': "TEXT DEFAULT 'Accepted'",
            'delivery_status': "VARCHAR DEFAULT 'OPEN'",
            'tracking_status': "VARCHAR DEFAULT 'PO_CREATED'",
            'customer_name_snapshot': 'TEXT',
            'business_name_snapshot': 'TEXT',
            'address_snapshot': 'TEXT',
            'customer_category_snapshot': 'TEXT',
            'customer_email_snapshot': 'TEXT',
            'customer_mobile_snapshot': 'TEXT'
        }
    }

    try:
        for table, cols in cols_to_add.items():
            for col, dtype in cols.items():
                if is_pg:
                    try:
                        c.execute(f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col} {dtype}")
                    except Exception as e:
                        print(f"Warning: Migrating {table}.{col} failed: {e}")
                else:
                    try:
                        c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {dtype}")
                    except Exception:
                        pass 

        # Ensure supabase_user_id exists on PostgreSQL even if older deployments skipped cols_to_add
        if is_pg:
            try:
                c.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS supabase_user_id UUID")
            except Exception as e:
                print(f"Warning: Ensuring customers.supabase_user_id failed: {e}")
        
        # New table: customer_categories
        if is_pg:
            c.execute('''
                CREATE TABLE IF NOT EXISTS customer_categories (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    is_active INTEGER DEFAULT 1
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS customer_categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    is_active INTEGER DEFAULT 1,
                    delivery_days TEXT,
                    parent_id INTEGER,
                    level INTEGER DEFAULT 1,
                    route_name TEXT
                )
            ''')
            
        # Migration: Ensure customer_categories has new columns (route_name, etc)
        if is_pg:
            try:
                c.execute("ALTER TABLE customer_categories ADD COLUMN IF NOT EXISTS delivery_days TEXT")
                c.execute("ALTER TABLE customer_categories ADD COLUMN IF NOT EXISTS route_name TEXT")
                c.execute("ALTER TABLE customer_categories ADD COLUMN IF NOT EXISTS parent_id INTEGER")
                c.execute("ALTER TABLE customer_categories ADD COLUMN IF NOT EXISTS level INTEGER DEFAULT 1")
            except Exception as e:
                print(f"Warning: Migrating customer_categories columns failed: {e}")
        else:
            try:
                c.execute("ALTER TABLE customer_categories ADD COLUMN delivery_days TEXT")
            except: pass
            try:
                c.execute("ALTER TABLE customer_categories ADD COLUMN route_name TEXT")
            except: pass
            try:
                c.execute("ALTER TABLE customer_categories ADD COLUMN parent_id INTEGER")
            except: pass
            try:
                c.execute("ALTER TABLE customer_categories ADD COLUMN level INTEGER DEFAULT 1")
            except: pass

        # New table: email_notifications
        if is_pg:
            c.execute('''
                CREATE TABLE IF NOT EXISTS email_notifications (
                    id SERIAL PRIMARY KEY,
                    event_type VARCHAR NOT NULL,
                    ref_id INTEGER NOT NULL,
                    recipient_role VARCHAR NOT NULL,
                    recipient_email VARCHAR NOT NULL,
                    status VARCHAR DEFAULT 'SENT',
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    error_message TEXT
                )
            ''')
        else:
            c.execute('''
                CREATE TABLE IF NOT EXISTS email_notifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    event_type TEXT NOT NULL,
                    ref_id INTEGER NOT NULL,
                    recipient_role TEXT NOT NULL,
                    recipient_email TEXT NOT NULL,
                    status TEXT DEFAULT 'SENT',
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    error_message TEXT
                )
            ''')

        # Migration: add email_notifications.error_message (safe, no drop/recreate)
        if is_pg:
            try:
                c.execute("ALTER TABLE email_notifications ADD COLUMN IF NOT EXISTS error_message TEXT")
            except Exception as e:
                print(f"Warning: Migrating email_notifications.error_message failed: {e}")
        else:
            try:
                c.execute("ALTER TABLE email_notifications ADD COLUMN error_message TEXT")
            except Exception:
                pass
        
        # Sync: Set default 'Active' for any customers missing it
        try:
            c.execute("UPDATE customers SET status = 'Active' WHERE status IS NULL")
        except:
            pass

        # Backfill snapshots for existing POs
        try:
            rows = c.execute("""
                SELECT po.id, c.name, c.business_name, c.address, c.category, c.email, c.mobile 
                FROM purchase_orders po
                JOIN customers c ON po.customer_email = c.email
                WHERE po.customer_name_snapshot IS NULL
            """).fetchall()
            for row in rows:
                if is_pg:
                    c.execute("""
                        UPDATE purchase_orders 
                        SET customer_name_snapshot = %s, business_name_snapshot = %s, 
                            address_snapshot = %s, customer_category_snapshot = %s,
                            customer_email_snapshot = %s, customer_mobile_snapshot = %s
                        WHERE id = %s
                    """, (row[1], row[2], row[3], row[4], row[5], row[6], row[0]))
                else:
                    c.execute("""
                        UPDATE purchase_orders 
                        SET customer_name_snapshot = ?, business_name_snapshot = ?, 
                            address_snapshot = ?, customer_category_snapshot = ?,
                            customer_email_snapshot = ?, customer_mobile_snapshot = ?
                        WHERE id = ?
                    """, (row[1], row[2], row[3], row[4], row[5], row[6], row[0]))
        except Exception as e:
            print(f"Warning: Backfilling snapshots failed: {e}")

        # OTP migrations removed - using Supabase native email verification

        # Admin Credentials
        current_email = get_setting('admin_email', ADMIN_EMAIL)
        current_pass = get_setting('admin_password', ADMIN_PASSWORD)
        
        if is_pg:
            c.execute("INSERT INTO settings (key, value) VALUES ('admin_email', %s) ON CONFLICT (key) DO NOTHING", (current_email,))
            c.execute("INSERT INTO settings (key, value) VALUES ('admin_password', %s) ON CONFLICT (key) DO NOTHING", (current_pass,))
        else:
            c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('admin_email', ?)", (current_email,))
            c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('admin_password', ?)", (current_pass,))

        # OTP migrations removed - using Supabase native email verification

        # ── [NEW WORKFLOW ADDITION] ── v12: workflow_type on purchase_requisitions
        # Safe ADD COLUMN migration (ignored if already exists)
        if is_pg:
            try:
                c.execute("ALTER TABLE purchase_requisitions ADD COLUMN IF NOT EXISTS workflow_type TEXT DEFAULT 'QUOTATION_APPROVAL'")
            except Exception as e:
                print(f"Warning: Migrating purchase_requisitions.workflow_type failed: {e}")
        else:
            try:
                c.execute("ALTER TABLE purchase_requisitions ADD COLUMN workflow_type TEXT DEFAULT 'QUOTATION_APPROVAL'")
            except Exception:
                pass  # Column already exists

        conn.commit()
        # Mark migrations as applied
        set_setting('migrations_applied', 'v12')
    except Exception as e:
        print(f"Error during migrations: {e}")
        try:
            conn.rollback()
        except:
            pass
    finally:
        conn.close()





def get_setting(key: str, default: str = "") -> str:
    try:
        conn = get_db_connection()
        res = conn.execute("SELECT value FROM settings WHERE key = ?", (key,))
        row = res.fetchone()
        conn.close()
        return row['value'] if row else default
    except:
        return default

def set_setting(key: str, value: str):
    try:
        conn = get_db_connection()
        if DATABASE_URL:
            conn.execute("INSERT INTO settings (key, value) VALUES (%s, %s) ON CONFLICT(key) DO UPDATE SET value = EXCLUDED.value", (key, value))
        else:
            conn.execute("INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value", (key, value))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error setting {key}: {e}")

# OTP functionality removed - using Supabase native email verification

# Initialize DB on startup
# ...
init_db()
apply_migrations()


# --- Dependencies & Helpers ---

def get_current_user(request: Request):
    user = request.session.get("user")
    if user:
        # Debug print to console (stderr)
        import sys
        print(f"DEBUG: Session User Type: {type(user)}", file=sys.stderr)
        
        if isinstance(user, dict):
            return user
            
        # Attempt conversion from sqlite3.Row or other mappings
        try:
            return dict(user)
        except Exception as e:
            print(f"DEBUG: Failed to convert user to dict: {e}", file=sys.stderr)
            
    return user

def require_admin(request: Request):
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=303, detail="Not authorized")
    return user

def require_customer(request: Request):
    user = get_current_user(request)
    # Handle Row objects or Dicts safely
    try:
        role = user["role"]
    except (KeyError, TypeError, IndexError):
        role = None
        
    if not user or role != "customer":
        raise HTTPException(status_code=303, detail="Not authorized")
    return user

# Translation helper
def translate_to_marathi(text: str) -> str:
    try:
        from deep_translator import GoogleTranslator
        # Use deep_translator which is more stable than googletrans
        return GoogleTranslator(source='auto', target='mr').translate(text)
    except Exception as e:
        print(f"Translation error: {e}")
        # Simple fallback mapping for common words
        fallback_map = {
            'rice': 'तांद',
            'wheat': 'गहू',
            'sugar': 'साखर',
            'oil': 'तेल',
            'milk': 'दूध',
            'water': 'पाणी',
            'salt': 'मीठ'
        }
        return fallback_map.get(text.lower(), "")

# Basic URL validation
def validate_image_url(url: str) -> bool:
    # Regex for a simple URL validation
    regex = re.compile(
        r'^(?:http|ftp)s?://' # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' # domain...
        r'localhost|' # localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
        r'(?::\d+)?' # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)
    return re.match(regex, url) is not None

def is_quotation_expired(created_at_str: str) -> bool:
    try:
        created_at = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
        return datetime.now() > (created_at + timedelta(days=3))
    except (ValueError, TypeError):
        return False

# --- API: Duplicate Checks ---

@app.get("/api/check-product-duplicate")
async def check_product_duplicate(name: str, unit: str):
    conn = get_db_connection()
    # Normalize by trimming and lowercase
    res = conn.execute("SELECT id FROM products WHERE LOWER(TRIM(name)) = LOWER(TRIM(?)) AND unit = ?", (name, unit)).fetchone()
    conn.close()
    return {"exists": res is not None}

@app.get("/api/check-category-duplicate")
async def check_category_duplicate(name: str):
    conn = get_db_connection()
    res = conn.execute("SELECT id FROM categories WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))", (name,)).fetchone()
    conn.close()
    return {"exists": res is not None}

@app.get("/api/check-customer-duplicate")
async def check_customer_duplicate(email: str):
    conn = get_db_connection()
    res = conn.execute("SELECT email FROM customers WHERE email = ?", (email.strip(),)).fetchone()
    conn.close()
    return {"exists": res is not None}

@app.post("/api/check-invoice-duplicate")
async def check_invoice_duplicate(request: Request):
    try:
        data = await request.json()
        customer_email = data.get("customer_email")
        items_list = data.get("items") # [{product_id, qty, rate}]
        total_amount = data.get("total_amount")
        
        # Date check (today)
        today = datetime.now().strftime("%Y-%m-%d")
        
        conn = get_db_connection()
        # Find all invoices for this customer today with same total
        orders = conn.execute("""
            SELECT items_json FROM purchase_orders 
            WHERE customer_email = ? 
            AND created_at LIKE ? 
            AND ABS(total_amount - ?) < 0.01
        """, (customer_email, f"{today}%", total_amount)).fetchall()
        
        exists = False
        import json
        for order in orders:
            try:
                order_items = json.loads(order['items_json'])
                # Compare products and quantities
                if len(order_items) != len(items_list):
                    continue
                
                # Sort both to compare easily
                canonical_existing = sorted([(int(itm['product_id']), float(itm['qty'])) for itm in order_items])
                canonical_new = sorted([(int(itm['product_id']), float(itm['qty'])) for itm in items_list])
                
                if canonical_existing == canonical_new:
                    exists = True
                    break
            except:
                continue
        
        conn.close()
        return {"exists": exists}
    except Exception as e:
        print(f"Error checking invoice duplicate: {e}")
        return {"exists": False}

# ============================================================================
# DELIVERY TRACKING FUNCTIONS
# ============================================================================

def calculate_next_delivery_date(po_created_at, delivery_days):
    """
    Calculate next delivery date based on route delivery days.
    
    Args:
        po_created_at: PO creation datetime
        delivery_days: List of weekday names (e.g., ["MONDAY", "FRIDAY"])
    
    Returns:
        Next delivery date (date object)
    """
    if not delivery_days:
        # Default: 7 days from PO creation
        return (po_created_at + timedelta(days=7)).date()
    
    # Convert weekday names to numbers (0=Monday, 6=Sunday)
    weekday_map = {
        "MONDAY": 0, "TUESDAY": 1, "WEDNESDAY": 2, "THURSDAY": 3,
        "FRIDAY": 4, "SATURDAY": 5, "SUNDAY": 6
    }
    
    delivery_weekdays = []
    for day in delivery_days:
        if day.upper() in weekday_map:
            delivery_weekdays.append(weekday_map[day.upper()])
    
    if not delivery_weekdays:
        return (po_created_at + timedelta(days=7)).date()
    
    # Find next delivery day
    current_date = po_created_at.date()
    current_weekday = po_created_at.weekday()
    
    # Check if today is a delivery day and it's before cutoff time (10 AM)
    if current_weekday in delivery_weekdays and po_created_at.hour < 10:
        return current_date
    
    # Find next delivery day
    for i in range(1, 8):
        next_date = current_date + timedelta(days=i)
        if next_date.weekday() in delivery_weekdays:
            return next_date
    
    # Fallback (should never reach here)
    return current_date + timedelta(days=7)


def get_current_delivery_stage(po, current_datetime=None):
    """
    Calculate current delivery stage based on dates.
    
    Args:
        po: Purchase order dict with expected_delivery_date
        current_datetime: Current datetime (defaults to now)
    
    Returns:
        (stage, timestamp) tuple
    """
    if current_datetime is None:
        current_datetime = datetime.now()
    
    current_date = current_datetime.date()
    
    # If invoice exists (delivered_at is set), always return DELIVERED
    if po.get('delivered_at'):
        return ('DELIVERED', po['delivered_at'])
    
    # Get expected delivery date
    expected_delivery_date = po.get('expected_delivery_date')
    if not expected_delivery_date:
        return ('ORDER_PLACED', po.get('order_placed_at') or po.get('created_at'))
    
    # Convert to date if string
    if isinstance(expected_delivery_date, str):
        expected_delivery_date = datetime.strptime(expected_delivery_date.split()[0], '%Y-%m-%d').date()
    
    D = expected_delivery_date
    
    # Stage logic based on expected delivery date
    if current_date >= D:
        # On or after delivery day → OUT_FOR_DELIVERY
        timestamp = datetime.combine(D, datetime.min.time())
        return ('OUT_FOR_DELIVERY', timestamp)
    elif current_date >= D - timedelta(days=1):
        # Day before delivery → SHIPPED
        timestamp = datetime.combine(D - timedelta(days=1), datetime.min.time())
        return ('SHIPPED', timestamp)
    elif current_date >= (po.get('order_placed_at') or po.get('created_at')).date() + timedelta(days=1):
        # At least 1 day after order → PACKAGED
        order_date = (po.get('order_placed_at') or po.get('created_at'))
        if isinstance(order_date, str):
            order_date = datetime.strptime(order_date, '%Y-%m-%d %H:%M:%S')
        timestamp = order_date + timedelta(days=1)
        return ('PACKAGED', timestamp)
    else:
        # Just placed → ORDER_PLACED
        return ('ORDER_PLACED', po.get('order_placed_at') or po.get('created_at'))


def update_delivery_timestamps(po_id):
    """
    Update delivery stage and timestamps for a PO.
    Called daily or on-demand.
    
    Args:
        po_id: Purchase order ID
    """
    conn = get_db_connection()
    
    try:
        po = conn.execute("""
            SELECT id, expected_delivery_date, delivery_stage, created_at,
                   order_placed_at, delivered_at
            FROM purchase_orders
            WHERE id = ?
        """, (po_id,)).fetchone()
        
        if not po:
            return
        
        # Don't update if already delivered
        if po['delivery_stage'] == 'DELIVERED':
            return
        
        # Get current stage
        stage, timestamp = get_current_delivery_stage(dict(po))
        
        # Update database
        update_fields = {
            'delivery_stage': stage
        }
        
        # Set appropriate timestamp
        if stage == 'PACKAGED' and not po.get('packaged_at'):
            update_fields['packaged_at'] = timestamp
        elif stage == 'SHIPPED' and not po.get('shipped_at'):
            update_fields['shipped_at'] = timestamp
        elif stage == 'OUT_FOR_DELIVERY' and not po.get('out_for_delivery_at'):
            update_fields['out_for_delivery_at'] = timestamp
        
        # Build UPDATE query
        set_clause = ', '.join([f"{k} = ?" for k in update_fields.keys()])
        values = list(update_fields.values()) + [po_id]
        
        conn.execute(f"""
            UPDATE purchase_orders
            SET {set_clause}
            WHERE id = ?
        """, values)
        
        conn.commit()
    finally:
        conn.close()


def update_all_delivery_stages():
    """Update delivery stages for all active POs (called by background job)"""
    conn = get_db_connection()
    
    try:
        active_pos = conn.execute("""
            SELECT id
            FROM purchase_orders
            WHERE delivery_stage != 'DELIVERED' AND is_active = 1
        """).fetchall()
        
        for po in active_pos:
            update_delivery_timestamps(po['id'])
    finally:
        conn.close()

# --- Routes: Auth ---
# Note: /login route is below


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    user = get_current_user(request)
    if user:
        if user["role"] == "admin":
            return RedirectResponse(url="/admin/dashboard", status_code=303)
        return RedirectResponse(url="/customer/dashboard", status_code=303)
    
    supplier_name = get_setting("supplier_name", "Hotel Supplier Inc.")
    return templates.TemplateResponse("landing.html", {"request": request, "supplier_name": supplier_name})

@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    user = get_current_user(request)
    if user:
         if user["role"] == "admin":
             return RedirectResponse(url="/admin/dashboard", status_code=303)
         return RedirectResponse(url="/customer/dashboard", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request})

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    user = get_current_user(request)
    if user:
         if user["role"] == "admin":
             return RedirectResponse(url="/admin/dashboard", status_code=303)
         return RedirectResponse(url="/customer/dashboard", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
async def login(request: Request, email: str = Form(...), password: str = Form(...)):
    email = email.lower().strip()
    # 1. Admin Check - use settings first, then env
    db_admin_email = get_setting("admin_email", ADMIN_EMAIL)
    db_admin_password = get_setting("admin_password", ADMIN_PASSWORD)
    
    if email == db_admin_email and password == db_admin_password:
        request.session["user"] = {"email": email, "role": "admin", "name": "Administrator"}
        return RedirectResponse(url="/admin/dashboard", status_code=303)
    
    # 2. Customer Check — Supabase Auth (mandatory)
    # SUPABASE FIX APPLIED - removed `if supabase:` guard and local DB fallback
    # Supabase is always available (enforced at startup); no plaintext fallback.
    try:
        res = supabase.auth.sign_in_with_password({"email": email, "password": password})
        
        # Check Email Verification
        if not res.user.email_confirmed_at:
            await supabase.auth.sign_out()
            return templates.TemplateResponse("login.html", {
                "request": request, 
                "error": "Please verify your email address to continue."
            })

        # Login Successful - Fetch Local Profile
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM customers WHERE email = ?", (email,)).fetchone()
        conn.close()
        
        if user:
            request.session["user"] = {"email": user["email"], "role": "customer", "name": user["name"]}
            return RedirectResponse(url="/customer/dashboard", status_code=303)
        else:
            return templates.TemplateResponse("login.html", {"request": request, "error": "Account not found. Please contact support."})

    except Exception as e:
        # Supabase login failed (wrong credentials, email not verified, etc.)
        print(f"[AUTH] Supabase login failed for {email}: {e}")
        return templates.TemplateResponse("login.html", {
            "request": request, 
            "error": "Invalid email or password"
        })

@app.get("/register")
async def register_page(request: Request):
    return RedirectResponse(url="/register-otp", status_code=303)

@app.get("/forgot-password")
async def forgot_password_page(request: Request):
    return RedirectResponse(url="/forgot-password-otp", status_code=303)

@app.post("/forgot-password")
async def forgot_password(request: Request, email: str = Form(...)):
    email = email.lower().strip()
    # 1. Check if Admin
    admin_email = get_setting("admin_email", ADMIN_EMAIL)
    
    if email == admin_email:
        # Admin reset flow still needs consideration if we remove OTP
        # But for now, let's focus on Customer flow as primary request
        # If admin needs reset, they might need manual DB access or use Supabase if we migrate admin too?
        # User said "REMOVE custom OTP".
        # I'll output a message for admin: "Please contact support" or similar if critical.
        # Or frankly, just let Supabase handle it if we migrated admin.
        # But admin is not in Supabase usually.
        # I will leave Admin flow as "Not Supported via Email" for now or just generic error 
        # to focus on Customer Supabase flow. 
        # User said "REMOVE custom OTP". 
        pass

    # 2. Supabase Reset Password (Customer)
    if supabase:
        try:
             # Supabase handles the email sending
             # redirect_to should point to a page that handles the hash fragment
             # or just the login page if it's a "magic link" login.
             supabase.auth.reset_password_email(email)
             
             return templates.TemplateResponse("forgot_password.html", {
                "request": request, 
                "message": "If an account exists, a password reset link has been sent to your email."
            })
        except Exception as e:
            print(f"Supabase Reset Error: {e}")
            # Don't reveal error details
            return templates.TemplateResponse("forgot_password.html", {
                "request": request, 
                "message": "If an account exists, a password reset link has been sent to your email."
            })
            
    return templates.TemplateResponse("forgot_password.html", {
        "request": request, 
        "error": "Password reset service unavailable."
    })

# Password reset handled by Supabase native email flow

# @app.post("/register")
# Route removed to disable legacy registration and prevent duplicates.
# Use /register-otp instead.

# REMOVED: /verify/register, /verify/register/resend (Supabase Native Auth)

@app.get("/logout")
async def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/", status_code=303)

# --- Routes: Admin ---

@app.get("/admin", include_in_schema=False)
async def admin_root(request: Request):
    return RedirectResponse(url="/admin/dashboard", status_code=303)

@app.get("/admin/dashboard", response_class=HTMLResponse)
async def admin_dashboard(request: Request):
    require_admin(request)
    supplier_name = get_setting("supplier_name", "Hotel Supplier Inc.")
    
    # Analytics
    conn = get_db_connection()
    stats = {
        "products": conn.execute("SELECT COUNT(*) FROM products").fetchone()[0],
        "total_prs": conn.execute("SELECT COUNT(*) FROM purchase_requisitions").fetchone()[0],
        "pending": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE status='PR'").fetchone()[0],
        "quoted": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE status='QT'").fetchone()[0],
        "accepted": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE status='ACCEPTED'").fetchone()[0],
        "rejected": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE status='REJECTED'").fetchone()[0],
        "partially_accepted": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE status='PARTIALLY_ACCEPTED'").fetchone()[0],
        # [NEW WORKFLOW ADDITION] Count PRs awaiting admin price review
        "pending_price_review": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE status='PENDING_PRICE_REVIEW'").fetchone()[0],
    }
    conn.close()
    
    return templates.TemplateResponse("admin_dashboard.html", {
        "request": request, 
        "supplier_name": supplier_name,
        "stats": stats
    })




@app.get("/admin/profile", response_class=HTMLResponse)
async def admin_profile(request: Request):
    require_admin(request)
    supplier_name = get_setting("supplier_name", "Hotel Supplier Inc.")
    admin_email = get_setting("admin_email", ADMIN_EMAIL)
    
    # Feature 2: Email Config
    otp_sender_email = get_setting("otp_sender_email", "")
    otp_password = get_setting("otp_password", "")
    otp_is_configured = bool(otp_password)
    
    return templates.TemplateResponse("admin_profile.html", {
        "request": request,     
        "supplier_name": supplier_name,
        "admin_email": admin_email,
        "otp_sender_email": otp_sender_email,
        "otp_is_configured": otp_is_configured
    })

@app.post("/admin/profile")
async def update_admin_profile(
    request: Request, 
    supplier_name: str = Form(...),
    admin_email: str = Form(...),
    current_password: str = Form(...),
    new_password: str = Form(None),
    otp_sender_email: str = Form(None),
    otp_password: str = Form(None),
    background_tasks: BackgroundTasks = None # Optional for backward compatibility but effectively required
):
    require_admin(request)
    
    # 1. Verify Current Password
    db_pass = get_setting("admin_password", ADMIN_PASSWORD)
    if current_password != db_pass:
        # Re-fetch settings to render page correctly
        otp_sender_email_db = get_setting("otp_sender_email", "")
        otp_is_configured = bool(get_setting("otp_password", ""))
        
        return templates.TemplateResponse("admin_profile.html", {
            "request": request, 
            "error": "Incorrect current password.",
            "supplier_name": supplier_name,
            "admin_email": admin_email,
            "otp_sender_email": otp_sender_email_db,
            "otp_is_configured": otp_is_configured
        })
        
    # 2. Update Basic Info (Non-Critical)
    set_setting("supplier_name", supplier_name)
    
    # Update OTP settings (Non-Critical, already protected by password check)
    if otp_sender_email:
        set_setting("otp_sender_email", otp_sender_email)
    
    if otp_password and otp_password.strip():
        set_setting("otp_password", otp_password)
        
    # 3. Check for Critical Changes (Email or Password)
    db_admin_email = get_setting("admin_email", ADMIN_EMAIL)
    critical_change = False
    
    if admin_email != db_admin_email:
        critical_change = True
        
    if new_password and new_password.strip():
        # Validate rules immediately
        errors = []
        if len(new_password) < 8:
            errors.append("Password < 8 chars.")
        if not re.search(r"[0-9]", new_password):
            errors.append("No number in password.")
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", new_password):
            errors.append("No special char in password.")
            
        if errors:
             # Re-fetch settings
             otp_sender_email_db = get_setting("otp_sender_email", "")
             otp_is_configured = bool(get_setting("otp_password", ""))
             
             return templates.TemplateResponse("admin_profile.html", {
                "request": request, 
                "error": "Invalid New Password: " + " ".join(errors),
                "supplier_name": supplier_name,
                "admin_email": admin_email,
                "otp_sender_email": otp_sender_email_db,
                "otp_is_configured": otp_is_configured
            })
        critical_change = True

    if critical_change:
        # Generate OTP
        otp = utils.generate_otp()
        current_email = db_admin_email # Send to CURRENT email
        # Store OTP
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        expires_at = (datetime.now() + timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")
        
        # Run cleanup
        cleanup_expired_otps()
        
        conn = get_db_connection()
        pending_data = json.dumps({
            "admin_email": admin_email,
            "new_password": new_password
        })
        
        if DATABASE_URL:
            conn.execute("""
                INSERT INTO otp_codes (email, otp, purpose, data, created_at, expires_at) 
                VALUES (%s, %s, %s, %s, %s, %s) 
                ON CONFLICT(email, purpose) DO UPDATE SET 
                otp=EXCLUDED.otp, data=EXCLUDED.data, created_at=EXCLUDED.created_at, expires_at=EXCLUDED.expires_at
            """, (current_email, otp, 'admin_update', pending_data, created_at, expires_at))
        else:
            conn.execute("""
                INSERT INTO otp_codes (email, otp, purpose, data, created_at, expires_at) 
                VALUES (?, ?, ?, ?, ?, ?) 
                ON CONFLICT(email, purpose) DO UPDATE SET 
                otp=excluded.otp, data=excluded.data, created_at=excluded.created_at, expires_at=excluded.expires_at
            """, (current_email, otp, 'admin_update', pending_data, created_at, expires_at))
        conn.commit()
        conn.close()
        
        # Send Email
        sender = get_setting("otp_sender_email", "")
        sender_pass = get_setting("otp_password", "")
        background_tasks.add_task(utils.send_email_otp, current_email, otp, sender, sender_pass)
            
        return RedirectResponse(url="/verify/admin-update", status_code=303)
        
    # If no critical changes
    return templates.TemplateResponse("admin_profile.html", {
        "request": request, 
        "success": "Profile updated successfully.",
        "supplier_name": supplier_name,
        "admin_email": admin_email,
        "otp_sender_email": get_setting("otp_sender_email", ""),
        "otp_is_configured": bool(get_setting("otp_password", ""))
    })

@app.get("/verify/admin-update", response_class=HTMLResponse)
async def verify_admin_update_page(request: Request):
    require_admin(request)
    # OTP sent to CURRENT email
    current_email = get_setting("admin_email", ADMIN_EMAIL)
    
    conn = get_db_connection()
    record = conn.execute("SELECT * FROM otp_codes WHERE email = ? AND purpose = ?", (current_email, 'admin_update')).fetchone()
    conn.close()
    
    if not record:
         return RedirectResponse(url="/admin/profile", status_code=303)
         
    masked = utils.mask_email(current_email)
    
    return templates.TemplateResponse("auth_verify_otp.html", {
        "request": request, 
        "email": masked,
        "target_url": "/verify/admin-update"
    })

@app.post("/verify/admin-update")
async def verify_admin_update(request: Request, otp: str = Form(...)):
    user = require_admin(request)
    current_email = get_setting("admin_email", ADMIN_EMAIL)
    
    # Verify OTP and get pending data from DB
    conn = get_db_connection()
    record = conn.execute("SELECT * FROM otp_codes WHERE email = ? AND purpose = ?", (current_email, 'admin_update')).fetchone()
    
    valid = False
    if record:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if record["otp"] == otp:
            if now < record["expires_at"]:
                valid = True
                
    if not valid or not record or not record["data"]:
        conn.close()
        return templates.TemplateResponse("auth_verify_otp.html", {
            "request": request, 
            "email": utils.mask_email(current_email),
            "target_url": "/verify/admin-update",
            "error": "Invalid or Expired OTP"
        })
        
    # Apply Critical Changes
    try:
        pending = json.loads(record["data"])
        if pending["admin_email"]:
            set_setting("admin_email", pending["admin_email"])
            user["email"] = pending["admin_email"]
            request.session["user"] = user
            
        if pending["new_password"] and pending["new_password"].strip():
            set_setting("admin_password", pending["new_password"])
        
        # Cleanup
        conn.execute("DELETE FROM otp_codes WHERE email = ? AND purpose = ?", (current_email, 'admin_update'))
        conn.commit()
        
    except Exception as e:
        print(f"Error applying admin update: {e}")
        
    conn.close()
    
    return RedirectResponse(url="/admin/profile?success=Credentials Updated", status_code=303)

# Kept for backward compatibility if any old links use it, but logic moved to Profile
@app.post("/admin/settings")
async def update_settings(request: Request, supplier_name: str = Form(...)):
    require_admin(request)
    set_setting("supplier_name", supplier_name)
    return RedirectResponse(url="/admin/dashboard", status_code=303)

@app.post("/admin/categories/add")
async def add_category(request: Request, category_name: str = Form(...)):
    require_admin(request)
    name = category_name.strip()
    if not name:
        # Just redirect back if empty
        return RedirectResponse(url="/admin/products", status_code=303)
        
    conn = get_db_connection()
    try:
        conn.execute("INSERT INTO categories (name) VALUES (?)", (name,))
        conn.commit()
    except (sqlite3.IntegrityError, psycopg2.Error):
        pass # Already exists
    conn.close()
    return RedirectResponse(url="/admin/products", status_code=303)

@app.get("/admin/products", response_class=HTMLResponse)
async def admin_products(request: Request):
    require_admin(request)
    conn = get_db_connection()
    # Left join categories to get name even if category_id is NULL or not found
    query = '''
        SELECT p.*, c.name as category_name 
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        ORDER BY p.id DESC
    '''
    products = conn.execute(query).fetchall()
    categories = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    conn.close()
    return templates.TemplateResponse("admin_products.html", {
        "request": request, 
        "products": products,
        "categories": categories
    })

@app.get("/admin/products/form", response_class=HTMLResponse)
async def admin_products_form(request: Request, id: int = None):
    require_admin(request)
    conn = get_db_connection()
    categories = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    
    product = None
    if id:
        product = conn.execute("SELECT * FROM products WHERE id = ?", (id,)).fetchone()
        
    conn.close()
    
    return templates.TemplateResponse("admin_product_form.html", {
        "request": request,
        "product": product,
        "categories": categories
    })

@app.post("/admin/products/add")
async def add_product(
    request: Request,
    name: str = Form(...),
    name_marathi: str = Form(""),
    unit: str = Form(...),
    rate: float = Form(...),
    category_id: str = Form(None),
    image: UploadFile = File(None),
    image_url: str = Form("")
):
    require_admin(request)
    
    # Handle category_id input (empty string -> None)
    cat_id = None
    if category_id and str(category_id).strip():
        cat_id = int(category_id)
    
    # Auto-translate if empty
    if not name_marathi and name:
        name_marathi = translate_to_marathi(name)
        
    # Validate image URL
    if image_url and not validate_image_url(image_url):
         pass

    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Handle optional image
    filename = None
    if image and image.filename:
        # Save to uploads
        original_name = image.filename
        safe_name = f"{int(datetime.now().timestamp())}_{original_name}"
        # Ensure upload dir
        import os
        upload_dir = "static/uploads"
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            
        file_location = os.path.join(upload_dir, safe_name)
        
        with open(file_location, "wb+") as file_object:
            shutil.copyfileobj(image.file, file_object)
            
        filename = f"uploads/{safe_name}"
    
    conn = get_db_connection()
    conn.execute("INSERT INTO products (name, name_marathi, unit, rate, category_id, image_path, image_url, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                 (name, name_marathi, unit, rate, cat_id, filename, image_url, created_at))
    conn.commit()
    conn.close()
    
    return RedirectResponse(url="/admin/products", status_code=303)

@app.post("/admin/products/delete")
async def delete_product(request: Request, product_id: int = Form(...)):
    require_admin(request)
    conn = get_db_connection()
    conn.execute("DELETE FROM products WHERE id = ?", (product_id,))
    conn.commit()
    conn.close()

# --- Customer Categories Management ---

@app.get("/admin/customer_categories", response_class=HTMLResponse)
async def admin_customer_categories(request: Request):
    require_admin(request)
    conn = get_db_connection()
    try:
        categories = conn.execute("SELECT * FROM customer_categories ORDER BY name").fetchall()
    except Exception:
        # Table might not exist yet
        categories = []
    conn.close()
    
    return templates.TemplateResponse("admin_customer_categories.html", {
        "request": request, 
        "categories": categories
    })

@app.post("/admin/customer_categories/add")
async def add_customer_category(request: Request, name: str = Form(...), parent_id: str = Form(None)):
    require_admin(request)
    conn = get_db_connection()
    try:
        form = await request.form()
        
        # 1. Parse Delivery Days
        days = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
        selected_days = []
        for day in days:
            if form.get(f"day_{day}"):
                selected_days.append(day)
        
        delivery_days_json = json.dumps(selected_days)
        
        # 2. Parse Parent ID
        pid = None
        if parent_id and str(parent_id).strip():
             pid = int(parent_id)

        name = name.strip()
        if name:
             # Postgres vs Sqlite syntax handled by driver usually for simple insert
             if DATABASE_URL:
                 conn.execute("INSERT INTO customer_categories (name, parent_id, delivery_days) VALUES (%s, %s, %s)", (name, pid, delivery_days_json))
             else:
                 conn.execute("INSERT INTO customer_categories (name, parent_id, delivery_days) VALUES (?, ?, ?)", (name, pid, delivery_days_json))
             conn.commit()
    except Exception as e:
        print(f"Error adding customer category: {e}")
    conn.close()
        
    return RedirectResponse(url="/admin/customer_categories", status_code=303)

@app.post("/admin/customer_categories/update")
async def update_customer_category(request: Request, category_name: str = Form(...)):
    require_admin(request)
    conn = get_db_connection()
    try:
        form = await request.form()
        
        # 1. Parse Delivery Days
        days = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
        selected_days = []
        for day in days:
            if form.get(f"day_{day}"):
                selected_days.append(day)
        
        delivery_days_json = json.dumps(selected_days)
        name = category_name.strip()
        
        if name:
             if DATABASE_URL:
                 conn.execute("UPDATE customer_categories SET delivery_days = %s WHERE name = %s", (delivery_days_json, name))
             else:
                 conn.execute("UPDATE customer_categories SET delivery_days = ? WHERE name = ?", (delivery_days_json, name))
             conn.commit()
    except Exception as e:
        print(f"Error updating customer category: {e}")
    conn.close()
    return RedirectResponse(url="/admin/customer_categories", status_code=303)

@app.post("/admin/customer_categories/delete")
async def delete_customer_category(request: Request, category_name: str = Form(...)):
    require_admin(request)
    conn = get_db_connection()
    try:
        if DATABASE_URL:
             conn.execute("DELETE FROM customer_categories WHERE name = %s", (category_name,))
        else:
             conn.execute("DELETE FROM customer_categories WHERE name = ?", (category_name,))
        conn.commit()
    except Exception as e:
        print(f"Error deleting customer category: {e}")
    conn.close()
    return RedirectResponse(url="/admin/customer_categories", status_code=303)
    return RedirectResponse(url="/admin/products", status_code=303)

@app.post("/admin/products/edit/{product_id}")
async def edit_product(
    request: Request, 
    product_id: int,
    name: str = Form(...),
    name_marathi: str = Form(""),
    unit: str = Form(...),
    rate: float = Form(0.0),
    favorite_price: Optional[float] = Form(None),
    category_id: str = Form(None),
    image_url: str = Form("")
):
    require_admin(request)
    try:
        # data = await request.json() <- ERROR SOURCE: Form sends form-data, not JSON
        
        # Handle Category ID
        cat_id = None
        if category_id and str(category_id).strip():
            cat_id = int(category_id)
        
        # Validate
        rate = float(rate)
        
        conn = get_db_connection()
        
        # Check if column favorite_price exists
        try:
             conn.execute("UPDATE products SET name=?, name_marathi=?, unit=?, rate=?, favorite_price=?, category_id=?, image_url=? WHERE id=?", 
                     (name, name_marathi, unit, rate, favorite_price, cat_id, image_url, product_id))
        except:
             # Fallback if favorite_price missing in schema
             conn.execute("UPDATE products SET name=?, name_marathi=?, unit=?, rate=?, category_id=?, image_url=? WHERE id=?", 
                     (name, name_marathi, unit, rate, cat_id, image_url, product_id))

        conn.commit()
        conn.close()
        
        # Return redirect or JSON? The form uses standard submit, so we should redirect or return JSON if AJX. 
        # Looking at admin_product_form.html, it's a standard form action. 
        # BUT current implementation was returning JSON. 
        # If the form has no JS interception, it needs a Redirect.
        # Let's check the template again. It has no onsubmit interception.
        # So we MUST return RedirectResponse.
        
        return RedirectResponse(url="/admin/products?success=Product+Updated", status_code=303)
        
    except Exception as e:
        print(f"Error editing product: {e}")
        # return JSONResponse(status_code=500, content={"error": str(e)}) # Avoid JSON for standard form
        return RedirectResponse(url=f"/admin/products/form?id={product_id}&error={str(e)}", status_code=303)

@app.post("/admin/products/bulk_import")
async def bulk_import_products(request: Request, file: UploadFile = File(...)):
    require_admin(request)
    try:
        content = await file.read()
        filename = file.filename.lower()
        
        if filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        elif filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(content))
        else:
            return JSONResponse(status_code=400, content={"error": "Unsupported file format. Use .csv or .xlsx"})

        conn = get_db_connection()
        # Get categories for mapping
        cats = conn.execute("SELECT id, name FROM categories").fetchall()
        cat_map = {str(c['name']).lower().strip(): c['id'] for c in cats}
        
        imported_count = 0
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Standardize column names (strip space)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Mapping template names to internal names
        col_map = {
            'Product Name': 'name',
            'Marathi Name': 'name_marathi',
            'Category': 'category',
            'Unit': 'unit',
            'Rate': 'rate',
            'Favorite Price': 'favorite_price',
            'Image URL': 'image_url'
        }
        
        # Rename columns if they match template names
        for col in df.columns:
            for template_name, internal_name in col_map.items():
                if col.lower() == template_name.lower():
                    df.rename(columns={col: internal_name}, inplace=True)
        
        for _, row in df.iterrows():
            name = str(row.get('name', '')).strip()
            if not name or name == 'nan' or name == '': continue
            
            name_marathi = str(row.get('name_marathi', '')).strip()
            if name_marathi == 'nan': name_marathi = ''
            
            unit = str(row.get('unit', 'pcs')).strip()
            if unit == 'nan': unit = 'pcs'
            
            rate = row.get('rate', 0)
            try: rate = float(rate) if str(rate) != 'nan' else 0
            except: rate = 0
            
            fav_price = row.get('favorite_price', None)
            try: fav_price = float(fav_price) if fav_price is not None and str(fav_price) != 'nan' else None
            except: fav_price = None
            
            img_url = str(row.get('image_url', '')).strip()
            if img_url == 'nan': img_url = ''
            
            cat_name = str(row.get('category', '')).lower().strip()
            cat_id = cat_map.get(cat_name)
            
            conn.execute("""
                INSERT INTO products (name, name_marathi, unit, rate, favorite_price, category_id, image_url, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, name_marathi, unit, rate, fav_price, cat_id, img_url, created_at))
            imported_count += 1
            
        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "imported": imported_count})
        
    except Exception as e:
        print(f"Bulk Import Error: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/admin/pr_list", response_class=HTMLResponse)
async def admin_pr_list(
    request: Request,
    search: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    require_admin(request)
    conn = get_db_connection()
    
    # Base query
    query = '''
        SELECT pr.*, c.name as customer_name, c.business_name, po.id as po_id
        FROM purchase_requisitions pr
        JOIN customers c ON pr.customer_email = c.email
        LEFT JOIN purchase_orders po ON pr.id = po.pr_id AND po.is_active = 1
    '''
    
    params = []
    where_clauses = []
    
    # 1. Status Filter
    if status:
        where_clauses.append("pr.status = ?")
        params.append(status)
    
    # 2. Date Filters
    if date_from:
        where_clauses.append("pr.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
    if date_to:
        where_clauses.append("pr.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
    
    # 3. Basic Field Search (ID, Customer Name, Email)
    # Note: Product name search is handled in Python later for robustness across DB types
    if search:
        search_term = f"%{search.lower().strip()}%"
        where_clauses.append("(CAST(pr.id AS TEXT) LIKE ? OR LOWER(c.name) LIKE ? OR LOWER(pr.customer_email) LIKE ?)")
        params.extend([search_term, search_term, search_term])
    
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    
    query += " ORDER BY pr.created_at DESC"
    
    try:
        rows = conn.execute(query, tuple(params)).fetchall()
    except Exception as e:
        print(f"Error executing PR list query: {e}")
        rows = []
    
    # 4. Product Name Filtering (In Python logic to handle JSON content safely)
    # Also calculate totals and item counts for UI
    
    # Fetch all products for rate estimation
    all_prods = conn.execute("SELECT id, rate FROM products").fetchall()
    prod_rates = {p['id']: p['rate'] for p in all_prods}
    
    filtered_prs = []
    search_lower = search.lower().strip() if search else ""
    
    for row in rows:
        pr = dict(row)
        pr['is_expired'] = is_quotation_expired(pr['created_at']) if pr['status'] == 'QT' else False
        
        # Parse items to calculate total and count
        try:
            items = json.loads(pr['items_json'])
            pr['item_count'] = len(items)
            
            # Calculate total amount
            total = 0
            for item in items:
                # Use quoted amount if available (QT/ACCEPTED statuses), else estimate from base rate
                if 'amount' in item and item['amount'] is not None:
                    total += float(item['amount'])
                elif 'quoted_rate' in item and item['quoted_rate'] is not None:
                    total += float(item['qty']) * float(item['quoted_rate'])
                else:
                    # Estimate based on current product rate
                    p_rate = prod_rates.get(item['product_id'], 0)
                    total += float(item['qty']) * float(p_rate)
            pr['total_amount'] = total
        except:
            pr['item_count'] = 0
            pr['total_amount'] = 0
            items = []

        if search_lower:
            # Check if already matched by ID, name or email
            match_basic = (
                search_lower in str(pr['id']) or 
                search_lower in pr['customer_name'].lower() or 
                search_lower in pr['customer_email'].lower()
            )
            
            if match_basic:
                filtered_prs.append(pr)
                continue
            
            # If not matched basic, check product names inside items_json
            try:
                for item in items:
                    name = item.get('name', '').lower()
                    if search_lower in name:
                        filtered_prs.append(pr)
                        break
            except:
                pass
        else:
            filtered_prs.append(pr)
    
    # Categorize PRs for the separate tables
    pending_prs = []
    sent_prs = []
    accepted_prs = []
    rejected_prs = []
    partially_accepted_prs = []
    
    for pr in filtered_prs:
        if pr['status'] == 'PR':
            pending_prs.append(pr)
        elif pr['status'] == 'QT':
            sent_prs.append(pr)
        elif pr['status'] == 'ACCEPTED':
            accepted_prs.append(pr)
        elif pr['status'] == 'REJECTED':
            rejected_prs.append(pr)
        elif pr['status'] == 'PARTIALLY_ACCEPTED':
            partially_accepted_prs.append(pr)
            
    conn.close()
    
    return templates.TemplateResponse("admin_pr_list.html", {
        "request": request, 
        "prs": filtered_prs,
        "pending_prs": pending_prs,
        "sent_prs": sent_prs,
        "accepted_prs": accepted_prs,
        "rejected_prs": rejected_prs,
        "partially_accepted_prs": partially_accepted_prs,
        "search": search or "",
        "status": status or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "has_results": len(filtered_prs) > 0
    })

@app.get("/admin/quotations/{pr_id}")
async def admin_quotation_plural_redirect(pr_id: int):
    return RedirectResponse(url=f"/admin/quotation/{pr_id}", status_code=303)

@app.get("/admin/quotation/{pr_id}", response_class=HTMLResponse)
async def admin_quotation(request: Request, pr_id: int):
    require_admin(request)
    conn = get_db_connection()
    
    # Fetch PR with Customer Name and Favorite Status
    query = '''
        SELECT pr.*, c.name as customer_name, c.business_name, c.is_favorite 
        FROM purchase_requisitions pr
        JOIN customers c ON pr.customer_email = c.email
        WHERE pr.id = ?
    '''
    pr = conn.execute(query, (pr_id,)).fetchone()
    
    if not pr:
        conn.close()
        return RedirectResponse(url="/admin/pr_list", status_code=303)
        
    items = json.loads(pr["items_json"])
    
    # Hydrate items with current product names
    hydrated_items = []
    for item in items:
        prod = conn.execute("SELECT * FROM products WHERE id = ?", (item['product_id'],)).fetchone()
        if prod:
            # Check for favorite price
            suggested_rate = prod['rate']
            
            # Safe check for favorite status and price
            is_favorite_applied = False
            try:
                is_fav = pr['is_favorite']
                fav_price = prod['favorite_price']
                if is_fav and fav_price and fav_price > 0:
                    suggested_rate = fav_price
                    is_favorite_applied = True
            except:
                pass
                
            hydrated_items.append({
                "product_id": item['product_id'],
                "name": prod['name'],
                "name_marathi": prod['name_marathi'],
                "qty": item['qty'],
                "unit": prod['unit'],
                "suggested_rate": suggested_rate,
                "is_favorite_applied": is_favorite_applied,
                "shop_stock": prod['current_stock'] if 'current_stock' in prod.keys() else 0,
                "reorder_level_shop": prod['reorder_level_shop'] if 'reorder_level_shop' in prod.keys() else 0
            })
            
    conn.close()
    return templates.TemplateResponse("admin_quotation.html", {"request": request, "pr": pr, "items": hydrated_items})

@app.post("/admin/quotation/{pr_id}/submit")
async def submit_quotation(request: Request, pr_id: int):
    require_admin(request)
    form_data = await request.form()
    
    # Process form data to build items with rates
    # Form keys: rate_{product_id}
    
    conn = get_db_connection()
    pr = conn.execute("SELECT * FROM purchase_requisitions WHERE id = ?", (pr_id,)).fetchone()
    if not pr:
        conn.close()
        return RedirectResponse(url="/admin/pr_list?error=PR+not+found", status_code=303)
        
    # Prevent duplicate quotation
    if pr['status'] != 'PR':
        conn.close()
        return RedirectResponse(url="/admin/pr_list?error=Quotation+already+sent+or+status+invalid", status_code=303)
        
    old_items = json.loads(pr["items_json"])
    new_items = []
    
    for item in old_items:
        pid = item['product_id']
        rate_key = f"rate_{pid}"
        if rate_key in form_data:
            new_rate = float(form_data[rate_key])
            # Hydrate name/unit for permanent record in PR (so if product deleted, record remains)
            prod = conn.execute("SELECT name, name_marathi, unit FROM products WHERE id = ?", (pid,)).fetchone()
            name = prod['name'] if prod else "Unknown"
            name_marathi = prod['name_marathi'] if prod else ""
            unit = prod['unit'] if prod else "N/A"
            
            new_items.append({
                "product_id": pid,
                "name": name,
                "name_marathi": name_marathi,
                "qty": item['qty'],
                "unit": unit,
                "quoted_rate": new_rate,
                "amount": item['qty'] * new_rate,
                "item_status": "PENDING",  # New field for item-level tracking
                "decision_timestamp": None,
                "is_locked": False
            })
            
    conn.execute("UPDATE purchase_requisitions SET status = 'QT', items_json = ? WHERE id = ?", 
                 (json.dumps(new_items), pr_id))
    conn.commit()

    try:
        from app.notifications import send_email_notification, dispatch_pending_notifications

        # Get supplier name for email
        supplier_name = get_setting("supplier_name", "Hotel Supplier Inc.")
        
        # ENHANCED EMAIL TEMPLATE WITH SUPPLIER INFO, PR NUMBER, AND DIRECT LINK
        subject = f"📄 New Quotation Received for PR #{pr_id}"
        message = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
    <div style="max-width: 600px; margin: 20px auto; padding: 20px; border: 1px solid #ddd; border-radius: 8px; background-color: #ffffff;">
        
        <!-- Header -->
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #2c3e50; margin: 0; font-size: 24px;">📄 New Quotation Received</h1>
            <p style="color: #7f8c8d; margin: 5px 0 0 0; font-size: 14px;">PR #{pr_id} - Action Required</p>
        </div>
        
        <!-- Supplier Information -->
        <div style="background-color: #e3f2fd; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #2196f3;">
            <h3 style="color: #1565c0; margin: 0 0 15px 0; font-size: 16px;">🏢 Supplier Details</h3>
            <p style="margin: 8px 0;"><strong>Supplier Name:</strong> {supplier_name}</p>
            <p style="margin: 8px 0;"><strong>PR Number:</strong> #{pr_id}</p>
        </div>
        
        <!-- Action Required -->
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #007bff;">
            <h3 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 16px;">📋 Quotation Received</h3>
            <p style="margin: 0 0 15px 0;">A quotation has been submitted by <strong>{supplier_name}</strong> for your purchase request PR #<strong>{pr_id}</strong>.</p>
            <p style="margin: 0 0 15px 0;">Please review the quotation details at your earliest convenience and decide whether to accept or reject the items.</p>
            
            <ul style="margin: 15px 0; padding-left: 20px;">
                <li style="margin: 8px 0;">Review quotation details and pricing</li>
                <li style="margin: 8px 0;">Accept or reject individual items</li>
                <li style="margin: 8px 0;">Confirm your decision</li>
            </ul>
        </div>
        
        <!-- CTA Button -->
        <div style="text-align: center; margin: 30px 0;">
            <a href="{APP_BASE_URL}/customer/quotation/{pr_id}" style="display: inline-block; background-color: #007bff; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 16px;">
                🚀 View Quotation
            </a>
        </div>
        
        <!-- Footer -->
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; text-align: center;">
            <p style="color: #6c757d; margin: 5px 0; font-size: 14px;">Thank you for your business!</p>
            <p style="color: #6c757d; margin: 5px 0; font-size: 12px;">This is an automated message. Please do not reply to this email.</p>
        </div>
        
    </div>
</body>
</html>
""".strip()

        notification_id = send_email_notification(
            event_type="QUOTATION_SENT",
            ref_id=int(pr_id),
            recipient_role="customer",
            recipient_email=pr['customer_email'],
            subject=subject,
            message=message,
            send_email=False,
        )

        if notification_id:
            print(f"[INFO] QUOTATION_SENT notification logged with ID: {notification_id}")
        else:
            print(f"[WARN] Failed to log QUOTATION_SENT notification")

        dispatch_pending_notifications()
    except Exception as e:
        print(f"[WARN] QUOTATION_SENT notification log failed (non-blocking): {e}")

    conn.close()
    
    return RedirectResponse(url="/admin/pr_list?success=Quotation+sent+successfully", status_code=303)

@app.get("/admin/rejected_items", response_class=HTMLResponse)
async def admin_rejected_items(
    request: Request,
    search: Optional[str] = None,
    customer_email: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    require_admin(request)
    conn = get_db_connection()
    
    # Get all customers for filter
    customers = conn.execute("SELECT DISTINCT email, name FROM customers ORDER BY name").fetchall()
    
    # Get all PRs with potential rejected items
    query = '''
        SELECT pr.*, c.name as customer_name 
        FROM purchase_requisitions pr
        JOIN customers c ON pr.customer_email = c.email
        WHERE pr.status IN ('REJECTED', 'PARTIALLY_ACCEPTED')
    '''
    params = []
    if customer_email:
        query += " AND pr.customer_email = ?"
        params.append(customer_email)
    
    if date_from:
        query += " AND pr.created_at >= ?"
        params.append(f"{date_from} 00:00:00")
        
    if date_to:
        query += " AND pr.created_at <= ?"
        params.append(f"{date_to} 23:59:59")
        
    query += " ORDER BY pr.created_at DESC"
    
    prs = conn.execute(query, tuple(params)).fetchall()
    
    # Extract rejected items from each PR and apply text search
    rejected_data = []
    search_lower = search.lower().strip() if search else None
    
    for pr in prs:
        items = json.loads(pr['items_json'])
        for item in items:
            if item.get('item_status') == 'REJECTED':
                # Apply search filter on PR ID, Item Name, or Customer Name if search provided
                match = True
                if search_lower:
                    match = (
                        search_lower in str(pr['id']) or
                        search_lower in item['name'].lower() or
                        search_lower in pr['customer_name'].lower() or
                        search_lower in pr['customer_email'].lower()
                    )
                
                if match:
                    rejected_data.append({
                        'pr_id': pr['id'],
                        'customer_name': pr['customer_name'],
                        'customer_email': pr['customer_email'],
                        'item_name': item['name'],
                        'qty': item['qty'],
                        'unit': item['unit'],
                        'quoted_rate': item.get('quoted_rate', 0),
                        'amount': item.get('amount', 0),
                        'decision_timestamp': item.get('decision_timestamp', 'N/A'),
                        'pr_created_at': pr['created_at']
                    })
    
    conn.close()
    return templates.TemplateResponse("admin_rejected_items.html", {
        "request": request,
        "rejected_items": rejected_data,
        "customers": customers,
        "search": search or "",
        "selected_customer": customer_email or "",
        "date_from": date_from or "",
        "date_to": date_to or ""
    })



# ============================================================
# [NEW WORKFLOW ADDITION] Admin Price Review Routes
# Workflow: PR → Admin Review & Edit Prices → Generate Invoice
# This is an ADDITIVE feature. Existing quotation workflow unchanged.
# ============================================================

@app.get("/admin/price_review/{pr_id}", response_class=HTMLResponse)
async def admin_price_review(request: Request, pr_id: int):
    """
    [NEW WORKFLOW ADDITION]
    Admin Price Review: Admin edits item prices and generates invoice directly.
    Only accessible for PRs with status=PENDING_PRICE_REVIEW.
    """
    require_admin(request)
    conn = get_db_connection()

    query = '''
        SELECT pr.*, c.name as customer_name, c.business_name, c.address, c.mobile, c.category
        FROM purchase_requisitions pr
        JOIN customers c ON pr.customer_email = c.email
        WHERE pr.id = ?
    '''
    pr = conn.execute(query, (pr_id,)).fetchone()

    if not pr:
        conn.close()
        return RedirectResponse(url="/admin/pr_list?error=PR+not+found", status_code=303)

    # Safety guard: only allow price review for PENDING_PRICE_REVIEW PRs
    if pr['status'] != 'PENDING_PRICE_REVIEW':
        conn.close()
        return RedirectResponse(
            url=f"/admin/quotation/{pr_id}?info=This+PR+uses+the+standard+quotation+workflow",
            status_code=303
        )

    items_raw = json.loads(pr["items_json"])

    # Hydrate each item with current product details
    hydrated_items = []
    for item in items_raw:
        prod = conn.execute(
            "SELECT id, name, name_marathi, unit, rate FROM products WHERE id = ?",
            (item['product_id'],)
        ).fetchone()
        if prod:
            hydrated_items.append({
                "product_id": item['product_id'],
                "name": prod['name'],
                "name_marathi": prod['name_marathi'] or "",
                "unit": prod['unit'],
                "qty": item['qty'],
                "suggested_rate": prod['rate'],  # Current catalogue rate as suggestion
            })

    conn.close()
    supplier_name = get_setting("supplier_name", "Hotel Supplier Inc.")
    return templates.TemplateResponse("admin_price_review.html", {
        "request": request,
        "pr": pr,
        "items": hydrated_items,
        "supplier_name": supplier_name,
    })


@app.post("/admin/price_review/{pr_id}/generate_invoice")
async def admin_price_review_generate_invoice(request: Request, pr_id: int):
    """
    [NEW WORKFLOW ADDITION]
    Admin Price Review: Generate invoice directly from PR after editing prices.
    Skips quotation step entirely. Reuses existing purchase_orders + invoices tables.
    Deducts stock via existing InventoryService (same as direct_sales).
    """
    require_admin(request)
    form_data = await request.form()
    conn = get_db_connection()

    # --- 1. Fetch and validate the PR ---
    pr = conn.execute(
        "SELECT * FROM purchase_requisitions WHERE id = ?", (pr_id,)
    ).fetchone()

    if not pr or pr['status'] != 'PENDING_PRICE_REVIEW':
        conn.close()
        return RedirectResponse(url="/admin/pr_list?error=PR+not+found+or+not+in+price+review+status", status_code=303)

    # --- 2. Fetch customer info for PO snapshots ---
    customer = conn.execute(
        "SELECT name, business_name, address, category, email, mobile FROM customers WHERE email = ?",
        (pr['customer_email'],)
    ).fetchone()

    if not customer:
        conn.close()
        return RedirectResponse(url=f"/admin/price_review/{pr_id}?error=Customer+not+found", status_code=303)

    # --- 3. Build items list from admin-edited form prices ---
    old_items = json.loads(pr["items_json"])
    new_items = []
    total_amount = 0.0
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for item in old_items:
        pid = item['product_id']
        rate_key = f"rate_{pid}"
        discount_key = f"discount_{pid}"

        # Read admin-entered rate (fall back to form if present, else 0)
        try:
            admin_rate = float(form_data.get(rate_key, 0))
        except (ValueError, TypeError):
            admin_rate = 0.0

        # Optional per-item discount percentage
        try:
            discount_pct = float(form_data.get(discount_key, 0))
        except (ValueError, TypeError):
            discount_pct = 0.0

        if admin_rate < 0:
            admin_rate = 0.0
        if not (0 <= discount_pct <= 100):
            discount_pct = 0.0

        # Apply discount
        effective_rate = admin_rate * (1 - discount_pct / 100.0)
        qty = float(item['qty'])
        amount = round(qty * effective_rate, 2)

        # Fetch current product metadata for the snapshot
        prod = conn.execute(
            "SELECT name, name_marathi, unit FROM products WHERE id = ?", (pid,)
        ).fetchone()
        name = prod['name'] if prod else item.get('name', 'Unknown')
        name_marathi = prod['name_marathi'] if prod else ""
        unit = prod['unit'] if prod else item.get('unit', '')

        new_items.append({
            "product_id": pid,
            "name": name,
            "name_marathi": name_marathi,
            "unit": unit,
            "qty": qty,
            "quoted_rate": round(effective_rate, 4),
            "original_rate": admin_rate,
            "discount_pct": discount_pct,
            "amount": amount,
            "item_status": "ACCEPTED",
            "decision_timestamp": created_at,
            "is_locked": True,
            # Audit: mark as generated via admin price review
            "workflow": "ADMIN_PRICE_REVIEW",
        })
        total_amount += amount

    if not new_items:
        conn.close()
        return RedirectResponse(url=f"/admin/price_review/{pr_id}?error=No+valid+items+found", status_code=303)

    total_amount = round(total_amount, 2)

    try:
        # --- 4. Create Purchase Order ---
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO purchase_orders (
                pr_id, customer_email, created_at, total_amount, items_json,
                invoice_source, delivery_status, tracking_status, status, is_active,
                customer_name_snapshot, business_name_snapshot, address_snapshot,
                customer_category_snapshot, customer_email_snapshot, customer_mobile_snapshot
            ) VALUES (?, ?, ?, ?, ?, 'ADMIN_PRICE_REVIEW', 'OPEN', 'PO_CREATED', 'Accepted', 1,
                      ?, ?, ?, ?, ?, ?)
        ''', (
            pr_id, pr['customer_email'], created_at, total_amount, json.dumps(new_items),
            customer['name'], customer['business_name'], customer['address'],
            customer['category'], customer['email'], customer['mobile']
        ))

        po_id = cursor.lastrowid
        display_id = f"APR-{po_id}"  # APR = Admin Price Review
        cursor.execute("UPDATE purchase_orders SET display_id = ? WHERE id = ?", (display_id, po_id))

        # --- 5. Auto-generate Invoice ---
        today_str = datetime.now().strftime("%Y%m%d")
        invoice_no = f"APR-INV-{today_str}-{po_id}"
        admin_notes = form_data.get("admin_notes", "").strip()

        cursor.execute("""
            INSERT INTO invoices (po_id, invoice_no, created_at, payment_mode, delivery_remarks, items_json, status)
            VALUES (?, ?, ?, ?, ?, ?, 'GENERATED')
        """, (
            po_id, invoice_no, created_at,
            form_data.get("payment_mode", "Credit"),
            admin_notes or "Admin Price Review",
            json.dumps(new_items)
        ))

        # --- 6. Update PR status to INVOICED_DIRECT ---
        if DATABASE_URL:
            conn.execute(
                "UPDATE purchase_requisitions SET status = 'INVOICED_DIRECT', items_json = %s WHERE id = %s",
                (json.dumps(new_items), pr_id)
            )
        else:
            conn.execute(
                "UPDATE purchase_requisitions SET status = 'INVOICED_DIRECT', items_json = ? WHERE id = ?",
                (json.dumps(new_items), pr_id)
            )

        # --- 7. Deduct Stock via InventoryService (same as direct_sales) ---
        try:
            inv_service = InventoryService(conn)
            for item in new_items:
                inv_service.record_sale_out(
                    product_id=int(item['product_id']),
                    qty=float(item['qty']),
                    reference_type='ADMIN_PRICE_REVIEW_PO',
                    reference_id=display_id,
                    commit=False  # Commit together below
                )
        except Exception as inv_err:
            print(f"[WARN] Stock deduction failed for APR PO {display_id}: {inv_err}")
            # We do NOT abort the invoice—stock issue is non-fatal for invoice creation
            # but log it clearly for admin awareness.

        conn.commit()

        # --- 8. Audit Notification ---
        try:
            admin_email_setting = get_setting("admin_email", ADMIN_EMAIL)
            create_notification(
                conn,
                admin_email_setting,
                "admin",
                f"[Price Review] Invoice {invoice_no} generated for PR #{pr_id} — Customer: {customer['name']}",
                f"/invoice/po/{po_id}"
            )
            conn.commit()

            # Notify customer that their order is invoiced
            create_notification(
                conn,
                pr['customer_email'],
                "customer",
                f"Your order (PR #{pr_id}) has been processed and invoiced by admin.",
                f"/invoice/po/{po_id}"
            )
            conn.commit()
        except Exception as notif_err:
            print(f"[WARN] Notification failed for APR invoice: {notif_err}")

        conn.close()
        return RedirectResponse(url=f"/invoice/po/{po_id}?success=Invoice+generated+via+Admin+Price+Review", status_code=303)

    except Exception as e:
        print(f"[ERROR] Admin Price Review invoice generation failed: {e}")
        try:
            conn.rollback()
        except Exception:
            pass
        conn.close()
        return RedirectResponse(
            url=f"/admin/price_review/{pr_id}?error=Invoice+generation+failed.+Please+try+again.",
            status_code=303
        )

# ============================================================
# [END NEW WORKFLOW ADDITION]
# ============================================================


# --- Routes: Customer ---

@app.get("/customer/dashboard", response_class=HTMLResponse)
async def customer_dashboard(request: Request):
    user = require_customer(request)
    conn = get_db_connection()
    
    stats = {
        "total_prs": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE customer_email=? AND status != 'DRAFT'", (user['email'],)).fetchone()[0],
        # ENHANCED: Quotation Inbox now includes BOTH:
        # - PRs awaiting supplier quotation (status='PR')
        # - PRs with quotations awaiting customer decision (status='QT')
        "received_qt": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE customer_email=? AND status IN ('PR', 'QT')", (user['email'],)).fetchone()[0],
        "accepted": conn.execute("SELECT COUNT(*) FROM purchase_requisitions WHERE customer_email=? AND status IN ('ACCEPTED', 'PARTIALLY_ACCEPTED')", (user['email'],)).fetchone()[0],
        "direct_orders": conn.execute("SELECT COUNT(*) FROM purchase_orders WHERE customer_email=? AND pr_id IS NULL AND is_active = 1", (user['email'],)).fetchone()[0],
    }
    # Adjust 'accepted' to include direct orders for the dashboard card
    stats["accepted"] += stats["direct_orders"]
    conn.close()
    
    return templates.TemplateResponse("customer_dashboard.html", {"request": request, "stats": stats})

@app.get("/customer/quotations", response_class=HTMLResponse)
async def customer_quotations(request: Request, view: Optional[str] = None):
    user = require_customer(request)
    customer_email = user['email']
    conn = get_db_connection()
    
    view_title = "My Quotations & Orders"
    status_filter = ""
    
    if view == "quotations":
        view_title = "Quotation Inbox"
        # ENHANCED: Quotation Inbox now shows full lifecycle:
        # - PR status: Awaiting supplier quotation
        # - QT status: Quotation received, awaiting customer decision
        status_filter = "AND pr.status IN ('PR', 'QT')"
    elif view == "orders":
        view_title = "Active Orders"
        # Active Orders: Only show items with PO (customer has accepted)
        status_filter = "AND pr.status IN ('ACCEPTED', 'PARTIALLY_ACCEPTED')"
    elif view == "rejected":
        view_title = "Rejected Quotations"
        status_filter = "AND pr.status = 'REJECTED'"
    else:
        # Default: show all non-PR, non-DRAFT items
        status_filter = "AND pr.status NOT IN ('PR', 'DRAFT')"

    # Fetch PRs
    prs_rows = conn.execute(f'''
        SELECT pr.*, 
               po.id as po_id
        FROM purchase_requisitions pr
        LEFT JOIN purchase_orders po ON pr.id = po.pr_id
        WHERE pr.customer_email = ? {status_filter}
        ORDER BY pr.created_at DESC
    ''', (customer_email,)).fetchall()
    
    # Fetch direct orders (only if not specifically viewing quotations or rejected)
    direct_orders_rows = []
    if view != "quotations" and view != "rejected":
        direct_orders_rows = conn.execute('''
            SELECT 
                NULL as id,
                customer_email,
                'ACCEPTED' as status,
                created_at,
                items_json,
                '' as admin_notes,
                id as po_id
            FROM purchase_orders po
            WHERE customer_email = ? AND pr_id IS NULL AND is_active = 1
            ORDER BY created_at DESC
        ''', (customer_email,)).fetchall()
    
    # Hydrate with product names for search
    all_products = conn.execute("SELECT id, name, name_marathi FROM products").fetchall()
    prod_map = {p['id']: p for p in all_products}
    
    all_rows = []
    for row in list(prs_rows) + list(direct_orders_rows):
        pr = dict(row)
        try:
            items = json.loads(pr['items_json'])
            names = []
            for item in items:
                prod = prod_map.get(item['product_id'])
                if prod:
                    nm = prod['name']
                    if prod['name_marathi']:
                        nm += f" / {prod['name_marathi']}"
                    names.append(nm)
            pr['product_names'] = ", ".join(names)
            pr['is_expired'] = is_quotation_expired(pr['created_at']) if pr['status'] == 'QT' else False
        except:
            pr['product_names'] = ""
            pr['is_expired'] = False
        all_rows.append(pr)
    
    # Sort by created_at desc
    all_rows.sort(key=lambda x: x['created_at'], reverse=True)
    
    conn.close()
    return templates.TemplateResponse("customer_quotation_list.html", {
        "request": request, 
        "prs": all_rows,
        "view_title": view_title
    })

@app.get("/customer/change-password", response_class=HTMLResponse)
async def customer_change_password_page(request: Request):
    require_customer(request)
    return templates.TemplateResponse("customer_change_password.html", {"request": request})

@app.post("/customer/change-password")
async def customer_change_password(request: Request, current_password: str = Form(...)):
    user = require_customer(request)
    email = user["email"]
    
    # Verify current password
    conn = get_db_connection()
    db_user = conn.execute("SELECT password FROM customers WHERE email = ?", (email,)).fetchone()
    conn.close()
    
    if not db_user or db_user["password"] != current_password:
        return templates.TemplateResponse("customer_change_password.html", {
            "request": request, "error": "Incorrect current password"
        })
    
    # Generate OTP
    otp = utils.generate_otp()
    
    # Store OTP
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    expires_at = (datetime.now() + timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")
    
    conn = get_db_connection()
    conn.execute("INSERT INTO otp_codes (email, otp, created_at, expires_at) VALUES (?, ?, ?, ?) ON CONFLICT(email) DO UPDATE SET otp=excluded.otp, created_at=excluded.created_at, expires_at=excluded.expires_at", 
                 (email, otp, created_at, expires_at))
    conn.commit()
    conn.close()
    
    # Send Email
    sender = get_setting("otp_sender_email", "")
    sender_pass = get_setting("otp_password", "")
    
    if sender and sender_pass:
        utils.send_email_otp(email, otp, sender, sender_pass)
        
    request.session["pending_cust_pass_change"] = True
    
    return RedirectResponse(url="/verify/customer-password", status_code=303)

@app.get("/verify/customer-password", response_class=HTMLResponse)
async def verify_customer_password_page(request: Request):
    user = require_customer(request)
    if not request.session.get("pending_cust_pass_change"):
        return RedirectResponse(url="/customer/change-password", status_code=303)
        
    masked = utils.mask_email(user["email"])
    return templates.TemplateResponse("auth_verify_otp.html", {
        "request": request, 
        "email": masked,
        "target_url": "/verify/customer-password"
    })

@app.post("/verify/customer-password")
async def verify_customer_password(request: Request, otp: str = Form(...)):
    user = require_customer(request)
    if not request.session.get("pending_cust_pass_change"):
        return RedirectResponse(url="/customer/change-password", status_code=303)
        
    email = user["email"]
    
    # Verify OTP
    conn = get_db_connection()
    record = conn.execute("SELECT * FROM otp_codes WHERE email = ?", (email,)).fetchone()
    conn.close()
    
    
    valid = False
    if record:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if record["otp"] == otp:
            # Check expiry. Note: string comparison works for ISO format
            if now < record["expires_at"]: 
                valid = True
                
    if not valid:
        return templates.TemplateResponse("auth_verify_otp.html", {
            "request": request, 
            "email": utils.mask_email(email),
            "target_url": "/verify/customer-password",
            "error": "Invalid or Expired OTP"
        })
        
    # Mark Verified
    request.session["pending_cust_pass_change"] = "VERIFIED"
    
    return RedirectResponse(url="/customer/set-new-password", status_code=303)

    return RedirectResponse(url="/customer/set-new-password", status_code=303)

# -------------------------------------------------------------------------
# NEW: OTP Based Password Change Flow (Additive)
# -------------------------------------------------------------------------

@app.get("/auth/change-password-otp", response_class=HTMLResponse)
async def auth_change_password_page_otp(request: Request):
    """Entry point for the new flow."""
    require_customer(request)
    # Clear any stale session state
    request.session.pop("otp_verified_email", None)
    return templates.TemplateResponse("auth_change_password_otp.html", {"request": request})

@app.post("/auth/request-password-otp")
async def request_password_otp_api(request: Request):
    """Step 1: Send OTP"""
    try:
        user = require_customer(request) # Ensure logged in
        data = await request.json()
        email = data.get("email", "").strip().lower()
        
        if not email:
            return JSONResponse({"success": False, "error": "Email is required"})
            
        if email != user['email'].lower():
             return JSONResponse({"success": False, "error": "Email does not match logged in user"})

        # Generate OTP
        otp = utils.generate_otp()
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        expires_at = (datetime.now() + timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")

        conn = get_db_connection()
        try:
            # Upsert OTP without is_used
            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                # Delete existing OTP for this email first
                conn.execute("DELETE FROM otp_codes WHERE email = %s", (email,))
                
                # Insert new OTP
                conn.execute("""
                    INSERT INTO otp_codes (email, otp, created_at, expires_at) 
                    VALUES (%s, %s, %s, %s)
                 """, (email, otp, created_at, expires_at))
            else:
                # Delete existing OTP for this email first
                conn.execute("DELETE FROM otp_codes WHERE email = ?", (email,))
                
                # Insert new OTP
                conn.execute("""
                    INSERT INTO otp_codes (email, otp, created_at, expires_at) 
                    VALUES (?, ?, ?, ?)
                 """, (email, otp, created_at, expires_at))
            
            conn.commit()
            
            # Send Email
            from app.notifications import call_supabase_email_function
            subject = "Password Change OTP"
            body = f"Your OTP for password change is {otp}. Valid for 10 minutes."
            
            # Non-blocking email call (best effort)
            try:
                call_supabase_email_function(None, email, subject, body)
            except Exception as e:
                print(f"OTP Email failed: {e}")
                
            return JSONResponse({"success": True, "message": "OTP sent successfully"})
            
        except Exception as e:
            if hasattr(conn, 'rollback'):
                conn.rollback()
            print(f"OTP Request Error: {e}")
            return JSONResponse({"success": False, "error": "Database error"})
        finally:
            conn.close()
            
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

@app.post("/auth/verify-password-otp")
async def verify_password_otp_api(request: Request):
    """Step 2: Verify OTP"""
    try:
        user = require_customer(request)
        data = await request.json()
        email = data.get("email", "").strip().lower()
        otp = data.get("otp", "").strip()
        
        if not email or not otp:
            return JSONResponse({"success": False, "error": "Email and OTP required"})

        if email != user['email'].lower():
             return JSONResponse({"success": False, "error": "Email mismatch"})

        conn = get_db_connection()
        try:
            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                row = conn.execute("SELECT * FROM otp_codes WHERE email = %s", (email,)).fetchone()
            else:
                row = conn.execute("SELECT * FROM otp_codes WHERE email = ?", (email,)).fetchone()
            
            if not row:
                return JSONResponse({"success": False, "error": "Invalid or Expired OTP"})
            
            # Check validity
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Check OTP match
            if str(row['otp']) != str(otp):
                 return JSONResponse({"success": False, "error": "Invalid OTP"})
            
            # Check Expiry
            if str(row['expires_at']) < now:
                return JSONResponse({"success": False, "error": "OTP Expired"})
                
            # SUCCESS: Delete OTP to prevent reuse
            if is_pg:
                conn.execute("DELETE FROM otp_codes WHERE email = %s", (email,))
            else:
                conn.execute("DELETE FROM otp_codes WHERE email = ?", (email,))
            
            conn.commit()
            
            # Set Session Flag to allow password update
            request.session["otp_verified_email"] = email
            
            return JSONResponse({"success": True, "message": "OTP Verified"})
            
        finally:
            conn.close()
            
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

@app.post("/auth/update-password")
async def update_password_api(request: Request):
    """Step 3: Update Password"""
    try:
        user = require_customer(request)
        data = await request.json()
        email = data.get("email", "").strip().lower()
        new_password = data.get("new_password", "")
        
        if not email or not new_password:
             return JSONResponse({"success": False, "error": "Missing fields"})
             
        # Security Check: Session flag must be present and match
        verified_email = request.session.get("otp_verified_email")
        if not verified_email or verified_email != email:
             return JSONResponse({"success": False, "error": "Session expired or unauthorized. Please verify OTP again."})
        
        if len(new_password) < 6:
            return JSONResponse({"success": False, "error": "Password too short"})

        conn = get_db_connection()
        try:
            conn.execute("UPDATE customers SET password = ? WHERE email = ?", (new_password, email))
            conn.commit()
            
            # Clear flag
            request.session.pop("otp_verified_email", None)
            
            return JSONResponse({"success": True, "message": "Password updated successfully"})
        finally:
            conn.close()

    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})
async def fix_database_schema(request: Request):
    require_admin(request)
    logs = []
    
    print("FIX-DB: Starting schema repair...")
    logs.append("Starting schema repair...")
    
    conn = None
    try:
        conn = get_db_connection()
        print("FIX-DB: Connected to DB")
        
        # 1. Determine columns
        columns = []
        is_postgres = False
        
        # Check if wrapped
        raw_conn = conn
        if hasattr(conn, 'conn'):
            raw_conn = conn.conn
            is_postgres = True
            print("FIX-DB: Detected PostgreSQL Wrapper")
        
        # Get cursor
        if hasattr(conn, 'cursor'):
            cursor = conn.cursor()
        else:
            cursor = raw_conn.cursor() # Fallback
            
        try:
            print("FIX-DB: Checking columns...")
            # Try generic SQL first or specific based on type
            # For Postgres, information_schema is best
            cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = 'products'")
            rows = cursor.fetchall()
            if rows:
                columns = [row[0] for row in rows]
                is_postgres = True
            else:
                # SQLite fallback
                cursor.execute("PRAGMA table_info(products)")
                rows = cursor.fetchall()
                columns = [row[1] for row in rows]
        except Exception as e:
            print(f"FIX-DB: Column check failed 1st attempt: {e}")
            # SQLite fallback attempt if 1st failed
            try:
                cursor.execute("PRAGMA table_info(products)")
                rows = cursor.fetchall()
                columns = [row[1] for row in rows]
            except Exception as e2:
                print(f"FIX-DB: Column check failed 2nd attempt: {e2}")
                columns = []

        print(f"FIX-DB: Current columns: {columns}")
        logs.append(f"Current columns: {columns}")
        
        # Close cursor to be clean
        try:
             cursor.close()
        except:
             pass
             
        # Commit to release any potential read locks before DDL
        if hasattr(conn, 'commit'):
            conn.commit()
        
        # 2. Prepare Updates
        missing_updates = []
        
        if 'current_stock' not in columns:
            missing_updates.append("ALTER TABLE products ADD COLUMN current_stock DECIMAL(10,2) DEFAULT 0")
        
        if 'reorder_level_shop' not in columns:
            missing_updates.append("ALTER TABLE products ADD COLUMN reorder_level_shop INTEGER DEFAULT 5")
            
        if 'reorder_level_godown' not in columns:
            missing_updates.append("ALTER TABLE products ADD COLUMN reorder_level_godown INTEGER DEFAULT 10")
            
        if 'favorite_price' not in columns:
            missing_updates.append("ALTER TABLE products ADD COLUMN favorite_price DECIMAL(10,2) DEFAULT NULL")
            
        print(f"FIX-DB: Found {len(missing_updates)} missing columns to add.")
        
        # 3. Execute Updates
        # Create new cursor
        if hasattr(conn, 'cursor'):
            cursor = conn.cursor()
        else:
            cursor = raw_conn.cursor()
            
        for i, sql in enumerate(missing_updates):
            try:
                print(f"FIX-DB: Executing [{i+1}/{len(missing_updates)}]: {sql}")
                # Use cursor.execute directly to avoid wrapper weirdness
                cursor.execute(sql)
                # Commit immediately after DDL
                if hasattr(conn, 'commit'):
                    conn.commit()
                logs.append(f"SUCCESS: Executed: {sql}")
                print(f"FIX-DB: Success")
            except Exception as e:
                print(f"FIX-DB: Failed: {e}")
                if hasattr(conn, 'rollback'):
                    conn.rollback()
                logs.append(f"ERROR: Failed to execute {sql}: {e}")
                
        # 4. Ledger Table logic
        print("FIX-DB: Checking inventory_ledger..")
        try:
            cursor.execute("SELECT count(*) FROM inventory_ledger")
            cursor.fetchone() # Consume result
            print("FIX-DB: inventory_ledger exists")
        except Exception:
             print("FIX-DB: inventory_ledger missing, creating...")
             if hasattr(conn, 'rollback'):
                 conn.rollback()
             
             logs.append("Attempting to create inventory_ledger table...")
             try:
                 sql = """
                 CREATE TABLE IF NOT EXISTS inventory_ledger (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL,
                    location_id INTEGER NOT NULL,
                    transaction_type VARCHAR(50) NOT NULL,
                    qty_change DECIMAL(10,2) NOT NULL, 
                    qty_in DECIMAL(10,2) DEFAULT 0,
                    qty_out DECIMAL(10,2) DEFAULT 0,
                    running_balance DECIMAL(10,2) DEFAULT 0, 
                    batch_id VARCHAR(50), 
                    reference_type VARCHAR(50), 
                    reference_id VARCHAR(50),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    created_by VARCHAR(50)
                );
                 """
                 if not is_postgres: 
                     # SQLite adjust
                     sql = sql.replace("SERIAL PRIMARY KEY", "INTEGER PRIMARY KEY AUTOINCREMENT")
                     sql = sql.replace("TIMESTAMP DEFAULT CURRENT_TIMESTAMP", "TEXT DEFAULT CURRENT_TIMESTAMP")
                 
                 cursor.execute(sql)
                 if hasattr(conn, 'commit'):
                    conn.commit()
                 logs.append("Created inventory_ledger table")
                 print("FIX-DB: Created inventory_ledger table")
             except Exception as e2:
                 logs.append(f"Failed to create ledger: {e2}")
                 print(f"FIX-DB: Failed to create ledger: {e2}")

    except Exception as e:
        print(f"FIX-DB: FATAL ERROR: {e}")
        logs.append(f"Fatal Error: {e}")
    finally:
        if conn:
            try:
                conn.close()
                print("FIX-DB: Connection closed")
            except:
                pass
        
    return JSONResponse({"status": "completed", "logs": logs})
    if record:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if record["otp"] == otp:
            if now < record["expires_at"]:
                valid = True
                
    if not valid:
        return templates.TemplateResponse("auth_verify_otp.html", {
            "request": request, 
            "email": utils.mask_email(email),
            "target_url": "/verify/customer-password",
            "error": "Invalid or Expired OTP"
        })
        
    # Mark Verified
    request.session["pending_cust_pass_change"] = "VERIFIED"
    
    # Clean OTP (optional here, or clean after next step)
    # Keeping OTP valid for few mins or until used is fine, but best to clean up to prevent reuse if we relied on OTP again.
    # Here we rely on session "VERIFIED" state.
    
    return RedirectResponse(url="/customer/set-new-password", status_code=303)

@app.get("/customer/set-new-password", response_class=HTMLResponse)
async def customer_set_password_page(request: Request):
    require_customer(request)
    if request.session.get("pending_cust_pass_change") != "VERIFIED":
        return RedirectResponse(url="/customer/change-password", status_code=303)
    return templates.TemplateResponse("customer_set_password.html", {"request": request})

@app.post("/customer/set-new-password")
async def customer_set_password(request: Request, new_password: str = Form(...), confirm_password: str = Form(...)):
    user = require_customer(request)
    if request.session.get("pending_cust_pass_change") != "VERIFIED":
        return RedirectResponse(url="/customer/change-password", status_code=303)
        
    if new_password != confirm_password:
        return templates.TemplateResponse("customer_set_password.html", {
            "request": request, "error": "Passwords do not match"
        })
        
    # Validate Rules
    errors = []
    if len(new_password) < 8:
        errors.append("Password < 8 chars.")
    if not re.search(r"[0-9]", new_password):
        errors.append("No number.")
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", new_password):
        errors.append("No special char.")
        
    if errors:
         return templates.TemplateResponse("customer_set_password.html", {
            "request": request, 
            "error": " ".join(errors)
        })
        
    conn = get_db_connection()
    conn.execute("UPDATE customers SET password = ? WHERE email = ?", (new_password, user["email"]))
    conn.commit()
    # Cleanup OTP just in case
    conn.execute("DELETE FROM otp_codes WHERE email = ?", (user["email"],))
    conn.commit()
    conn.close()
    
    request.session.pop("pending_cust_pass_change", None)
    
    return RedirectResponse(url="/customer/dashboard?success=Password Changed", status_code=303)

@app.get("/customer/profile", response_class=HTMLResponse)
async def customer_profile(request: Request):
    user = require_customer(request)
    conn = get_db_connection()
    # Fetch fresh user data
    db_user = conn.execute("SELECT * FROM customers WHERE email = ?", (user["email"],)).fetchone()
    conn.close()
    
    return templates.TemplateResponse("customer_profile.html", {"request": request, "user": db_user})

@app.post("/customer/profile")
async def update_customer_profile(
    request: Request,
    name: str = Form(...),
    mobile: Optional[str] = Form(None),
    address: Optional[str] = Form(None)
):
    user = require_customer(request)
    conn = get_db_connection()
    
    try:
        conn.execute(
            "UPDATE customers SET name = ?, mobile = ?, address = ? WHERE email = ?", 
            (name, mobile, address, user["email"])
        )
        conn.commit()
        # Update session name
        request.session["user"]["name"] = name
        success = "Profile updated successfully"
    except Exception as e:
        print(f"Error updating profile: {e}")
        success = None
        
    # Fetch fresh user data for re-render
    db_user = conn.execute("SELECT * FROM customers WHERE email = ?", (user["email"],)).fetchone()
    conn.close()
    
    return templates.TemplateResponse("customer_profile.html", {
        "request": request, 
        "user": db_user, 
        "success": success
    })



@app.get("/admin/reports", response_class=HTMLResponse)
async def admin_reports(request: Request):
    require_admin(request)
    return templates.TemplateResponse("admin_reports_dashboard.html", {"request": request})

@app.get("/admin/reports/customers", response_class=HTMLResponse)
async def report_customers(
    request: Request,
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = 1,
    page_size: int = 20
):
    require_admin(request)
    conn = get_db_connection()
    
    query = "SELECT * FROM customers"
    params = []
    where_clauses = []
    
    if search:
        s = f"%{search.lower().strip()}%"
        where_clauses.append("(LOWER(name) LIKE ? OR LOWER(email) LIKE ? OR mobile LIKE ?)")
        params.extend([s, s, s])
    
    if category:
        where_clauses.append("category = ?")
        params.append(category)
        
    if status:
        where_clauses.append("status = ?")
        params.append(status)

    if date_from:
        where_clauses.append("created_at >= ?")
        params.append(f"{date_from} 00:00:00")
        
    if date_to:
        where_clauses.append("created_at <= ?")
        params.append(f"{date_to} 23:59:59")
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
        
    # Count for pagination
    count_query = query.replace("SELECT *", "SELECT COUNT(*)")
    total_count = conn.execute(count_query, tuple(params)).fetchone()[0]
    
    # Sort and Paginate
    query += " ORDER BY created_at DESC, name ASC"
    query += f" LIMIT {page_size} OFFSET {(page - 1) * page_size}"
    
    customers = conn.execute(query, tuple(params)).fetchall()
    
    # Get all categories from MASTER table
    category_list = conn.execute("SELECT name FROM customer_categories WHERE is_active = 1 ORDER BY name ASC").fetchall()
    total_pages = (total_count + page_size - 1) // page_size
    
    return templates.TemplateResponse("admin_report_customers.html", {
        "request": request,
        "customers": customers,
        "search": search or "",
        "category": category or "",
        "status": status or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "category_list": category_list,
        "page": page,
        "total_pages": total_pages,
        "total_count": total_count
    })

@app.post("/admin/reports/customers/update")
async def update_customer_report(request: Request):
    require_admin(request)
    form = await request.form()
    email = form.get("email")
    new_category = form.get("category")
    new_status = form.get("status")
    
    if email:
        conn = get_db_connection()
        conn.execute("UPDATE customers SET category = ?, status = ? WHERE email = ?", (new_category, new_status, email))
        conn.commit()
        conn.close()
        
    return RedirectResponse(url=request.headers.get("referer", "/admin/reports/customers"), status_code=303)

@app.get("/admin/reports/invoices", response_class=HTMLResponse)
async def report_invoices(
    request: Request,
    search: Optional[str] = None,
    customer: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    export: Optional[str] = None
):
    """Invoice Report - Shows only invoices"""
    require_admin(request)
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    conn = get_db_connection()
    
    # Query: Invoices only (no PO join needed)
    query = '''
        SELECT 
            i.id,
            i.invoice_no,
            i.created_at,
            i.payment_mode,
            i.delivery_remarks,
            i.status,
            i.po_id,
            po.total_amount,
            po.amount_received,
            po.display_id,
            c.name as customer_name,
            c.business_name,
            c.category as customer_category
        FROM invoices i
        JOIN purchase_orders po ON i.po_id = po.id
        JOIN customers c ON po.customer_email = c.email
        WHERE po.is_active = 1
    '''
    
    params = []
    where_clauses = []
    
    if search:
        s = f"%{search.lower().strip()}%"
        where_clauses.append("(LOWER(i.invoice_no) LIKE ? OR LOWER(po.display_id) LIKE ? OR LOWER(c.name) LIKE ?)")
        params.extend([s, s, s])
        
    if customer:
        where_clauses.append("c.email = ?")
        params.append(customer)
        
    if status:
        where_clauses.append("LOWER(i.status) = LOWER(?)")
        params.append(status)
        
    if date_from:
        where_clauses.append("i.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
        
    if date_to:
        where_clauses.append("i.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
        
    if where_clauses:
        query += " AND " + " AND ".join(where_clauses)
        
    query += " ORDER BY i.created_at DESC"
    
    invoices = conn.execute(query, tuple(params)).fetchall()
    
    # Convert to dict and Handle Alias Manually if SQL alias fails or for clarity
    invoices_data = []
    for row in invoices:
        d = dict(row)
        # Fallback if alias didn't catch or just to be safe
        if 'display_id' in d and 'po_display_id' not in d:
            d['po_display_id'] = d['display_id']
        invoices_data.append(d)
    
    conn.close()
    
    if export:
        return export_invoices(invoices_data)
    
    return templates.TemplateResponse("admin_report_invoices.html", {
        "request": request,
        "invoices": invoices_data,
        "search": search or "",
        "customer": customer or "",
        "status": status or "",
        "date_from": date_from or "",
        "date_to": date_to or ""
    })

@app.get("/admin/reports/pos", response_class=HTMLResponse)
async def report_pos(
    request: Request,
    search: Optional[str] = None,
    customer: Optional[str] = None,
    status: Optional[str] = None,
    delivery_status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    export: Optional[str] = None
):
    """PO Report - Shows only purchase orders"""
    require_admin(request)
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    
    conn = get_db_connection()
    
    # Query: POs only (no invoice join needed)
    query = '''
        SELECT 
            po.id, 
            po.display_id,
            po.created_at,
            po.total_amount,
            po.status as pr_status,
            po.delivery_status,
            c.name as customer_name,
            c.business_name,
            c.category as customer_category
        FROM purchase_orders po
        JOIN customers c ON po.customer_email = c.email
        WHERE po.is_active = 1
    '''
    
    params = []
    where_clauses = []
    
    if search:
        s = f"%{search.lower().strip()}%"
        where_clauses.append("(LOWER(po.display_id) LIKE ? OR LOWER(c.name) LIKE ?)")
        params.extend([s, s])
        
    if customer:
        where_clauses.append("c.email = ?")
        params.append(customer)
        
    if status:
        where_clauses.append("LOWER(po.status) = LOWER(?)")
        params.append(status)
        
    if delivery_status:
        where_clauses.append("LOWER(po.delivery_status) = LOWER(?)")
        params.append(delivery_status)
        
    if date_from:
        where_clauses.append("po.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
        
    if date_to:
        where_clauses.append("po.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
        
    if where_clauses:
        query += " AND " + " AND ".join(where_clauses)
        
    query += " ORDER BY po.created_at DESC"
    
    pos = conn.execute(query, tuple(params)).fetchall()
    
    # Convert sqlite3.Row to dict
    pos_data = [dict(row) for row in pos]
    
    conn.close()
    
    if export:
        return export_pos(pos_data)
    
    return templates.TemplateResponse("admin_report_pos.html", {
        "request": request,
        "pos": pos,
        "search": search or "",
        "customer": customer or "",
        "status": status or "",
        "delivery_status": delivery_status or "",
        "date_from": date_from or "",
        "date_to": date_to or ""
    })

@app.get("/admin/profile", response_class=HTMLResponse)
async def admin_profile(request: Request):
    require_admin(request)
    
    # Fetch settings
    supplier_name = get_setting("supplier_name", "Hotel Supplier")
    admin_email = get_setting("admin_email", ADMIN_EMAIL)
    otp_sender_email = get_setting("otp_sender_email", "")
    otp_password = get_setting("otp_password", "")
    
    return templates.TemplateResponse("admin_profile.html", {
        "request": request,
        "supplier_name": supplier_name,
        "admin_email": admin_email,
        "otp_sender_email": otp_sender_email,
        "otp_is_configured": bool(otp_password),
        "success": request.query_params.get("success"),
        "error": request.query_params.get("error")
    })

@app.post("/admin/profile")
async def update_admin_profile(
    request: Request,
    supplier_name: str = Form(...),
    admin_email: str = Form(...),
    current_password: str = Form(...),
    new_password: Optional[str] = Form(None),
    otp_sender_email: Optional[str] = Form(None),
    otp_password: Optional[str] = Form(None)
):
    require_admin(request)
    
    # Verify current password
    db_admin_password = get_setting("admin_password", ADMIN_PASSWORD)
    if current_password != db_admin_password:
        return RedirectResponse(url="/admin/profile?error=Incorrect Password", status_code=303)
    
    # Update settings
    set_setting("supplier_name", supplier_name)
    set_setting("admin_email", admin_email)
    
    if otp_sender_email:
        set_setting("otp_sender_email", otp_sender_email)
    
    if otp_password and otp_password.strip():
        set_setting("otp_password", otp_password)
        
    if new_password and new_password.strip():
        set_setting("admin_password", new_password)
        
    return RedirectResponse(url="/admin/profile?success=Profile Updated Successfully", status_code=303)

@app.get("/admin/reports/item_sales", response_class=HTMLResponse)
async def report_item_sales(
    request: Request,
    search: Optional[str] = None,
    item_name: Optional[str] = None,
    customer_name: Optional[str] = None,
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    export: Optional[str] = None
):
    require_admin(request)
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    conn = get_db_connection()
    
    # Base query for POs with customer info
    query = '''
        SELECT po.*, c.name as customer_name, c.business_name, c.category as customer_category
        FROM purchase_orders po
        JOIN customers c ON po.customer_email = c.email
        INNER JOIN invoices i ON po.id = i.po_id
        WHERE po.is_active = 1
    '''
    params = []
    where_clauses = []
    
    if date_from:
        where_clauses.append("po.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
    if date_to:
        where_clauses.append("po.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
        
    if where_clauses:
        query += " AND " + " AND ".join(where_clauses)
    
    query += " ORDER BY po.created_at DESC"
    
    po_rows = conn.execute(query, tuple(params)).fetchall()
    
    # Flatten items and apply remaining filters
    sales_data = []
    
    # Pre-fetch all product names for the dropdown
    all_products = conn.execute("SELECT name FROM products ORDER BY name ASC").fetchall()
    item_name_list = [p['name'] for p in all_products]

    search_lower = search.lower().strip() if search else None
    
    for po in po_rows:
        try:
            items = json.loads(po['items_json'])
            for item in items:
                # Apply filters
                match = True
                if item_name and item['name'] != item_name: match = False
                if customer_name and customer_name.lower() not in po['customer_name'].lower(): match = False
                if category and po['customer_category'] != category: match = False
                
                if search_lower:
                    search_match = (
                        search_lower in item['name'].lower() or
                        search_lower in po['customer_name'].lower()
                    )
                    if not search_match: match = False
                
                if match:
                    sales_data.append({
                        "id": po['id'],
                        "display_id": po['display_id'] or f"PO-{po['id']}",
                        "date": po['created_at'].split(' ')[0],
                        "item_name": item['name'],
                        "qty": item['qty'],
                        "unit": item.get('unit', ''),
                        "customer_name": po['customer_name'],
                        "business_name": po['business_name'],
                        "customer_category": po['customer_category'],
                        "rate": item.get('quoted_rate', 0),
                        "amount": item.get('amount', 0)
                    })
        except:
            continue
            
    # Summary
    total_qty = sum(s['qty'] for s in sales_data)
    total_amount = sum(s['amount'] for s in sales_data)
    
    # Get categories for filter
    categories_rows = conn.execute("SELECT DISTINCT category FROM customers WHERE category IS NOT NULL AND category != ''").fetchall()
    category_list_names = [c[0] for c in categories_rows]
    
    conn.close()
    
    if export:
        return export_item_sales(sales_data)
    
    return templates.TemplateResponse("admin_report_item_sales.html", {
        "request": request,
        "sales": sales_data,
        "total_qty": total_qty,
        "total_amount": total_amount,
        "search": search or "",
        "item_name": item_name or "",
        "customer_name": customer_name or "",
        "category": category or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "item_name_list": item_name_list,
        "category_list": category_list_names
    })

@app.get("/admin/pr_view/{pr_id}", response_class=HTMLResponse)
async def admin_pr_view(request: Request, pr_id: int):
    require_admin(request)
    conn = get_db_connection()
    
    query = '''
        SELECT pr.*, c.name as customer_name, c.business_name
        FROM purchase_requisitions pr
        JOIN customers c ON pr.customer_email = c.email
        WHERE pr.id = ?
    '''
    pr_row = conn.execute(query, (pr_id,)).fetchone()
    
    if not pr_row:
        conn.close()
        return RedirectResponse(url="/admin/reports/prs", status_code=303)
        
    pr = dict(pr_row)
    items = json.loads(pr["items_json"])
    
    # Hydrate items
    hydrated_items = []
    total = 0.0
    for item in items:
        prod = conn.execute("SELECT name, unit FROM products WHERE id = ?", (item['product_id'],)).fetchone()
        item_name = prod['name'] if prod else f"Unknown Product ({item['product_id']})"
        item_unit = prod['unit'] if prod else "-"
        
        quoted_rate = item.get('quoted_rate', 0)
        amt = item.get('amount', 0)
            
        hydrated_items.append({
            "product_id": item['product_id'],
            "name": item_name,
            "unit": item_unit,
            "qty": item['qty'],
            "quoted_rate": quoted_rate,
            "amount": amt
        })
        total += amt
            
    conn.close()
    return templates.TemplateResponse("admin_pr_view.html", {
        "request": request, 
        "pr": pr, 
        "items": hydrated_items,
        "total": total
    })

@app.get("/admin/reports/prs", response_class=HTMLResponse)
async def report_prs(
    request: Request,
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    export: Optional[str] = None
):
    require_admin(request)
    export = str(export).lower() in ('true', '1', 'on', 'yes') if export else False
    conn = get_db_connection()
    
    query = '''
        SELECT pr.*, c.name as customer_name, c.business_name, c.category as customer_category, po.display_id as po_id
        FROM purchase_requisitions pr
        JOIN customers c ON pr.customer_email = c.email
        LEFT JOIN purchase_orders po ON pr.id = po.pr_id AND po.is_active = 1
    '''
    
    params = []
    where_clauses = []
    
    if search:
        s = f"%{search.lower().strip()}%"
        where_clauses.append("(CAST(pr.id AS TEXT) LIKE ? OR LOWER(c.name) LIKE ?)")
        params.extend([s, s])
        
    if category:
        where_clauses.append("c.category = ?")
        params.append(category)
        
    if status:
        status_lower = status.lower().strip()
        if status_lower == 'submitted': where_clauses.append("LOWER(pr.status) = 'pr'")
        elif status_lower == 'quotation sent': where_clauses.append("LOWER(pr.status) = 'qt'")
        elif status_lower in ('invoiced', 'accepted'): where_clauses.append("LOWER(pr.status) IN ('accepted', 'partially_accepted')")
        elif status_lower == 'rejected': where_clauses.append("LOWER(pr.status) = 'rejected'")
        else:
            where_clauses.append("LOWER(pr.status) = ?")
            params.append(status_lower)
            
    if date_from:
        where_clauses.append("pr.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
        
    if date_to:
        where_clauses.append("pr.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
        
    query += " ORDER BY pr.created_at DESC"
    
    prs_rows = conn.execute(query, tuple(params)).fetchall()
    
    prs = []
    for row in prs_rows:
        pr = dict(row)
        try:
            items = json.loads(pr['items_json'])
            total = sum(item.get('amount', 0) for item in items)
            pr['total_amount'] = total
        except:
            pr['total_amount'] = 0
            
        db_status = pr['status']
        if db_status == 'PR': pr['display_status'] = 'Submitted'
        elif db_status == 'QT': pr['display_status'] = 'Quotation Sent'
        elif db_status == 'ACCEPTED': pr['display_status'] = 'Invoiced'
        elif db_status == 'PARTIALLY_ACCEPTED': pr['display_status'] = 'Invoiced (Partial)'
        elif db_status == 'REJECTED': pr['display_status'] = 'Rejected'
        else: pr['display_status'] = db_status
        
        prs.append(pr)
        
    categories = conn.execute("SELECT DISTINCT category FROM customers WHERE category IS NOT NULL AND category != ''").fetchall()
    category_list_names = [c[0] for c in categories]
    
    conn.close()
    
    if export:
        return export_prs(prs)
    
    return templates.TemplateResponse("admin_report_prs.html", {
        "request": request,
        "prs": prs,
        "search": search or "",
        "category": category or "",
        "status": status or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "category_list": category_list_names
    })



@app.get("/admin/reports/outstanding", response_class=HTMLResponse)
async def report_outstanding(request: Request, search: Optional[str] = None):
    require_admin(request)
    conn = get_db_connection()
    
    # Logic: Outstanding = SUM(invoiced_amount) - SUM(invoice_receipt) - SUM(amount_received)
    # ONLY for generated invoices.
    
    # 1. Get all customers
    # 2. Get aggregated outstanding per customer
    
    query = '''
        SELECT 
            c.email,
            c.name,
            c.business_name,
            SUM(po.total_amount) as total_invoiced,
            SUM(po.invoice_receipt) as total_receipts,
            SUM(po.amount_received) as total_received
        FROM customers c
        INNER JOIN purchase_orders po ON c.email = po.customer_email
        INNER JOIN invoices i ON po.id = i.po_id
        WHERE po.is_active = 1
        GROUP BY c.email
    '''
    
    rows = conn.execute(query).fetchall()
    
    outstanding_list = []
    total_receivable = 0.0
    
    search_lower = search.lower().strip() if search else None
    
    for row in rows:
        invoiced = row['total_invoiced'] or 0.0
        receipts = row['total_receipts'] or 0.0
        received = row['total_received'] or 0.0
        
        balance = invoiced - receipts - received
        
        # Only show if balance is non-zero (or strictly positive? User said "Outstanding", usually implies > 0, 
        # but seeing credits is also useful. Let's show all non-zero).
        # User requirement: "all businesses with outstanding amounts"
        
        if abs(balance) > 0.01:
            customer_match = True
            if search_lower:
                match_str = f"{row['name']} {row['business_name']} {row['email']}".lower()
                if search_lower not in match_str:
                    customer_match = False
            
            if customer_match:
                outstanding_list.append({
                    "email": row['email'],
                    "name": row['name'],
                    "business_name": row['business_name'],
                    "balance": balance
                })
                total_receivable += balance

    # Sort by balance descending (highest debt first)
    outstanding_list.sort(key=lambda x: x['balance'], reverse=True)
    
    conn.close()
    
    return templates.TemplateResponse("admin_report_outstanding.html", {
        "request": request,
        "outstanding_list": outstanding_list,
        "total_receivable": total_receivable,
        "search": search or ""
    })

@app.get("/admin/reports/ledger", response_class=HTMLResponse)
async def report_ledger(
    request: Request,
    customer_email: Optional[str] = None,
    customer: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    export: Optional[str] = None
):
    require_admin(request)
    conn = get_db_connection()
    
    # Get all customers for the dropdown
    all_customers = conn.execute("SELECT email, name, business_name FROM customers ORDER BY name ASC").fetchall()
    
    ledger_entries = []
    selected_customer = None
    email_to_use = customer_email or customer
    
    if email_to_use:
        selected_customer = conn.execute("SELECT * FROM customers WHERE email = ?", (email_to_use,)).fetchone()
        
        query = '''
            SELECT po.*, c.category as customer_category, i.invoice_no
            FROM purchase_orders po
            JOIN customers c ON po.customer_email = c.email
            INNER JOIN invoices i ON po.id = i.po_id
            WHERE po.customer_email = ? AND po.is_active = 1
        '''
        params = [email_to_use]
        query += " ORDER BY po.created_at ASC"
        
        all_pos = conn.execute(query, tuple(params)).fetchall()
        
        running_balance = 0.0
        for po in all_pos:
            opening_bal = running_balance
            inv_amt = po['total_amount']
            inv_rec = po['invoice_receipt'] or 0.0
            amt_rec = po['amount_received'] or 0.0
            closing_bal = opening_bal + inv_amt - inv_rec - amt_rec
            
            entry = {
                "id": po['id'],
                "display_id": po['invoice_no'] or po['display_id'] or f"PO-{po['id']}",
                "date": po['created_at'].split(' ')[0],
                "category": po['customer_category'],
                "opening_bal": opening_bal,
                "inv_amt": inv_amt,
                "inv_rec": inv_rec,
                "amt_rec": amt_rec,
                "closing_bal": closing_bal
            }
            
            show = True
            if date_from and entry['date'] < date_from: show = False
            if date_to and entry['date'] > date_to: show = False
            if show: ledger_entries.append(entry)
            running_balance = closing_bal
            
    conn.close()
    
    if export:
        return export_ledger(selected_customer['name'] if selected_customer else "Unknown", ledger_entries)
    
    return templates.TemplateResponse("admin_report_ledger.html", {
        "request": request,
        "customers": all_customers,
        "selected_email": email_to_use or "",
        "selected_customer": selected_customer,
        "ledger": ledger_entries,
        "date_from": date_from or "",
        "date_to": date_to or ""
    })

@app.post("/admin/reports/ledger/update")
async def update_ledger_payment(request: Request):
    require_admin(request)
    form = await request.form()
    po_id = form.get("po_id")
    try:
        amt_rec = float(form.get("amount_received", 0))
    except:
        amt_rec = 0.0
        
    if po_id:
        conn = get_db_connection()
        # Fetch current invoice total and previous balance to validate
        po = conn.execute("SELECT total_amount, invoice_receipt, customer_email, created_at, delivery_status, tracking_status FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
        if po:
            # Re-calculate opening balance to properly validate max amount
            all_pos = conn.execute("SELECT total_amount, amount_received, invoice_receipt, delivery_status FROM purchase_orders WHERE customer_email = ? AND created_at < ? AND is_active = 1 ORDER BY created_at ASC", 
                                   (po['customer_email'], po['created_at'])).fetchall()
            opening_bal = sum((p['total_amount'] - (p['invoice_receipt'] or 0.0) - (p['amount_received'] or 0.0)) for p in all_pos)
            
            # Validation
            if amt_rec < 0:
                amt_rec = 0.0
            
            # Allow any amount (including overpayment/credit)
            # if amt_rec > max_allowed:
            #     amt_rec = max_allowed
                
            conn.execute("UPDATE purchase_orders SET amount_received = ? WHERE id = ?", (amt_rec, po_id))
            conn.commit()
        conn.close()
        
    return RedirectResponse(url=request.headers.get("referer", "/admin/reports/ledger"), status_code=303)

# --- Admin: Customer Category Management ---

@app.get("/admin/customer_categories", response_class=HTMLResponse)
async def admin_customer_categories(request: Request):
    with open("debug_route.txt", "w") as f:
        f.write("Route HIT\n")
    try:
        # require_admin(request) # Commented out to isolate
        conn = get_db_connection()
        
        # Get all customer categories with their delivery settings
        categories = conn.execute("""
            SELECT name, delivery_days, route_name
            FROM customer_categories
            ORDER BY name
        """).fetchall()
        
        # Get customer count for each category
        category_data = []
        for cat in categories:
            customer_count = conn.execute(
                "SELECT COUNT(*) FROM customers WHERE category = ?", 
                (cat['name'],)
            ).fetchone()[0]
            
            delivery_days = []
            if cat['delivery_days']:
                try:
                    delivery_days = json.loads(cat['delivery_days'])
                except:
                    delivery_days = []
            
            category_data.append({
                'name': cat['name'],
                'route_name': cat['route_name'] or '',
                'delivery_days': delivery_days,
                'customer_count': customer_count
            })
        
        conn.close()
        
        return templates.TemplateResponse("admin_customer_categories.html", {
            "request": request,
            "categories": category_data
        })
    except Exception as e:
        import traceback
        import sys
        
        # Print to stderr for visibility in console
        print(f"\n[CRITICAL ERROR] in admin_customer_categories: {e}\n", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        
        # Also try to log to file as backup with absolute path
        try:
            with open("C:/Users/usery/.gemini/antigravity/scratch/hotel_app/debug_error_v2.txt", "w") as f:
                f.write(str(e) + "\n")
                traceback.print_exc(file=f)
        except Exception as file_err:
             print(f"Could not write log file: {file_err}", file=sys.stderr)
             
        raise e

@app.post("/admin/customer_categories/update")
async def update_customer_category(request: Request):
    require_admin(request)
    form = await request.form()
    category_name = form.get("category_name")
    route_name = form.get("route_name", "")
    
    # Get selected delivery days from form (multi-select)
    delivery_days = []
    for day in ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']:
        if form.get(f"day_{day}"):
            delivery_days.append(day)
    
    if category_name:
        conn = get_db_connection()
        conn.execute("""
            UPDATE customer_categories
            SET delivery_days = ?, route_name = ?
            WHERE name = ?
        """, (json.dumps(delivery_days), route_name, category_name))
        conn.commit()
        conn.close()
    
    return RedirectResponse(url="/admin/customer-categories", status_code=303)

# --- Export Helpers ---

def export_prs(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "PRS"
    
    headers = ["PR No", "Date", "Customer Name", "Customer Category", "Total Amount", "Status", "Items Count"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row.get('id'), 
            row.get('created_at'), 
            row.get('customer_name'), 
            row.get('customer_category'),
            row.get('total_amount'), 
            row.get('display_status'), 
            len(json.loads(row.get('items_json', '[]')))
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"PRS_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_pos(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"
    
    headers = ["PO No", "Date", "Customer Name", "Customer Category", "Total Amount", "Status", "Delivery Status"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row.get('display_id') or f"PO-{row.get('id')}",
            row.get('created_at'),
            row.get('customer_name'),
            row.get('customer_category'),
            row.get('total_amount'),
            row.get('pr_status'),
            row.get('delivery_status')
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"POS_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_invoices(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    headers = ["Invoice No", "Date", "PO No", "Customer Name", "Amount", "Payment Mode", "Status"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row.get('invoice_no'),
            row.get('created_at'),
            row.get('po_display_id'),
            row.get('customer_name'),
            row.get('total_amount'),
            row.get('payment_mode'),
            row.get('status')
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Invoices_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_ledger(customer_name, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    
    ws.append([f"Customer Ledger: {customer_name}"])
    ws.append([])
    
    headers = ["Date", "PO Ref", "Category", "Opening Balance", "Invoice Amount", "Adjustments", "Paid Amount", "Closing Balance"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row.get('date'),
            row.get('display_id'),
            row.get('category'),
            row.get('opening_bal'),
            row.get('inv_amt'),
            row.get('inv_rec'),
            row.get('amt_rec'),
            row.get('closing_bal')
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Ledger_{customer_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

def export_item_sales(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Item Wise Sales"
    
    headers = ["Date", "PO Ref", "Item Name", "Unit", "Qty", "Rate", "Amount", "Customer Name", "Category"]
    ws.append(headers)
    
    for row in data:
        ws.append([
            row.get('date'),
            row.get('display_id'),
            row.get('item_name'),
            row.get('unit'),
            row.get('qty'),
            row.get('rate'),
            row.get('amount'),
            row.get('customer_name'),
            row.get('customer_category')
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Item_Sales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/customer/create_pr", response_class=HTMLResponse)
async def create_pr(request: Request):
    user = require_customer(request)
    conn = get_db_connection()
    # Join with categories
    query = '''
        SELECT p.*, c.name as category_name 
        FROM products p
        LEFT JOIN categories c ON p.category_id = c.id
        ORDER BY c.name, p.name
    '''
    products = conn.execute(query).fetchall()
    categories = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    # Fetch current draft if any
    draft_items = {}
    draft = conn.execute("SELECT items_json FROM purchase_requisitions WHERE customer_email = ? AND status = 'DRAFT' ORDER BY created_at DESC LIMIT 1", (user['email'],)).fetchone()
    if draft:
        try:
            items = json.loads(draft['items_json'])
            draft_items = {item['product_id']: item['qty'] for item in items}
        except: pass

    conn.close()
    return templates.TemplateResponse("create_pr.html", {
        "request": request, 
        "products": products,
        "categories": categories,
        "draft_items": draft_items
    })

@app.post("/customer/create_pr")
async def submit_pr(request: Request):
    user = require_customer(request)
    form_data = await request.form()
    
    # Parse quantities
    # Form keys: qty_{product_id}. If > 0, include in PR.
    items = []
    for key, value in form_data.items():
        if key.startswith("qty_") and value:
            try:
                qty = float(value)
                if qty > 0:
                    product_id = int(key.split("_")[1])
                    items.append({"product_id": product_id, "qty": qty})
            except ValueError:
                continue
                
    if not items:
        # Should probably show error, but keeping it simple
        return RedirectResponse(url="/customer/create_pr", status_code=303)
        
    conn = get_db_connection()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Check for existing draft to update instead of creating new one
    existing_draft = conn.execute("SELECT id FROM purchase_requisitions WHERE customer_email = ? AND status = 'DRAFT' ORDER BY created_at DESC LIMIT 1", (user['email'],)).fetchone()
    
    cur = conn.cursor()
    if existing_draft:
        pr_id = existing_draft['id']
        if DATABASE_URL:
            conn.execute("UPDATE purchase_requisitions SET items_json = %s, created_at = %s WHERE id = %s", (json.dumps(items), created_at, pr_id))
        else:
            conn.execute("UPDATE purchase_requisitions SET items_json = ?, created_at = ? WHERE id = ?", (json.dumps(items), created_at, pr_id))
    else:
        # Insert PR with DRAFT status
        if DATABASE_URL:
            cur.execute(
                "INSERT INTO purchase_requisitions (customer_email, status, created_at, items_json) VALUES (%s, %s, %s, %s)",
                (user['email'], 'DRAFT', created_at, json.dumps(items))
            )
            # Fetch the generated ID for PostgreSQL
            cur.execute("SELECT LASTVAL()")
            pr_id = cur.fetchone()[0]
        else:
            cur.execute(
                "INSERT INTO purchase_requisitions (customer_email, status, created_at, items_json) VALUES (?, ?, ?, ?)",
                (user['email'], 'DRAFT', created_at, json.dumps(items))
            )
            pr_id = cur.lastrowid
        
    conn.commit()
    conn.close()

    # Redirect to Draft Review Page
    return RedirectResponse(url=f"/customer/draft_pr/{pr_id}", status_code=303)

@app.get("/customer/cart", response_class=HTMLResponse)
async def customer_cart(request: Request):
    user = require_customer(request)
    conn = get_db_connection()
    # Find most recent draft
    draft = conn.execute("SELECT id FROM purchase_requisitions WHERE customer_email = ? AND status = 'DRAFT' ORDER BY created_at DESC LIMIT 1", (user['email'],)).fetchone()
    
    if draft:
        conn.close()
        return RedirectResponse(url=f"/customer/draft_pr/{draft['id']}", status_code=303)
    
    # No draft found - fetch last 3 submitted PRs for reuse section
    if DATABASE_URL:
        recent_prs = conn.execute(
            "SELECT id, created_at, items_json FROM purchase_requisitions WHERE customer_email = %s AND status != 'DRAFT' ORDER BY created_at DESC LIMIT 3",
            (user['email'],)
        ).fetchall()
    else:
        recent_prs = conn.execute(
            "SELECT id, created_at, items_json FROM purchase_requisitions WHERE customer_email = ? AND status != 'DRAFT' ORDER BY created_at DESC LIMIT 3",
            (user['email'],)
        ).fetchall()
    
    recent_prs_list = []
    for rpr in recent_prs:
        try:
            r_items = json.loads(rpr['items_json'])
            recent_prs_list.append({
                "id": rpr['id'],
                "date": rpr['created_at'].split()[0], # YYYY-MM-DD
                "count": len(r_items)
            })
        except: continue
    conn.close()

    return templates.TemplateResponse("customer_draft_review.html", {
        "request": request,
        "draft": None,
        "items": [],
        "total_items": 0,
        "recent_prs": recent_prs_list
    })

@app.get("/customer/draft_pr/{pr_id}", response_class=HTMLResponse)
async def view_draft_pr(request: Request, pr_id: int):
    user = require_customer(request)
    conn = get_db_connection()
    
    # Fetch Draft
    if DATABASE_URL:
        pr = conn.execute(
            "SELECT * FROM purchase_requisitions WHERE id = %s AND customer_email = %s AND status = 'DRAFT'", 
            (pr_id, user['email'])
        ).fetchone()
    else:
        pr = conn.execute(
            "SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ? AND status = 'DRAFT'", 
            (pr_id, user['email'])
        ).fetchone()
        
    if not pr:
        conn.close()
        return RedirectResponse(url="/customer/dashboard?error=Draft not found or already submitted", status_code=303)
        
    items = json.loads(pr['items_json'])
    
    # Enrich items with product details
    enriched_items = []
    if items:
        product_ids = [item['product_id'] for item in items]
        placeholders = ','.join(['%s' if DATABASE_URL else '?'] * len(product_ids))
        
        query = f"SELECT id, name, unit, image_url FROM products WHERE id IN ({placeholders})"
        products_db = conn.execute(query, tuple(product_ids)).fetchall()
        product_map = {p['id']: p for p in products_db}
        
        for item in items:
            p = product_map.get(item['product_id'])
            if p:
                enriched_items.append({
                    "product_id": item['product_id'],
                    "qty": item['qty'],
                    "name": p['name'],
                    "unit": p['unit'],
                    "image_url": p['image_url']
                })
    
    # Fetch Last 3 Submitted PRs for Reuse
    if DATABASE_URL:
        recent_prs = conn.execute(
            "SELECT id, created_at, items_json FROM purchase_requisitions WHERE customer_email = %s AND status != 'DRAFT' ORDER BY created_at DESC LIMIT 3",
            (user['email'],)
        ).fetchall()
    else:
        recent_prs = conn.execute(
            "SELECT id, created_at, items_json FROM purchase_requisitions WHERE customer_email = ? AND status != 'DRAFT' ORDER BY created_at DESC LIMIT 3",
            (user['email'],)
        ).fetchall()
    
    recent_prs_list = []
    for rpr in recent_prs:
        try:
            r_items = json.loads(rpr['items_json'])
            recent_prs_list.append({
                "id": rpr['id'],
                "date": rpr['created_at'].split()[0], # YYYY-MM-DD
                "count": len(r_items)
            })
        except: continue

    conn.close()
    
    return templates.TemplateResponse("customer_draft_review.html", {
        "request": request,
        "draft": pr,
        "items": enriched_items,
        "total_items": len(enriched_items),
        "recent_prs": recent_prs_list
    })

@app.post("/customer/draft_pr/{pr_id}/reuse")
async def reuse_items_in_draft(request: Request, pr_id: int):
    user = require_customer(request)
    form = await request.form()
    old_pr_id = int(form.get("old_pr_id"))

    conn = get_db_connection()
    
    # 1. Fetch old PR
    old_pr = conn.execute("SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ?", (old_pr_id, user['email'])).fetchone()
    if not old_pr:
        conn.close()
        return RedirectResponse(url=f"/customer/draft_pr/{pr_id}?error=Old PR not found", status_code=303)
    
    old_items = json.loads(old_pr['items_json'])
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 2. Handle Existing Draft vs New Draft
    if pr_id == 0:
        # Create new DRAFT PR
        cur = conn.cursor()
        if DATABASE_URL:
            cur.execute(
                "INSERT INTO purchase_requisitions (customer_email, status, created_at, items_json) VALUES (%s, %s, %s, %s)",
                (user['email'], 'DRAFT', created_at, old_pr['items_json'])
            )
            cur.execute("SELECT LASTVAL()")
            new_pr_id = cur.fetchone()[0]
        else:
            cur.execute(
                "INSERT INTO purchase_requisitions (customer_email, status, created_at, items_json) VALUES (?, ?, ?, ?)",
                (user['email'], 'DRAFT', created_at, old_pr['items_json'])
            )
            new_pr_id = cur.lastrowid
        conn.commit()
        conn.close()
        return RedirectResponse(url=f"/customer/draft_pr/{new_pr_id}?success=Order started from PR #{old_pr_id}", status_code=303)
    
    # Fetch current draft
    pr = conn.execute("SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ? AND status = 'DRAFT'", (pr_id, user['email'])).fetchone()
    if not pr:
        conn.close()
        return RedirectResponse(url="/customer/dashboard?error=Draft not found", status_code=303)
    
    current_items = json.loads(pr['items_json'])
    
    # Merge items: if exists in current, update; else add
    item_map = {item['product_id']: item for item in current_items}
    for o_item in old_items:
        if o_item['product_id'] in item_map:
            item_map[o_item['product_id']]['qty'] = o_item['qty'] 
        else:
            item_map[o_item['product_id']] = o_item
    
    new_items_json = json.dumps(list(item_map.values()))
    
    if DATABASE_URL:
        conn.execute("UPDATE purchase_requisitions SET items_json = %s WHERE id = %s", (new_items_json, pr_id))
    else:
        conn.execute("UPDATE purchase_requisitions SET items_json = ? WHERE id = ?", (new_items_json, pr_id))
    
    conn.commit()
    conn.close()
    return RedirectResponse(url=f"/customer/draft_pr/{pr_id}?success=Items added from PR #{old_pr_id}", status_code=303)

@app.post("/customer/draft_pr/{pr_id}/update")
async def update_draft_pr(request: Request, pr_id: int):
    user = require_customer(request)
    form = await request.form()
    action = form.get("action")
    product_id = int(form.get("product_id"))
    qty = float(form.get("qty", 0))
    
    conn = get_db_connection()
    pr = conn.execute(
        "SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ? AND status = 'DRAFT'",
        (pr_id, user['email'])
    ).fetchone()
    
    if not pr:
        conn.close()
        return RedirectResponse(url="/customer/dashboard?error=Draft not found", status_code=303)
        
    items = json.loads(pr['items_json'])
    new_items = []
    
    if action == "remove":
        new_items = [i for i in items if i['product_id'] != product_id]
    elif action == "update":
        for i in items:
            if i['product_id'] == product_id:
                if qty > 0:
                    i['qty'] = qty
                    new_items.append(i)
                # If qty <= 0, it's effectively removed
            else:
                new_items.append(i)
                
    # Update DB
    if DATABASE_URL:
        conn.execute("UPDATE purchase_requisitions SET items_json = %s WHERE id = %s", (json.dumps(new_items), pr_id))
    else:
        conn.execute("UPDATE purchase_requisitions SET items_json = ? WHERE id = ?", (json.dumps(new_items), pr_id))
        
    conn.commit()
    conn.close()
    
    return RedirectResponse(url=f"/customer/draft_pr/{pr_id}", status_code=303)

@app.post("/customer/draft_pr/{pr_id}/delete")
async def delete_draft_pr(request: Request, pr_id: int):
    user = require_customer(request)
    conn = get_db_connection()
    
    if DATABASE_URL:
        conn.execute("DELETE FROM purchase_requisitions WHERE id = %s AND customer_email = %s AND status = 'DRAFT'", (pr_id, user['email']))
    else:
        conn.execute("DELETE FROM purchase_requisitions WHERE id = ? AND customer_email = ? AND status = 'DRAFT'", (pr_id, user['email']))
        
    conn.commit()
    conn.close()
    return RedirectResponse(url="/customer/dashboard?success=Draft deleted", status_code=303)

@app.post("/customer/draft_pr/{pr_id}/submit")
async def submit_draft_pr(request: Request, pr_id: int):
    user = require_customer(request)
    conn = get_db_connection()
    
    # Verify Draft
    pr = conn.execute(
        "SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ? AND status = 'DRAFT'", 
        (pr_id, user['email'])
    ).fetchone()
    
    if not pr:
        conn.close()
        return RedirectResponse(url="/customer/dashboard?error=Draft not found or already submitted", status_code=303)
        
    # Update Status to PR
    submitted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # ── [NEW WORKFLOW ADDITION] ── Read workflow_type from the submitted form
    # Defaults to QUOTATION_APPROVAL to preserve backward compatibility
    form_data_submit = await request.form()
    workflow_type = form_data_submit.get("workflow_type", "QUOTATION_APPROVAL").strip()
    if workflow_type not in ("QUOTATION_APPROVAL", "ADMIN_PRICE_REVIEW"):
        workflow_type = "QUOTATION_APPROVAL"  # Sanitize unknown values

    # Determine new PR status based on workflow type
    # ADMIN_PRICE_REVIEW → PENDING_PRICE_REVIEW (skips quotation entirely)
    # QUOTATION_APPROVAL → PR (existing flow unchanged)
    new_pr_status = "PENDING_PRICE_REVIEW" if workflow_type == "ADMIN_PRICE_REVIEW" else "PR"

    if DATABASE_URL:
        conn.execute(
            "UPDATE purchase_requisitions SET status = %s, created_at = %s, workflow_type = %s WHERE id = %s",
            (new_pr_status, submitted_at, workflow_type, pr_id)
        )
    else:
        conn.execute(
            "UPDATE purchase_requisitions SET status = ?, created_at = ?, workflow_type = ? WHERE id = ?",
            (new_pr_status, submitted_at, workflow_type, pr_id)
        )
        
    conn.commit()
    
    # --- Email Notification Logic (Same as before) ---
    try:
        from app.notifications import send_email_notification, dispatch_pending_notifications

        supplier_email = get_setting("admin_email", ADMIN_EMAIL)
        customer_row = conn.execute("SELECT name, full_name, business_name FROM customers WHERE email = ?", (user['email'],)).fetchone()
        customer_name = None
        if customer_row:
            customer_name = customer_row.get('business_name') or customer_row.get('full_name') or customer_row.get('name')
        if not customer_name:
            customer_name = user.get('name') or user.get('email')

        # ENHANCED EMAIL TEMPLATE WITH CUSTOMER INFO, PR NUMBER, AND DIRECT LINK
        subject = f"🔔 New Purchase Request from {customer_name or 'Customer'}"
        body = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
    <div style="max-width: 600px; margin: 20px auto; padding: 20px; border: 1px solid #ddd; border-radius: 8px; background-color: #ffffff;">
        
        <!-- Header -->
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #2c3e50; margin: 0; font-size: 24px;">🛒 New Purchase Request</h1>
            <p style="color: #7f8c8d; margin: 5px 0 0 0; font-size: 14px;">PR #{pr_id} - Immediate Action Required</p>
        </div>
        
        <!-- Customer Information -->
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #3498db;">
            <h3 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 16px;">📋 Customer Details</h3>
            <p style="margin: 8px 0;"><strong>Customer Name:</strong> {customer_name or 'N/A'}</p>
            <p style="margin: 8px 0;"><strong>Customer Email:</strong> {user['email']}</p>
            <p style="margin: 8px 0;"><strong>PR Number:</strong> #{pr_id}</p>
        </div>
        
        <!-- Action Required -->
        <div style="background-color: #fff3cd; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #ffc107;">
            <h3 style="color: #856404; margin: 0 0 15px 0; font-size: 16px;">⚡ Action Required</h3>
            <p style="margin: 0 0 15px 0;">A new purchase request has been created and requires your immediate attention. Please review the details and submit your competitive quotation.</p>
            
            <ul style="margin: 0; padding-left: 20px;">
                <li style="margin: 8px 0;">Review purchase request details</li>
                <li style="margin: 8px 0;">Submit your competitive quotation</li>
            </ul>
        </div>
        
        <!-- CTA Button -->
        <div style="text-align: center; margin: 30px 0;">
            <a href="{APP_BASE_URL}/admin/login" style="display: inline-block; background-color: #007bff; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 16px;">
                🚀 View Purchase Request
            </a>
        </div>
        
        <!-- Footer -->
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; text-align: center;">
            <p style="color: #6c757d; margin: 5px 0; font-size: 14px;">Prompt response is appreciated to ensure timely processing.</p>
            <p style="color: #6c757d; margin: 5px 0; font-size: 12px;">This is an automated message. Please do not reply to this email.</p>
        </div>
        
    </div>
</body>
</html>
        """.strip()

        notification_id = send_email_notification(
            event_type="PR_CREATED",
            ref_id=int(pr_id),
            recipient_role="supplier",
            recipient_email=supplier_email,
            subject=subject,
            message=body,
        )
        dispatch_pending_notifications()
    except Exception as e:
        print(f"[WARN] PR_CREATED notification failed: {e}")

    conn.close()
    
    return templates.TemplateResponse("pr_submitted.html", {"request": request})

# Removed duplicate customer/quotations route

@app.get("/customer/quotation/{pr_id}", response_class=HTMLResponse)
async def view_quotation(request: Request, pr_id: int):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user.get("role") != "customer":
        if user.get("role") == "admin":
             return RedirectResponse(url=f"/admin/quotation/{pr_id}", status_code=303)
        return JSONResponse({"detail": "Not authorized"}, status_code=403)

    conn = get_db_connection()
    pr = conn.execute("SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ?", 
                      (pr_id, user['email'])).fetchone()
    
    if not pr:
        conn.close()
        return RedirectResponse(url="/customer/quotations", status_code=303)
    
    items = json.loads(pr['items_json'])
    total = sum([item.get('amount', 0) for item in items])
    
    # Fetch active PO with delivery tracking data
    po_row = conn.execute("""
        SELECT id, delivery_status, tracking_status,
               expected_delivery_date, delivery_stage,
               order_placed_at, packaged_at, shipped_at, 
               out_for_delivery_at, delivered_at
        FROM purchase_orders 
        WHERE pr_id = ? AND is_active = 1
    """, (pr_id,)).fetchone()
    
    # Convert Row to dict to support .get() method (needed for SQLite)
    po = dict(po_row) if po_row else None
    
    po_id = po['id'] if po else None
    delivery_data = None
    
    # Prepare delivery tracking data if PO exists and order is accepted
    if po and pr['status'] in ['ACCEPTED', 'PARTIALLY_ACCEPTED']:
        current_stage = po.get('delivery_stage', 'ORDER_PLACED')
        delivery_data = {
            'expected_delivery_date': po.get('expected_delivery_date'),
            'current_stage': current_stage,
            'stages': {
                'ORDER_PLACED': {
                    'label': 'Order Placed',
                    'timestamp': po.get('order_placed_at'),
                    'completed': bool(po.get('order_placed_at')),
                    'active': current_stage == 'ORDER_PLACED'
                },
                'SCHEDULED': {
                    'label': 'Scheduled',
                    'timestamp': po.get('expected_delivery_date'),
                    'completed': bool(po.get('expected_delivery_date')),
                    'active': current_stage == 'SCHEDULED'
                },
                'PACKED': {
                    'label': 'Packed',
                    'timestamp': po.get('packaged_at'),
                    'completed': bool(po.get('packaged_at')),
                    'active': current_stage == 'PACKED'
                },
                'OUT_FOR_DELIVERY': {
                    'label': 'Out for Delivery',
                    'timestamp': po.get('out_for_delivery_at'),
                    'completed': bool(po.get('out_for_delivery_at')),
                    'active': current_stage == 'OUT_FOR_DELIVERY'
                },
                'DELIVERED': {
                    'label': 'Delivered',
                    'timestamp': po.get('delivered_at'),
                    'completed': bool(po.get('delivered_at')),
                    'active': current_stage == 'DELIVERED'
                }
            }
        }
    
    conn.close()
    is_expired = is_quotation_expired(pr['created_at']) if pr['status'] == 'QT' else False
    return templates.TemplateResponse("customer_quotation_view.html", 
                                      {"request": request, "pr": pr, "items": items, "total": total, 
                                       "po_id": po_id, "is_expired": is_expired, "delivery_data": delivery_data})

@app.post("/customer/quotation/{pr_id}/action")
async def quotation_action(request: Request, pr_id: int):
    try:
        print(f"[DEBUG] Quotation action called for PR ID: {pr_id}")
        user = require_customer(request)
        form_data = await request.form()
        print(f"[DEBUG] Form data received: {dict(form_data)}")
        
        conn = get_db_connection()
        pr = conn.execute("SELECT * FROM purchase_requisitions WHERE id = ? AND customer_email = ?", 
                          (pr_id, user['email'])).fetchone()
        
        if not pr:
            print(f"[DEBUG] Quotation not found or access denied for PR ID: {pr_id}, user: {user['email']}")
            conn.close()
            return RedirectResponse(url="/customer/quotations", status_code=303)
        
        print(f"[DEBUG] Quotation found - ID: {pr['id']}, Status: {pr['status']}")
        
        if pr['status'] != 'QT' or is_quotation_expired(pr['created_at']):
            print(f"[DEBUG] Quotation status check failed - Status: {pr['status']}, Expired: {is_quotation_expired(pr['created_at'])}")
            conn.close()
            return RedirectResponse(url="/customer/quotations", status_code=303)
        
        items = json.loads(pr['items_json'])
        
        # Check if items are already locked
        if items and items[0].get('is_locked', False):
            conn.close()
            return RedirectResponse(url="/customer/quotations", status_code=303)
        
        # Process item-level decisions
        # Form keys: item_{product_id}_action = "accept" or "reject"
        decision_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        accepted_count = 0
        rejected_count = 0
        
        for item in items:
            pid = item['product_id']
            action_key = f"item_{pid}_action"
            
            if action_key in form_data:
                action = form_data[action_key]
                if action == "accept":
                    item['item_status'] = "ACCEPTED"
                    accepted_count += 1
                elif action == "reject":
                    item['item_status'] = "REJECTED"
                    rejected_count += 1
                
                item['decision_timestamp'] = decision_timestamp
                item['is_locked'] = True
        
        # Validate: at least one item must be selected
        if accepted_count == 0 and rejected_count == 0:
            conn.close()
            return RedirectResponse(url=f"/customer/quotation/{pr_id}?error=no_selection", status_code=303)
        
        # Determine overall PR status
        if accepted_count > 0 and rejected_count == 0:
            new_status = 'ACCEPTED'
        elif rejected_count > 0 and accepted_count == 0:
            new_status = 'REJECTED' # Rejection is still rejection, but 'Cancelled' might be for POs.
        else:
            new_status = 'PARTIALLY_ACCEPTED'
        
        # Update PR with new status and item decisions
        conn.execute("UPDATE purchase_requisitions SET status = ?, items_json = ? WHERE id = ?", 
                     (new_status, json.dumps(items), pr_id))
        
        if accepted_count > 0:
            # Fetch customer info for snapshots
            customer_row = conn.execute("SELECT name, business_name, address, category, email, mobile FROM customers WHERE email = ?", (user['email'],)).fetchone()
            
            # Create Purchase Order
            accepted_items = [i for i in items if i.get('item_status') == 'ACCEPTED']
            total_amount = sum(float(i.get('amount', 0)) for i in accepted_items)
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Snapshots
            cust_name = customer_row['name'] if customer_row else user['name']
            cust_biz = customer_row['business_name'] if customer_row else ""
            cust_addr = customer_row['address'] if customer_row else ""
            cust_cat = customer_row['category'] if customer_row else ""
            cust_email = customer_row['email'] if customer_row else user['email']
            cust_mobile = customer_row['mobile'] if customer_row else ""

            # Insert PO
            # Note: We must handle both SQLite and PostgreSQL syntax differences via the wrapper if needed, 
            # but standard placeholders usually work if the wrapper handles it.
            # Our wrapper does ? -> %s conversion.
            
            cursor = conn.execute("""
                INSERT INTO purchase_orders (
                    pr_id, customer_email, created_at, total_amount, items_json, 
                    amount_received, invoice_source, status, delivery_status,
                    customer_name_snapshot, business_name_snapshot, address_snapshot, 
                    customer_category_snapshot, customer_email_snapshot, customer_mobile_snapshot,
                    is_active
                ) VALUES (?, ?, ?, ?, ?, 0, 'PR', ?, 'OPEN', ?, ?, ?, ?, ?, ?, 1)
            """, (
                pr_id, user['email'], current_time, total_amount, json.dumps(accepted_items), new_status,
                cust_name, cust_biz, cust_addr, cust_cat, cust_email, cust_mobile
            ))
            
            # Commit the transaction!
            conn.commit()
            
            # ===== DELIVERY TRACKING INTEGRATION =====
            # Get PO ID from cursor
            po_id = cursor.lastrowid if hasattr(cursor, 'lastrowid') else cursor.fetchone()[0]
            
            # Get customer category for delivery route
            delivery_info = conn.execute("""
                SELECT cc.name as category_name, cc.delivery_days, cc.route_name
                FROM customer_categories cc
                WHERE cc.name = ?
            """, (cust_cat,)).fetchone()
            
            delivery_days = []
            if delivery_info and delivery_info['delivery_days']:
                try:
                    delivery_days = json.loads(delivery_info['delivery_days'])
                except:
                    delivery_days = []
            
            # Calculate expected delivery date
            po_created_datetime = datetime.strptime(current_time, '%Y-%m-%d %H:%M:%S')
            expected_delivery_date = calculate_next_delivery_date(po_created_datetime, delivery_days)
            
            # Update PO with delivery tracking fields
            conn.execute("""
                UPDATE purchase_orders
                SET expected_delivery_date = ?,
                    delivery_stage = 'ORDER_PLACED',
                    order_placed_at = ?,
                    customer_category_snapshot = ?
                WHERE id = ?
            """, (expected_delivery_date, current_time, delivery_info['category_name'] if delivery_info else cust_cat, po_id))
            
            conn.commit()
            # ===== END DELIVERY TRACKING =====
            
            from app.notifications import send_email_notification, dispatch_pending_notifications
            from app.notifications import send_email_notification, dispatch_pending_notifications

            supplier_email = get_setting("admin_email", ADMIN_EMAIL)
            
            # Get customer name for email
            customer_row = conn.execute("SELECT name, full_name, business_name FROM customers WHERE email = ?", (user['email'],)).fetchone()
            customer_name = None
            if customer_row:
                customer_name = customer_row['business_name'] or customer_row['full_name'] or customer_row['name']
            if not customer_name:
                customer_name = user['name'] or user['email']
            
            # ENHANCED EMAIL TEMPLATES WITH CUSTOMER INFO, PR NUMBER, AND DIRECT LINK
            if new_status == 'ACCEPTED':
                subject = f"✅ Quotation Accepted by {customer_name}"
                message = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
    <div style="max-width: 600px; margin: 20px auto; padding: 20px; border: 1px solid #ddd; border-radius: 8px; background-color: #ffffff;">
        
        <!-- Header -->
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #27ae60; margin: 0; font-size: 24px;">✅ Quotation Accepted</h1>
            <p style="color: #7f8c8d; margin: 5px 0 0 0; font-size: 14px;">PR #{pr_id} - Congratulations!</p>
        </div>
        
        <!-- Customer Information -->
        <div style="background-color: #d4edda; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #28a745;">
            <h3 style="color: #155724; margin: 0 0 15px 0; font-size: 16px;">📋 Acceptance Details</h3>
            <p style="margin: 8px 0;"><strong>Customer Name:</strong> {customer_name or 'N/A'}</p>
            <p style="margin: 8px 0;"><strong>PR Number:</strong> #{pr_id}</p>
            <p style="margin: 8px 0;"><strong>Status:</strong> <span style="color: #155724; font-weight: bold;">Accepted</span></p>
        </div>
        
        <!-- Action Required -->
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #007bff;">
            <h3 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 16px;">🎉 Congratulations!</h3>
            <p style="margin: 0 0 15px 0;">Customer <strong>{customer_name or 'N/A'}</strong> has <strong>accepted</strong> your quotation for PR #<strong>{pr_id}</strong>.</p>
            <p style="margin: 0 0 15px 0;">Your quotation met the customer's requirements and has been approved. Please proceed with the next steps as per your standard business process.</p>
        </div>
        
        <!-- CTA Button -->
        <div style="text-align: center; margin: 30px 0;">
            <a href="{APP_BASE_URL}/admin/login" style="display: inline-block; background-color: #28a745; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 16px;">
                🚀 View Dashboard
            </a>
        </div>
        
        <!-- Footer -->
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; text-align: center;">
            <p style="color: #6c757d; margin: 5px 0; font-size: 14px;">Thank you for your excellent service!</p>
            <p style="color: #6c757d; margin: 5px 0; font-size: 12px;">This is an automated message. Please do not reply to this email.</p>
        </div>
        
    </div>
</body>
</html>
                """.strip()
            elif new_status == 'Rejected':
                subject = f"❌ Quotation Rejected by {customer_name}"
                message = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
    <div style="max-width: 600px; margin: 20px auto; padding: 20px; border: 1px solid #ddd; border-radius: 8px; background-color: #ffffff;">
        
        <!-- Header -->
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #dc3545; margin: 0; font-size: 24px;">❌ Quotation Rejected</h1>
            <p style="color: #7f8c8d; margin: 5px 0 0 0; font-size: 14px;">PR #{pr_id} - Review Required</p>
        </div>
        
        <!-- Customer Information -->
        <div style="background-color: #f8d7da; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #dc3545;">
            <h3 style="color: #721c24; margin: 0 0 15px 0; font-size: 16px;">📋 Rejection Details</h3>
            <p style="margin: 8px 0;"><strong>Customer Name:</strong> {customer_name or 'N/A'}</p>
            <p style="margin: 8px 0;"><strong>PR Number:</strong> #{pr_id}</p>
            <p style="margin: 8px 0;"><strong>Status:</strong> <span style="color: #721c24; font-weight: bold;">Rejected</span></p>
        </div>
        
        <!-- Action Required -->
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #007bff;">
            <h3 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 16px;">📝 Status Update</h3>
            <p style="margin: 0 0 15px 0;">Customer <strong>{customer_name or 'N/A'}</strong> has <strong>rejected</strong> your quotation for PR #<strong>{pr_id}</strong>.</p>
            <p style="margin: 0 0 15px 0;">The customer has decided not to proceed with this quotation at this time. We encourage you to continue submitting competitive quotations for future opportunities.</p>
        </div>
        
        <!-- CTA Button -->
        <div style="text-align: center; margin: 30px 0;">
            <a href="{APP_BASE_URL}/admin/login" style="display: inline-block; background-color: #dc3545; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 16px;">
                🚀 View Dashboard
            </a>
        </div>
        
        <!-- Footer -->
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; text-align: center;">
            <p style="color: #6c757d; margin: 5px 0; font-size: 14px;">Thank you for your participation.</p>
            <p style="color: #6c757d; margin: 5px 0; font-size: 12px;">This is an automated message. Please do not reply to this email.</p>
        </div>
        
    </div>
</body>
</html>
                """.strip()
            else:  # Partially Accepted
                subject = f"⚠️ Quotation Partially Accepted by {customer_name}"
                message = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 0;">
    <div style="max-width: 600px; margin: 20px auto; padding: 20px; border: 1px solid #ddd; border-radius: 8px; background-color: #ffffff;">
        
        <!-- Header -->
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #ffc107; margin: 0; font-size: 24px;">⚠️ Quotation Partially Accepted</h1>
            <p style="color: #7f8c8d; margin: 5px 0 0 0; font-size: 14px;">PR #{pr_id} - Action Required</p>
        </div>
        
        <!-- Customer Information -->
        <div style="background-color: #fff3cd; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #ffc107;">
            <h3 style="color: #856404; margin: 0 0 15px 0; font-size: 16px;">📋 Partial Acceptance Details</h3>
            <p style="margin: 8px 0;"><strong>Customer Name:</strong> {customer_name or 'N/A'}</p>
            <p style="margin: 8px 0;"><strong>PR Number:</strong> #{pr_id}</p>
            <p style="margin: 8px 0;"><strong>Status:</strong> <span style="color: #856404; font-weight: bold;">Partially Accepted</span></p>
        </div>
        
        <!-- Action Required -->
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 6px; margin: 20px 0; border-left: 4px solid #007bff;">
            <h3 style="color: #2c3e50; margin: 0 0 15px 0; font-size: 16px;">📋 Partial Acceptance Explained</h3>
            <p style="margin: 0 0 15px 0;">Customer <strong>{customer_name or 'N/A'}</strong> has <strong>partially accepted</strong> your quotation for PR #<strong>{pr_id}</strong>.</p>
            <p style="margin: 0 0 15px 0;">This means the customer has accepted some items from your quotation while rejecting others. Please review the detailed breakdown in your dashboard to see which items were accepted and which were rejected.</p>
            <p style="margin: 0 0 15px 0;">Please proceed with the accepted items as per your standard business process.</p>
        </div>
        
        <!-- CTA Button -->
        <div style="text-align: center; margin: 30px 0;">
            <a href="{APP_BASE_URL}/admin/login" style="display: inline-block; background-color: #ffc107; color: #212529; padding: 12px 30px; text-decoration: none; border-radius: 6px; font-weight: bold; font-size: 16px;">
                🚀 View Dashboard
            </a>
        </div>
        
        <!-- Footer -->
        <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; text-align: center;">
            <p style="color: #6c757d; margin: 5px 0; font-size: 14px;">Thank you for your participation.</p>
            <p style="color: #6c757d; margin: 5px 0; font-size: 12px;">This is an automated message. Please do not reply to this email.</p>
        </div>
        
    </div>
</body>
</html>
                """.strip()
            
            notification_id = send_email_notification(
                event_type="QUOTATION_ACCEPTED" if new_status == 'Accepted' else "QUOTATION_REJECTED" if new_status == 'Rejected' else "QUOTATION_PARTIALLY_ACCEPTED",
                ref_id=int(pr_id),
                recipient_role="supplier",
                recipient_email=supplier_email,
                subject=subject,
                message=message,
                send_email=False,
            )

            if notification_id:
                print(f"[INFO] Quotation status notification logged with ID: {notification_id}")
            else:
                print(f"[WARN] Failed to log quotation status notification")

            # NOTIFICATION: PO Created / Quotation Status
            notif_msg = f"Quotation #{pr_id} status updated to {new_status}"
            create_notification(conn, supplier_email, "supplier", f"Customer updated PR #{pr_id} to {new_status}", f"/admin/quotation/{pr_id}") # Supplier
            # Also notify customer if needed? Usually they know their own action.
            # But let's notify the Customer for confirmation
            # create_notification(conn, user['email'], "customer", f"You {new_status} Quotation #{pr_id}", f"/customer/quotation/{pr_id}") # Passive confirmation
            
            dispatch_pending_notifications()
        
        # Ensure transaction is committed
        conn.commit()
        conn.close()
        
        if accepted_count > 0:
            return templates.TemplateResponse("po.html", {"request": request, "message": "Purchase Order Created Successfully!"})
        else:
            return RedirectResponse(url="/customer/quotations", status_code=303)
            
    except Exception as e:
        print(f"[ERROR] Quotation action failed: {e}")
        if 'conn' in locals():
            conn.close()
        return RedirectResponse(url="/customer/quotations", status_code=303)

@app.get("/customer/invoices", response_class=HTMLResponse)
async def customer_invoices(
    request: Request,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    """Customer Invoice List - Shows only generated invoices (not just POs)"""
    user = require_customer(request)
    conn = get_db_connection()
    
    # Query: Get invoices for customer's POs (only actual invoices, not just POs)
    query = '''
        SELECT 
            i.id,
            i.invoice_no,
            i.created_at,
            po.total_amount,
            po.amount_received,
            i.status,
            i.payment_mode,
            po.display_id,
            po.invoice_source
        FROM purchase_orders po
        INNER JOIN invoices i ON i.po_id = po.id
        WHERE po.customer_email = ? AND po.is_active = 1
    '''
    
    params = [user['email']]
    where_clauses = []
    
    if search:
        s = f"%{search.lower().strip()}%"
        where_clauses.append("(LOWER(i.invoice_no) LIKE ? OR LOWER(po.display_id) LIKE ?)")
        params.extend([s, s])
        
    if date_from:
        where_clauses.append("po.created_at >= ?")
        params.append(f"{date_from} 00:00:00")
        
    if date_to:
        where_clauses.append("po.created_at <= ?")
        params.append(f"{date_to} 23:59:59")
        
    if where_clauses:
        query += " AND " + " AND ".join(where_clauses)
        
    query += " ORDER BY po.created_at DESC"
    
    invoices = conn.execute(query, tuple(params)).fetchall()
    conn.close()
    
    return templates.TemplateResponse("customer_report_invoices.html", {
        "request": request,
        "invoices": invoices,
        "search": search or "",
        "date_from": date_from or "",
        "date_to": date_to or ""
    })

@app.get("/customer/statement", response_class=HTMLResponse)
async def customer_statement(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    """Customer Account Statement / Ledger"""
    logging.error("DEBUG_LOG: >>> ENTERING CUSTOMER STATEMENT ROUTE <<<")
    try:
        user = require_customer(request)
        logging.error(f"DEBUG_LOG: User authenticated: {user['email']}")
        conn = get_db_connection()
        
        ledger_entries = []
        
        # Get all POs for this customer that have been invoiced, sorted OLD to NEW to calculate running balance
        query = '''
            SELECT po.*, c.category as customer_category, i.invoice_no
            FROM purchase_orders po
            JOIN customers c ON po.customer_email = c.email
            INNER JOIN invoices i ON i.po_id = po.id
            WHERE po.customer_email = ? AND po.is_active = 1
        '''
        params = [user['email']]
        
        # We need ALL POs to calculate the opening balance correctly
        query += " ORDER BY po.created_at ASC"
        
        if DATABASE_URL:
            # Use %s for PostgreSQL
            pg_query = query.replace('?', '%s')
            all_pos = conn.execute(pg_query, tuple(params)).fetchall()
        else:
            all_pos = conn.execute(query, tuple(params)).fetchall()
        
        running_balance = 0.0
        for po_row in all_pos:
            # ABSOLUTE FIX: Convert to dict to ensure .get() works on SQLite Rows
            po = dict(po_row)
            
            opening_bal = running_balance
            inv_amt = float(po.get('total_amount') or 0.0) # Handle None/Null
            inv_rec = float(po.get('invoice_receipt') or 0.0)
            amt_rec = float(po.get('amount_received') or 0.0)
            
            closing_bal = opening_bal + inv_amt - inv_rec - amt_rec
            
            # Handle date whether string or datetime object
            created_at = po.get('created_at')
            if hasattr(created_at, 'strftime'):
                 date_str = created_at.strftime('%Y-%m-%d')
            else:
                 date_str = str(created_at).split(' ')[0] if created_at else ""

            entry = {
                "id": po['id'],
                "display_id": po.get('invoice_no') or po.get('display_id') or f"PO-{po['id']}",
                "pr_id": po.get('pr_id'),
                "date": date_str,
                "opening_bal": opening_bal,
                "inv_amt": inv_amt,
                "inv_rec": inv_rec,
                "amt_rec": amt_rec,
                "closing_bal": closing_bal
            }
            
            # Apply date filter for DISPLAY only
            show = True
            if date_from and entry['date'] < date_from: show = False
            if date_to and entry['date'] > date_to: show = False
            
            if show:
                ledger_entries.append(entry)
                
            running_balance = closing_bal

        conn.close()
        
        return templates.TemplateResponse("customer_statement.html", {
            "request": request,
            "ledger": ledger_entries,
            "date_from": date_from or "",
            "date_to": date_to or ""
        })
    except Exception as e:
        if 'conn' in locals(): conn.close()
        import traceback
        traceback.print_exc()
        return HTMLResponse(content=f"<h1>Inner Error: {str(e)}</h1><pre>{traceback.format_exc()}</pre>", status_code=500)

@app.get("/admin/direct_sales", response_class=HTMLResponse)
async def admin_direct_sales_form(request: Request):
    require_admin(request)
    conn = get_db_connection()
    
    # 1. Customers with Favorite Status
    customers = conn.execute("SELECT email, name, business_name, is_favorite FROM customers WHERE status='Active' ORDER BY name").fetchall()
    
    # 2. Products with Stock and Favorite Price
    # Use InventoryService to get accurate stock
    inv_service = InventoryService(conn)
    stock_summary = inv_service.get_stock_summary() # Returns list of dicts with shop_stock
    
    # Create a map for quick access
    stock_map = {p['id']: p['shop_stock'] for p in stock_summary}
    
    # Fetch base product details including favorite_price
    # Note: 'favorite_price' might be missing in some older schemas, handle gracefully if needed, 
    # but based on user feedback it implies it exists.
    products_db = conn.execute("SELECT id, name, unit, rate, favorite_price FROM products ORDER BY name").fetchall()
    
    products = []
    for p in products_db:
        products.append({
            "id": p["id"],
            "name": p["name"],
            "unit": p["unit"],
            "rate": p["rate"],
            "favorite_price": p["favorite_price"],
            "shop_stock": stock_map.get(p["id"], 0)
        })
        
    conn.close()
    return templates.TemplateResponse("admin_direct_sales.html", {
        "request": request,
        "customers": customers,
        "products": products
    })

@app.post("/admin/direct_sales")
async def admin_direct_sales_submit(
    request: Request,
    customer_email: str = Form(...),
    product_id: List[int] = Form(...),
    qty: List[float] = Form(...),
    rate: List[float] = Form(...)
):
    require_admin(request)
    conn = get_db_connection()
    
    # Validation
    if not product_id or len(product_id) == 0:
        conn.close()
        return RedirectResponse(url="/admin/direct_sales?error=no_items", status_code=303)
        
    items = []
    total_amount = 0
    
    # Fetch customer info for snapshots
    customer = conn.execute("SELECT name, business_name, address, category, email, mobile FROM customers WHERE email = ?", (customer_email,)).fetchone()
    if not customer:
        conn.close()
        return RedirectResponse(url="/admin/direct_sales?error=customer_not_found", status_code=303)

    # Fetch product info for the items_json
    p_ids = ",".join([str(pid) for pid in product_id])
    products_rows = conn.execute(f"SELECT id, name, name_marathi, unit FROM products WHERE id IN ({p_ids})").fetchall()
    prod_map = {p['id']: p for p in products_rows}
    
    for i in range(len(product_id)):
        pid = product_id[i]
        q = qty[i]
        r = rate[i]
        if q > 0:
            prod = prod_map.get(pid)
            if prod:
                amt = q * r
                items.append({
                    "product_id": pid,
                    "name": prod['name'],
                    "name_marathi": prod['name_marathi'],
                    "unit": prod['unit'],
                    "qty": q,
                    "quoted_rate": r,
                    "amount": amt,
                    "item_status": "Accepted",
                    "decision_timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "is_locked": True
                })
                total_amount += amt
                
    if not items:
        conn.close()
        return RedirectResponse(url="/admin/direct_sales?error=invalid_items", status_code=303)
        
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Insert Direct Purchase Order (no pr_id)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO purchase_orders 
        (pr_id, customer_email, created_at, total_amount, items_json, invoice_source, delivery_status, tracking_status, status, 
         customer_name_snapshot, business_name_snapshot, address_snapshot, customer_category_snapshot,
         customer_email_snapshot, customer_mobile_snapshot) 
        VALUES (NULL, ?, ?, ?, ?, 'DIRECT', 'OPEN', 'PO_CREATED', 'Accepted', ?, ?, ?, ?, ?, ?)
    ''', (customer_email, created_at, total_amount, json.dumps(items), 
          customer['name'], customer['business_name'], customer['address'], customer['category'],
          customer['email'], customer['mobile']))
    
    print(f"[DEBUG] Direct PO created with delivery_status=OPEN")
    
    po_id = cursor.lastrowid
    display_id = f"PO-{po_id}"
    cursor.execute("UPDATE purchase_orders SET display_id = ? WHERE id = ?", (display_id, po_id))
    
    # --- AUTO-GENERATE INVOICE ---
    try:
        today_str = datetime.now().strftime("%Y%m%d")
        invoice_no = f"DS-INV-{today_str}-{po_id}"
        
        cursor.execute("""
            INSERT INTO invoices (po_id, invoice_no, created_at, payment_mode, delivery_remarks, items_json, status, amount_received)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (po_id, invoice_no, created_at, 'Credit', 'Direct Sale', json.dumps(items), 'GENERATED', 0))
        
        print(f"[DEBUG] Auto-generated Invoice {invoice_no} for Direct Sale PO {po_id}")
    except Exception as e:
        print(f"[ERROR] Failed to auto-generate invoice for Direct Sale: {e}")
        # We don't rollback the PO creation, as the sale is valid, but logging the error is crucial.
        
    # --- DEDUCT STOCK (Direct Sales = Immediate Deduction) ---
    try:
        inv_service = InventoryService(conn)
        for item in items:
            product_id = int(item['product_id'])
            qty = float(item['qty'])
            # Assuming Direct Sales deducts from SHOP (Location ID 2)
            # Using record_sale_out which wraps reduce_stock for shop
            inv_service.record_sale_out(
                product_id=product_id,
                qty=qty,
                reference_type='DIRECT_PO',
                reference_id=display_id,
                commit=False # Commit at the end with the main transaction
            )
        print(f"[DEBUG] Stock deducted for Direct Sale PO {display_id}")
    except Exception as e:
        print(f"[ERROR] Failed to deduct stock for Direct Sale: {e}")
        # Note: If stock deduction fails (e.g., negative stock not allowed), 
        # the transaction commit below might fail or we should rollback?
        # Ideally, we should rollback everything if stock fails.
        conn.rollback() 
        conn.close()
        return RedirectResponse(url="/admin/direct_sales?error=stock_error", status_code=303)
    
    conn.commit()
    conn.close()
    
    return RedirectResponse(url=f"/invoice/po/{po_id}?success=direct_sales", status_code=303)

@app.get("/admin/orders/revise/{po_id}", response_class=HTMLResponse)
async def revise_invoice_form(request: Request, po_id: int):
    require_admin(request)
    conn = get_db_connection()
    po = conn.execute("SELECT *, delivery_status, tracking_status FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
    if not po:
        conn.close()
        return Response("PO not found", status_code=404)
        
    products = conn.execute("SELECT id, name, unit, rate FROM products ORDER BY name").fetchall()
    conn.close()
    
    items = json.loads(po['items_json'])
    
    return templates.TemplateResponse("admin_revise_invoice.html", {
        "request": request,
        "po": po,
        "items": items,
        "products": products
    })

@app.post("/admin/orders/revise/{po_id}")
async def revise_invoice_submit(
    request: Request, 
    po_id: int,
    product_id: List[int] = Form(...),
    qty: List[float] = Form(...),
    rate: List[float] = Form(...),
    revision_reason: Optional[str] = Form(None)
):
    require_admin(request)
    conn = get_db_connection()
    
    # 1. Fetch Original PO info
    original_po = conn.execute("SELECT *, delivery_status, tracking_status FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
    if not original_po:
        conn.close()
        return Response("PO not found", status_code=404)

    # 2. Build New Items List
    p_ids = ",".join([str(pid) for pid in product_id])
    products_rows = conn.execute(f"SELECT id, name, name_marathi, unit FROM products WHERE id IN ({p_ids})").fetchall()
    prod_map = {p['id']: dict(p) for p in products_rows}
    
    # Fetch customer info for snapshots (Re-snapshot during revision to reflect current customer details for the NEW RPO)
    customer = conn.execute("SELECT name, business_name, address, category, email, mobile FROM customers WHERE email = ?", (original_po['customer_email'],)).fetchone()

    items = []
    total_amount = 0
    for i in range(len(product_id)):
        pid = product_id[i]
        q = qty[i]
        r = rate[i]
        if q > 0:
            prod = prod_map.get(pid)
            if prod:
                amt = q * r
                items.append({
                    "product_id": pid,
                    "name": prod['name'],
                    "name_marathi": prod.get('name_marathi', ''),
                    "unit": prod['unit'],
                    "qty": q,
                    "quoted_rate": r,
                    "amount": amt,
                    "item_status": "Accepted", # Revisions are implicitly accepted
                    "is_locked": True
                })
                total_amount += amt

    # 3. Supersede Old PO
    conn.execute("UPDATE purchase_orders SET is_active = 0 WHERE id = ?", (po_id,))
    
    # 4. Determine Root Revision ID
    # If the PO being revised is itself a revision, it has a revision_of_id. Use that.
    # If it's a root PO, its revision_of_id is None, so use its ID.
    root_id = original_po['revision_of_id'] if original_po['revision_of_id'] else po_id
    
    # 5. Generate RPO Number
    # Count how many RPOs exist globally to maintain series RPO-1, RPO-2...
    rpo_count = conn.execute("SELECT count(*) FROM purchase_orders WHERE revision_of_id IS NOT NULL").fetchone()[0]
    next_rpo_num = rpo_count + 1
    display_id = f"RPO-{next_rpo_num}"
    
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 6. Insert Revised PO
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO purchase_orders 
        (pr_id, customer_email, created_at, total_amount, items_json, invoice_source, revision_of_id, display_id, is_active, revision_reason, amount_received, delivery_status, tracking_status, status,
         customer_name_snapshot, business_name_snapshot, address_snapshot, customer_category_snapshot,
         customer_email_snapshot, customer_mobile_snapshot) 
        VALUES (?, ?, ?, ?, ?, 'Revised', ?, ?, 1, ?, ?, 'OPEN', 'PO_CREATED', 'Accepted', ?, ?, ?, ?, ?, ?)
    ''', (original_po['pr_id'], original_po['customer_email'], created_at, total_amount, json.dumps(items), root_id, display_id, revision_reason, original_po['amount_received'],
          customer['name'], customer['business_name'], customer['address'], customer['category'],
          customer['email'], customer['mobile']))
    
    print(f"[DEBUG] Revised PO created with delivery_status=OPEN for PO #{root_id}")
    
    new_po_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return RedirectResponse(url=f"/invoice/po/{new_po_id}", status_code=303)


@app.get("/invoice/po/{po_id}", response_class=HTMLResponse)
async def view_invoice_by_po(request: Request, po_id: int):
    user = get_current_user(request)
    if not user:
        return RedirectResponse(url="/", status_code=303)
        
    conn = get_db_connection()
    po_row = conn.execute("SELECT *, delivery_status, tracking_status FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
    
    if not po_row:
        conn.close()
        return Response("Invoice not found", status_code=404)
        
    po = dict(po_row)
    # Fetch actual invoice number if exists
    inv = conn.execute("SELECT invoice_no FROM invoices WHERE po_id = ?", (po_id,)).fetchone()
    if inv:
        po['invoice_no'] = inv['invoice_no']

    if user['role'] == 'customer':
        if po['customer_email'] != user['email']:
            conn.close()
            return Response("Unauthorized", status_code=403)
        # Customer should NOT see inactive invoices if they are superseded?
        # Requirement: "Customer can view: ONLY the latest revised invoice (never the original once revised)"
        if po['is_active'] == 0:
             # Find the active one?
             # But if they click a link to an old one, maybe show a warning or redirect?
             # For now, let's allow viewing but maybe show "Superseded" banner.
             # Strict reading: "never the original".
             # Let's try to find the latest revision for this root.
             # Find the active revision for this root
             root_id = po['revision_of_id'] if po['revision_of_id'] else po['id']
             active_po = conn.execute('''
                SELECT id, delivery_status, tracking_status FROM purchase_orders 
                WHERE (id = ? OR revision_of_id = ?) AND is_active = 1
                LIMIT 1
             ''', (root_id, root_id)).fetchone()
             
             if active_po:
                 conn.close()
                 return RedirectResponse(url=f"/invoice/po/{active_po['id']}", status_code=303)

    raw_customer = conn.execute("SELECT * FROM customers WHERE email = ?", (po['customer_email'],)).fetchone()
    
    # Snapshot fallback logic
    customer = {
        "name": po['customer_name_snapshot'] if po['customer_name_snapshot'] else (raw_customer['name'] if raw_customer else "Unknown"),
        "business_name": po['business_name_snapshot'] if po['business_name_snapshot'] else (raw_customer['business_name'] if raw_customer else ""),
        "address": po['address_snapshot'] if po['address_snapshot'] else (raw_customer['address'] if raw_customer else ""),
        "category": po['customer_category_snapshot'] if po['customer_category_snapshot'] else (raw_customer['category'] if raw_customer else ""),
        "email": po['customer_email_snapshot'] if po['customer_email_snapshot'] else (raw_customer['email'] if raw_customer else ""),
        "mobile": po['customer_mobile_snapshot'] if po['customer_mobile_snapshot'] else (raw_customer['mobile'] if raw_customer else (raw_customer['phone'] if raw_customer else ""))
    }

    supplier_name = get_setting("supplier_name", "Hotel Supplier Inc.")
    items = json.loads(po['items_json'])
    
    # Calculate Outstanding Balance (Dynamic)
    # "Previous Balance" = Sum of (Total - Received) for all Previous POs
    
    # 1. Fetch all previous purchase orders for this customer
    # Strict inequality to exclude current timestamp (if created sequentially)
    # But usually unique ID is helper if timestamps match. Using created_at first.
    prev_pos = conn.execute('''
        SELECT total_amount, amount_received 
        FROM purchase_orders 
        WHERE customer_email = ? AND created_at < ? AND is_active = 1
    ''', (po['customer_email'], po['created_at'])).fetchall()
    
    previous_balance = 0.0
    for p in prev_pos:
        amt = p['total_amount']
        rec = p['amount_received'] or 0.0
        previous_balance += (amt - rec)
        
    # "Total Outstanding" = Previous Balance + This Invoice Amount
    # (Assuming this invoice has not been paid yet. If amount_received > 0 on this invoice, 
    # the 'Net Payable' might be different, but 'Total Outstanding' implies Gross Debt context usually)
    # The requirement: "Total Outstanding Balance = Previous Balance + Invoice Amount"
    
    total_outstanding = previous_balance + po['total_amount']

    amount_in_words = number_to_words(total_outstanding)
    
    conn.close()
    
    return templates.TemplateResponse("invoice.html", {
        "request": request, 
        "po": po, 
        "customer": customer, 
        "items": items, 
        "supplier_name": supplier_name,
        "total": po['total_amount'],
        "previous_balance": previous_balance,
        "outstanding_balance": total_outstanding,
        "amount_in_words": amount_in_words,
        "is_revision_mode": (po['is_active'] == 0), 
        "is_admin": (user['role'] == 'admin')
    })

# --- Routes: Order Tracking ---

@app.post("/admin/orders/update-tracking/{po_id}")
async def update_order_tracking(request: Request, po_id: int):
    """Supplier-only endpoint to update tracking status"""
    require_admin(request)
    
    form_data = await request.form()
    new_tracking_status = form_data.get("tracking_status")
    
    # Validate tracking status
    valid_statuses = ['PO_CREATED', 'PROCESSING', 'OUT_FOR_DELIVERY', 'DELIVERED']
    if new_tracking_status not in valid_statuses:
        return JSONResponse(
            status_code=400, 
            content={"error": f"Invalid tracking status. Must be one of: {', '.join(valid_statuses)}"}
        )
    
    conn = get_db_connection()
    
    # Check if PO exists and is OPEN
    po = conn.execute("SELECT id, delivery_status, tracking_status FROM purchase_orders WHERE id = ?", (po_id,)).fetchone()
    if not po:
        conn.close()
        return JSONResponse(status_code=404, content={"error": "Purchase Order not found"})
    
    if po['delivery_status'] != 'OPEN':
        conn.close()
        return JSONResponse(status_code=400, content={"error": "Cannot update tracking status. Order is not OPEN."})
    
    # Validate status progression (can only move forward)
    status_order = ['PO_CREATED', 'PROCESSING', 'OUT_FOR_DELIVERY', 'DELIVERED']
    current_index = status_order.index(po['tracking_status'])
    new_index = status_order.index(new_tracking_status)
    
    if new_index <= current_index:
        conn.close()
        return JSONResponse(status_code=400, content={"error": "Can only move tracking status forward."})
    
    # Update tracking status
    try:
        conn.execute("UPDATE purchase_orders SET tracking_status = ? WHERE id = ?", (new_tracking_status, po_id))
        conn.commit()
        
        print(f"[DEBUG] Tracking status updated for PO #{po_id}: {po['tracking_status']} → {new_tracking_status}")
        
        conn.close()
        return JSONResponse(content={
            "success": True,
            "po_id": po_id,
            "old_status": po['tracking_status'],
            "new_status": new_tracking_status
        })
    except Exception as e:
        conn.close()
        print(f"[ERROR] Failed to update tracking status for PO #{po_id}: {e}")
        return JSONResponse(status_code=500, content={"error": "Failed to update tracking status"})

# --- Routes: Invoice Creation at Delivery ---

@app.get("/admin/orders/{po_id}", response_class=HTMLResponse)
async def admin_view_po(request: Request, po_id: int):
    user = require_admin(request)
    conn = get_db_connection()
    
    # Fetch PO with customer info
    po_row = conn.execute('''
        SELECT po.*, c.name as customer_name, c.business_name, c.address, c.mobile
        FROM purchase_orders po
        LEFT JOIN customers c ON po.customer_email = c.email
        WHERE po.id = ? AND po.is_active = 1
    ''', (po_id,)).fetchone()
    
    if not po_row:
        conn.close()
        return templates.TemplateResponse("admin_dashboard.html", {
            "request": request,
            "error": "Purchase Order not found"
        })
        
    po = dict(po_row)
    
    # Snapshot fallback logic
    customer = {
        "name": po['customer_name_snapshot'] if po['customer_name_snapshot'] else po.get('customer_name', "Unknown"),
        "business_name": po['business_name_snapshot'] if po['business_name_snapshot'] else po.get('business_name', ""),
        "address": po['address_snapshot'] if po['address_snapshot'] else po.get('address', ""),
        "mobile": po['customer_mobile_snapshot'] if po['customer_mobile_snapshot'] else po.get('mobile', "")
    }

    try:
        items = json.loads(po['items_json'])
    except:
        items = []
        
    conn.close()
    
    return templates.TemplateResponse("admin_po_details.html", {
        "request": request,
        "po": po,
        "customer": customer,
        "items": items,
        "is_admin": True
    })

@app.get("/admin/orders/create-invoice/{po_id}", response_class=HTMLResponse)
async def create_invoice_form(request: Request, po_id: int):
    """Show invoice creation form for supplier"""
    user = require_admin(request)
    conn = get_db_connection()
    
    # Fetch PO with validation
    po = conn.execute('''
        SELECT po.*, c.name as customer_name, c.business_name
        FROM purchase_orders po
        LEFT JOIN customers c ON po.customer_email = c.email
        WHERE po.id = ? AND po.is_active = 1 AND po.delivery_status = 'OPEN'
    ''', (po_id,)).fetchone()
    
    if not po:
        conn.close()
        return templates.TemplateResponse("admin_dashboard.html", {
            "request": request,
            "error": "Purchase Order not found or cannot be invoiced"
        })
    
    # Parse items for pre-filling
    try:
        items = json.loads(po['items_json'])
        # Ensure ID exists for frontend JS using product_id fallback
        for item in items:
            if not item.get('id') and item.get('product_id'):
                item['id'] = item.get('product_id')
    except:
        items = []
    
    # Fetch all active products for "Add Item" feature
    products = conn.execute('SELECT id, name, unit, rate, name_marathi FROM products ORDER BY name').fetchall()
    
    conn.close()
    
    return templates.TemplateResponse("create_invoice.html", {
        "request": request,
        "po": dict(po), # Convert to dict to be safe
        "items": items,
        "products": [dict(p) for p in products]
    })

@app.post("/admin/orders/create-invoice/{po_id}")
async def create_invoice_submit(request: Request, po_id: int):
    """Process invoice creation at delivery with dynamic item modification"""
    user = require_admin(request)
    form_data = await request.form()
    action = form_data.get('action', 'create')
    
    conn = get_db_connection()
    
    try:
        # Validate PO is still open
        po = conn.execute('''
            SELECT id, delivery_status, customer_email FROM purchase_orders 
            WHERE id = ? AND is_active = 1
        ''', (po_id,)).fetchone()
        
        if not po or po['delivery_status'] != 'OPEN':
            conn.close()
            return JSONResponse(status_code=400, content={"error": "Purchase Order cannot be modified or invoiced"})

        cursor = conn.cursor()
        
        # Handle Cancellation
        if action == 'cancel':
            cursor.execute('''
                UPDATE purchase_orders 
                SET delivery_status = 'CANCELLED'
                WHERE id = ?
            ''', (po_id,))
            conn.commit()
            conn.close()
            print(f"[INFO] PO #{po_id} cancelled by admin")
            return RedirectResponse(url="/admin/reports/pos?success=po_cancelled", status_code=303)
        
        # Process Items (Dynamic list from form arrays)
        # Form fields: item_id[], item_qty[], item_rate[], item_name[], item_unit[]
        item_ids = form_data.getlist('item_id[]')
        item_names = form_data.getlist('item_name[]')
        item_units = form_data.getlist('item_unit[]')
        item_qtys = form_data.getlist('item_qty[]')
        item_rates = form_data.getlist('item_rate[]')
        
        updated_items = []
        grand_total = 0.0
        
        for i in range(len(item_ids)):
            qty = float(item_qtys[i] or 0)
            if qty <= 0: continue # Skip zero qty items
            
            rate = float(item_rates[i] or 0)
            amount = qty * rate
            grand_total += amount
            
            updated_items.append({
                "id": int(item_ids[i]),
                "name": item_names[i],
                "unit": item_units[i],
                "qty": qty,  # Final delivered qty
                "delivered_qty": qty,
                "quoted_rate": rate,
                "amount": amount
            })

        if not updated_items:
            conn.close()
            return JSONResponse(status_code=400, content={"error": "Cannot create empty invoice. Please add items or cancel PO."})
        
        # Create invoice record
        invoice_no = f"INV-{po_id}-{int(datetime.now().timestamp())}"
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        amount_received = float(form_data.get('amount_received', 0) or 0)
        
        cursor.execute('''
            INSERT INTO invoices 
            (po_id, invoice_no, created_at, payment_mode, delivery_remarks, items_json, status)
            VALUES (?, ?, ?, ?, ?, ?, 'PENDING')
        ''', (
            po_id, invoice_no, created_at, 
            form_data.get('payment_mode', 'CASH'), 
            form_data.get('delivery_remarks', ''), 
            json.dumps(updated_items)
        ))
        
        # Update PO status and final items/totals
        # Fetch customer details for snapshot
        customer = conn.execute("SELECT name, business_name, address, category, email, mobile FROM customers WHERE email = ?", (po['customer_email'],)).fetchone()
        
        # Update PO status and final items/totals WITH SNAPSHOTS
        cursor.execute('''
            UPDATE purchase_orders 
            SET delivery_status = 'INVOICED', 
                delivered_at = ?, 
                items_json = ?, 
                total_amount = ?,
                invoice_receipt = ?,
                customer_name_snapshot = ?,
                business_name_snapshot = ?,
                address_snapshot = ?,
                customer_category_snapshot = ?,
                customer_email_snapshot = ?,
                customer_mobile_snapshot = ?
            WHERE id = ?
        ''', (
            created_at, json.dumps(updated_items), grand_total, 
            float(form_data.get('amount_received', 0) or 0),
            customer['name'] if customer else None,
            customer['business_name'] if customer else None,
            customer['address'] if customer else None,
            customer['category'] if customer else None,
            customer['email'] if customer else None,
            customer['mobile'] if customer else None,
            po_id
        ))
        
        conn.commit()
        
        # NOTIFICATION: Invoice Generated
        create_notification(conn, customer['email'], "customer", f"New Invoice {invoice_no} generated", f"/customer/invoices?search={invoice_no}")
        conn.commit()
        
        conn.close()
        
        print(f"[INFO] Invoice generated: {invoice_no} (Total: {grand_total})")
        return RedirectResponse(url=f"/invoice/po/{po_id}?success=invoice_created", status_code=303)
        
    except Exception as e:
        if conn: conn.close()
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": f"Internal Server Error: {str(e)}"})

# Amazon-Style Delivery Tracking Routes Management
@app.get("/admin/delivery_routes", response_class=HTMLResponse)
async def admin_delivery_routes(request: Request):
    require_admin(request)
    conn = get_db_connection()
    
    # Get all categories with hierarchical structure
    categories = conn.execute("""
        SELECT 
            cc.id, cc.name, cc.delivery_days, cc.route_name,
            cc.parent_id, cc.level,
            parent.name as parent_name,
            cc.is_active
        FROM customer_categories cc
        LEFT JOIN customer_categories parent ON cc.parent_id = parent.id
        ORDER BY cc.level, cc.name
    """).fetchall()
    
    conn.close()
    return templates.TemplateResponse("admin_delivery_routes.html", {
        "request": request,
        "categories": categories
    })

@app.post("/admin/delivery_routes/update")
async def update_delivery_route(request: Request):
    require_admin(request)
    form = await request.form()
    category_id = form.get("category_id")
    delivery_days = form.getlist("delivery_days")  # ["MONDAY", "FRIDAY"]
    route_name = form.get("route_name", "")
    
    conn = get_db_connection()
    
    # Update category
    if category_id and delivery_days:
        conn.execute("""
            UPDATE customer_categories
            SET delivery_days = ?, route_name = ?
            WHERE id = ?
        """, (json.dumps(delivery_days), route_name, category_id))
        
        # If updating parent, update all children
        if category_id and form.get("update_children") == "true":
            conn.execute("""
                UPDATE customer_categories
                SET parent_id = ?
                WHERE parent_id = ?
            """, (category_id, category_id))
    
    conn.commit()
    conn.close()
    
    return RedirectResponse(url="/admin/delivery_routes", status_code=303)

# Daily Background Job for Delivery Stage Updates
@app.on_event("startup")
async def startup_event():
    print("Starting background tasks...")
    try:
        # Initialize DB (Safe to run multiple times)
        init_db()

        # SUPABASE FIX APPLIED - validate Supabase connection and tables on startup
        check_supabase_connection()
        check_supabase_tables()

        # daily_delivery_stage_update is defined at the end of this file, 
        # so we can call it here as this function runs after full file load
        asyncio.create_task(daily_delivery_stage_update())
        print("Background delivery update task started.")
    except Exception as e:
        print(f"Error starting background task: {e}")
        raise  # Re-raise so startup failure is visible in logs

async def daily_delivery_stage_update():
    while True:
        try:
            print("Running scheduled delivery update...")
            update_all_delivery_stages()
        except Exception as e:
            print(f"Error in delivery update: {e}")
        await asyncio.sleep(3600)  # Run every hour

def update_all_delivery_stages():
    conn = get_db_connection()
    try:
        active_pos = conn.execute("""
            SELECT id, expected_delivery_date, delivery_stage, created_at
            FROM purchase_orders
            WHERE delivery_stage != 'DELIVERED' AND is_active = 1
        """).fetchall()
        
        for po in active_pos:
            update_delivery_timestamps(po['id'])
        
        conn.commit()
    finally:
        conn.close()


# --- Customer Favorite Toggle API ---
class FavoriteRequest(BaseModel):
    email: str
    is_favorite: bool

@app.post("/admin/customers/toggle_favorite")
async def toggle_favorite(request: Request, data: FavoriteRequest):
    require_admin(request)
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("UPDATE customers SET is_favorite = ? WHERE email = ?", (data.is_favorite, data.email))
        conn.commit()
        return {"status": "success", "is_favorite": data.is_favorite}
    except Exception as e:
        print(f"Error toggling favorite: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# -------------------------------------------------------------------------
# NEW: OTP Based Registration Flow (Additive)
# -------------------------------------------------------------------------

@app.get("/register-otp", response_class=HTMLResponse)
async def register_otp_page(request: Request):
    """Entry point for the new OTP-based registration flow."""
    return templates.TemplateResponse("register_otp.html", {"request": request})

@app.post("/auth/request-register-otp")
async def request_register_otp(request: Request):
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        name = data.get("name", "").strip()
        
        if not email or not name:
             return JSONResponse({"success": False, "error": "Email and Name are required"})
             
        conn = get_db_connection()
        try:
            # Check if user already exists
            existing = conn.execute("SELECT email FROM customers WHERE email = ?", (email,)).fetchone()
            if existing:
                 return JSONResponse({"success": False, "error": "User already exists. Please login."})
        
            # Generate OTP
            otp = utils.generate_otp()
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            expires_at = (datetime.now() + timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")

            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                conn.execute("DELETE FROM otp_codes WHERE email = %s", (email,))
                conn.execute("""
                    INSERT INTO otp_codes (email, otp, created_at, expires_at) 
                    VALUES (%s, %s, %s, %s)
                 """, (email, otp, created_at, expires_at))
            else:
                conn.execute("DELETE FROM otp_codes WHERE email = ?", (email,))
                conn.execute("""
                    INSERT INTO otp_codes (email, otp, created_at, expires_at) 
                    VALUES (?, ?, ?, ?)
                 """, (email, otp, created_at, expires_at))
            
            conn.commit()
            
            # Send Email
            from app.notifications import call_supabase_email_function
            subject = "Registration OTP"
            body = f"Hello {name}, your OTP for registration is {otp}. Valid for 10 minutes."
            
            try:
                call_supabase_email_function(None, email, subject, body)
            except Exception as e:
                print(f"OTP Email failed: {e}")
                
            return JSONResponse({"success": True, "message": "OTP sent successfully"})
            
        finally:
            conn.close()
            
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

@app.post("/auth/verify-register-otp")
async def verify_register_otp(request: Request):
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        otp = data.get("otp", "").strip()
        
        if not email or not otp:
            return JSONResponse({"success": False, "error": "Email and OTP required"})

        conn = get_db_connection()
        try:
            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                row = conn.execute("SELECT * FROM otp_codes WHERE email = %s", (email,)).fetchone()
            else:
                row = conn.execute("SELECT * FROM otp_codes WHERE email = ?", (email,)).fetchone()
            
            if not row:
                return JSONResponse({"success": False, "error": "Invalid or Expired OTP"})
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if str(row['otp']) != str(otp):
                 return JSONResponse({"success": False, "error": "Invalid OTP"})
            if str(row['expires_at']) < now:
                return JSONResponse({"success": False, "error": "OTP Expired"})
                
            # SUCCESS: Delete OTP
            if is_pg:
                conn.execute("DELETE FROM otp_codes WHERE email = %s", (email,))
            else:
                conn.execute("DELETE FROM otp_codes WHERE email = ?", (email,))
            conn.commit()
            
            # Set Session Flag
            request.session["registration_verified"] = email
            return JSONResponse({"success": True, "message": "OTP Verified"})
            
        finally:
            conn.close()
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

@app.post("/auth/complete-registration")
async def complete_registration(request: Request):
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        password = data.get("password", "")
        name = data.get("name", "").strip()
        mobile = data.get("mobile", "").strip()
        address = data.get("address", "").strip()
        business_name = data.get("business_name", "").strip()
        category = data.get("category", "General")
        
        # Security Check
        verified_email = request.session.get("registration_verified")
        if not verified_email or verified_email != email:
             return JSONResponse({"success": False, "error": "Session expired or unauthorized. Verify OTP again."})
        
        # Password Complexity Check
        # At least 8 chars, 1 number, 1 special char
        if len(password) < 8:
             return JSONResponse({"success": False, "error": "Password must be at least 8 characters"})
        if not re.search(r"\d", password):
             return JSONResponse({"success": False, "error": "Password must contain at least one number"})
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
             return JSONResponse({"success": False, "error": "Password must contain at least one special character"})

        if not name or not mobile or not address or not business_name:
             return JSONResponse({"success": False, "error": "All fields are required"})

        # SUPABASE FIX APPLIED - create Supabase Auth user FIRST so we have the real UID
        # Uses service role key (admin) so email is auto-confirmed (OTP was already verified)
        import uuid
        supabase_uid = str(uuid.uuid4())  # fallback UUID if Supabase auth fails
        
        try:
            sb_user_resp = supabase.auth.admin.create_user({
                "email": email,
                "password": password,
                "email_confirm": True  # OTP already verified above
            })
            if sb_user_resp and sb_user_resp.user:
                supabase_uid = sb_user_resp.user.id
                print(f"[OK] Supabase Auth user created for {email} (uid={supabase_uid})")
            else:
                print(f"[WARN] Supabase Auth create_user returned no user for {email}")
        except Exception as sb_err:
            err_str = str(sb_err)
            # If user already exists in Supabase, that's OK — use a UUID placeholder
            if "already been registered" in err_str or "already exists" in err_str.lower():
                print(f"[INFO] Supabase Auth user already exists for {email}: {sb_err}")
            else:
                print(f"[ERROR] Supabase Auth user creation failed for {email}: {sb_err}")
                return JSONResponse({"success": False, "error": f"Auth account creation failed: {err_str}"})

        # Insert Customer into local DB (with real Supabase UID)
        conn = get_db_connection()
        try:
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            is_pg = os.getenv("DATABASE_URL") is not None
            
            sql = """
                INSERT INTO customers (email, password, name, phone, full_name, business_name, mobile, address, category, created_at, status, supabase_user_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active', ?)
            """
            
            if is_pg:
                sql = sql.replace("?", "%s")
                
            # Map fields: name -> name, mobile -> phone & mobile (legacy), name -> full_name
            conn.execute(sql, (email, password, name, mobile, name, business_name, mobile, address, category, created_at, supabase_uid))
            conn.commit()
            
            # Clear flag
            request.session.pop("registration_verified", None)
            
            # NOTIFICATION: Welcome
            create_notification(conn, email, "customer", f"Welcome to HotelSys, {name}! Your account is active.")
            conn.commit()
            
            return JSONResponse({"success": True, "message": "Registration Successful"})
        finally:
            conn.close()
            
    except Exception as e:
        # Handle duplicates if any race condition
        if "UNIQUE" in str(e).upper():
             return JSONResponse({"success": False, "error": "User already exists"})
        return JSONResponse({"success": False, "error": str(e)})

# -------------------------------------------------------------------------
# NEW: OTP Based Forgot Password Flow (Additive)
# -------------------------------------------------------------------------

@app.get("/forgot-password-otp", response_class=HTMLResponse)
async def forgot_password_otp_page(request: Request):
    """Entry point for the new OTP-based forgot password flow."""
    return templates.TemplateResponse("forgot_password_otp.html", {"request": request})

@app.post("/auth/request-forgot-password-otp")
async def request_forgot_password_otp(request: Request):
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        
        if not email:
             return JSONResponse({"success": False, "error": "Email is required"})
             
        conn = get_db_connection()
        try:
            # Check if user exists (Required for password reset)
            existing = conn.execute("SELECT email FROM customers WHERE email = ?", (email,)).fetchone()
            if not existing:
                 # Security: Could return success to avoid enumeration, but Requirement says "Check email EXISTS"
                 return JSONResponse({"success": False, "error": "Email not found."})
        
            # Generate OTP
            otp = utils.generate_otp()
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            expires_at = (datetime.now() + timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")

            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                conn.execute("DELETE FROM otp_codes WHERE email = %s", (email,))
                conn.execute("""
                    INSERT INTO otp_codes (email, otp, created_at, expires_at) 
                    VALUES (%s, %s, %s, %s)
                 """, (email, otp, created_at, expires_at))
            else:
                conn.execute("DELETE FROM otp_codes WHERE email = ?", (email,))
                conn.execute("""
                    INSERT INTO otp_codes (email, otp, created_at, expires_at) 
                    VALUES (?, ?, ?, ?)
                 """, (email, otp, created_at, expires_at))
            
            conn.commit()
            
            # Send Email
            from app.notifications import call_supabase_email_function
            subject = "Password Reset OTP"
            body = f"Your OTP for password reset is {otp}. Valid for 10 minutes."
            
            try:
                call_supabase_email_function(None, email, subject, body)
            except Exception as e:
                print(f"OTP Email failed: {e}")
                
            return JSONResponse({"success": True, "message": "OTP sent successfully"})
            
        finally:
            conn.close()
            
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

@app.post("/auth/verify-forgot-password-otp")
async def verify_forgot_password_otp(request: Request):
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        otp = data.get("otp", "").strip()
        
        if not email or not otp:
            return JSONResponse({"success": False, "error": "Email and OTP required"})

        conn = get_db_connection()
        try:
            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                row = conn.execute("SELECT * FROM otp_codes WHERE email = %s", (email,)).fetchone()
            else:
                row = conn.execute("SELECT * FROM otp_codes WHERE email = ?", (email,)).fetchone()
            
            if not row:
                return JSONResponse({"success": False, "error": "Invalid or Expired OTP"})
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if str(row['otp']) != str(otp):
                 return JSONResponse({"success": False, "error": "Invalid OTP"})
            if str(row['expires_at']) < now:
                return JSONResponse({"success": False, "error": "OTP Expired"})
                
            # SUCCESS: Delete OTP
            if is_pg:
                conn.execute("DELETE FROM otp_codes WHERE email = %s", (email,))
            else:
                conn.execute("DELETE FROM otp_codes WHERE email = ?", (email,))
            conn.commit()
            
            # Set Session Flag
            request.session["password_reset_verified"] = email
            return JSONResponse({"success": True, "message": "OTP Verified"})
            
        finally:
            conn.close()
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

@app.post("/auth/reset-password")
async def reset_password_api(request: Request):
    try:
        data = await request.json()
        email = data.get("email", "").strip().lower()
        new_password = data.get("new_password", "")
        
        if not email or not new_password:
             return JSONResponse({"success": False, "error": "Missing fields"})
             
        # Security Check
        verified_email = request.session.get("password_reset_verified")
        if not verified_email or verified_email != email:
             return JSONResponse({"success": False, "error": "Session expired or unauthorized. Please verify OTP again."})
        
        if len(new_password) < 6:
            return JSONResponse({"success": False, "error": "Password too short"})

        conn = get_db_connection()
        try:
            conn.execute("UPDATE customers SET password = ? WHERE email = ?", (new_password, email))
            conn.commit()
            
            # Clear flag
            request.session.pop("password_reset_verified", None)
            
            return JSONResponse({"success": True, "message": "Password reset successfully"})
        finally:
            conn.close()

    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)})

# -------------------------------------------------------------------------
# NEW: Notification APIs
# -------------------------------------------------------------------------

@app.get("/api/notifications")
async def api_get_notifications(request: Request):
    """Get unread/recent notifications for the current user."""
    user = None
    try:
        user = require_customer(request)
        role = 'customer'
        email = user['email']
    except:
        user = get_current_user(request)
        if user and user.get("role") == "admin":
             role = 'admin'
             # Admin session might not have email, fallback to settings
             email = user.get('email')
             if not email:
                 email = get_setting("admin_email", ADMIN_EMAIL)
        else:
             return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
    
    conn = get_db_connection()
    try:
        # If admin, ensure email is valid (or just use a placeholder if truly missing)
        if role == 'admin' and not email:
             email = "admin@example.com"
             
        notifs = get_notifications(conn, email)
        data = []
        for n in notifs:
            n_dict = dict(n)
            # Serialize datetime objects
            for k, v in n_dict.items():
                if isinstance(v, datetime):
                    n_dict[k] = v.isoformat()
            data.append(n_dict)
            
        return JSONResponse({"status": "success", "notifications": data})
    finally:
        conn.close()

@app.post("/api/notifications/{n_id}/read")
async def api_mark_notification_read(request: Request, n_id: int):
    """Mark a notification as read."""
    user = None
    try:
        user = require_customer(request)
        email = user['email']
    except:
        user = get_current_user(request)
        if user and user.get("role") == "admin":
             email = user['email']
        else:
             return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)
             
    conn = get_db_connection()
    try:
        mark_notification_read(conn, n_id, email)
        conn.commit()
        return JSONResponse({"status": "success"})
    finally:
        conn.close()

@app.post("/api/push/subscribe")
async def api_push_subscribe(request: Request):
    """Save web push subscription."""
    try:
        data = await request.json()
        endpoint = data.get("endpoint")
        keys = data.get("keys", {})
        p256dh = keys.get("p256dh")
        auth = keys.get("auth")
        
        if not endpoint or not p256dh or not auth:
            return JSONResponse({"status": "error", "message": "Invalid subscription data"})

        user = None
        try:
            user = require_customer(request)
            email = user['email']
        except:
            user = get_current_user(request)
            if user and user.get("role") == "admin":
                 email = user['email']
            else:
                 return JSONResponse({"status": "error", "message": "Unauthorized"}, status_code=401)

        conn = get_db_connection()
        try:
            is_pg = os.getenv("DATABASE_URL") is not None
            if is_pg:
                sql = """
                    INSERT INTO push_subscriptions (user_email, endpoint, p256dh, auth)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (endpoint) DO UPDATE SET user_email = EXCLUDED.user_email
                """
                conn.execute(sql, (email, endpoint, p256dh, auth))
            else:
                count = conn.execute("SELECT count(*) FROM push_subscriptions WHERE endpoint = ?", (endpoint,)).fetchone()[0]
                if count > 0:
                    conn.execute("UPDATE push_subscriptions SET user_email = ? WHERE endpoint = ?", (email, endpoint))
                else:
                    conn.execute("INSERT INTO push_subscriptions (user_email, endpoint, p256dh, auth) VALUES (?, ?, ?, ?)", 
                                 (email, endpoint, p256dh, auth))
            
            conn.commit()
            return JSONResponse({"status": "success", "message": "Subscribed"})
        finally:
            conn.close()
    except Exception as e:
        print(f"Push subscribe error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)
